import openpyxl
import os

file_path = "Lead CRM ApplicationData.xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
else:
    try:
        print("Loading workbook (read_only)...")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\nScanning sheet: {sheet_name}")
            ws = wb[sheet_name]
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = list(row)
                break
            print(f"Headers: {headers}")
            
    except Exception as e:
        print(f"Error: {e}")
