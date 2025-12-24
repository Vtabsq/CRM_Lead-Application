from fastapi import FastAPI, HTTPException, Body, BackgroundTasks, UploadFile, File, Query
from fastapi.responses import StreamingResponse, FileResponse, Response
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import range_boundaries
import gspread
from google.oauth2.service_account import Credentials
import os # Trigger Reload Fix
from datetime import datetime
import re
import uuid
import random
import csv
import json
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv
import requests
import httpx
from file_manager import save_upload, process_data_file
import pandas as pd
from dropdown_helpers import (
    get_dropdown_option_sheet,
    initialize_dropdown_sheet,
    get_all_dropdown_options,
    get_dropdown_options_for_field,
    add_dropdown_option,
    delete_dropdown_option,
    sync_dropdown_options_to_schema,
    sync_schema_to_dropdown_sheet
)

# Import invoice routes
try:
    from invoice_routes import router as invoice_router
    INVOICE_MODULE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Invoice module not available: {e}")
    INVOICE_MODULE_AVAILABLE = False

# Import catalog routes
try:
    from catalog_routes import router as catalog_router
    CATALOG_MODULE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Catalog module not available: {e}")
    CATALOG_MODULE_AVAILABLE = False

# Import dashboard routes
try:
    from dashboard_routes import router as dashboard_router
    DASHBOARD_MODULE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Dashboard module not available: {e}")
    DASHBOARD_MODULE_AVAILABLE = False

# Import dropdown routes
try:
    from dropdown_routes import router as dropdown_router
    DROPDOWN_MODULE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Dropdown module not available: {e}")
    DROPDOWN_MODULE_AVAILABLE = False

# Import home care routes
try:
    from homecare_routes import router as homecare_router
    HOMECARE_MODULE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Home Care module not available: {e}")
    HOMECARE_MODULE_AVAILABLE = False

# Import home care scheduler
try:
    from homecare_scheduler import start_billing_scheduler, stop_billing_scheduler
    HOMECARE_SCHEDULER_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Home Care scheduler not available: {e}")
    HOMECARE_SCHEDULER_AVAILABLE = False

# Import patient admission routes
try:
    from patientadmission_routes import router as patientadmission_router
    PATIENTADMISSION_MODULE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Patient Admission module not available: {e}")
    PATIENTADMISSION_MODULE_AVAILABLE = False

# Import patient admission scheduler
try:
    from patientadmission_scheduler import start_billing_scheduler as start_pa_billing_scheduler, stop_billing_scheduler as stop_pa_billing_scheduler
    PATIENTADMISSION_SCHEDULER_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Patient Admission scheduler not available: {e}")
    PATIENTADMISSION_SCHEDULER_AVAILABLE = False


# Load environment variables from .env file
# Trigger reload for schema update
load_dotenv()

# ============== Excel Template is OPTIONAL for cloud deployment ==============
# The app will use cached schema or fallback schema if Excel file is not present.
# This allows deployment without uploading the Excel file as a secret.
# ============================================================================

app = FastAPI()

SETTINGS_FILE = "settings.json"

class Settings(BaseModel):
    single_bed_price: float = 1000.0
    twin_bed_price: float = 2000.0
    icu_bed_price: float = 5000.0
    general_bed_price: float = 0.0
    nurse_fee: float = 0.0
    bed_service_charge: float = 0.0
    consultation_fee: float = 0.0
    registration_fee: float = 0.0
    medication_fee: float = 0.0
    miscellaneous_fee: float = 0.0

def load_settings() -> Settings:
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                data = json.load(f)
            return Settings(**data)
        except Exception:
            return Settings()
    return Settings()

def save_settings(settings: Settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings.dict(), f, indent=2)

# CORS middleware to allow frontend requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for development
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include invoice router
if INVOICE_MODULE_AVAILABLE:
    app.include_router(invoice_router, prefix="/api", tags=["invoices"])
    print("[Invoice Module] Loaded successfully")
else:
    print("[Invoice Module] Not loaded - module unavailable")

# Include catalog router
if CATALOG_MODULE_AVAILABLE:
    app.include_router(catalog_router)
    print("[Catalog Module] Loaded successfully")
else:
    print("[Catalog Module] Not loaded - module unavailable")

# Include dashboard router
if DASHBOARD_MODULE_AVAILABLE:
    app.include_router(dashboard_router)
    print("[Dashboard Module] Loaded successfully")
else:
    print("[Dashboard Module] Not loaded - module unavailable")

# Include dropdown router
if DROPDOWN_MODULE_AVAILABLE:
    app.include_router(dropdown_router)
    print("[Dropdown Module] Loaded successfully")
else:
    print("[Dropdown Module] Not loaded - module unavailable")

# Include home care router
if HOMECARE_MODULE_AVAILABLE:
    app.include_router(homecare_router, prefix="/api", tags=["homecare"])
    print("[Home Care Module] Loaded successfully")
else:
    print("[Home Care Module] Not loaded - module unavailable")

# Include patient admission router
if PATIENTADMISSION_MODULE_AVAILABLE:
    app.include_router(patientadmission_router, prefix="/api", tags=["patientadmission"])
    print("[Patient Admission Module] Loaded successfully")
else:
    print("[Patient Admission Module] Not loaded - module unavailable")


# Configuration  
EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH", "Lead CRM ApplicationData.xlsx")
EXCEL_SHEET_NAME = "Lead CRM"  # Sheet name to use from the Excel file
ENQUIRY_SHEET_NAME = "Enquiry"
ADMISSION_SHEET_NAME = "Patient Admission"
CHARGE_SHEET_NAME = "Charge Summary"
COMPLAINT_SHEET_NAME = "Complaints"
FEEDBACK_SHEET_NAME = "Feedback"
LIST_BOX_SHEET = "List box"  # Sheet containing dropdown options (legacy)
DROPDOWN_OPTION_SHEET = "DropdownOption"  # New centralized dropdown management sheet
GOOGLE_SHEET_NAME = "Sheet1"  # Keep as default variable but irrelevant for data storage now
CREDENTIALS_FILE = os.getenv("CREDENTIALS_FILE", "google_credentials.json")
CSV_FILE_PATH = "CRM Leads - Sheet1.csv"
FIELD_SCHEMA_FILE = "field_schema.json"
ADMISSION_SCHEMA_FILE = "admission_schema.json"
SCHEMA_CACHE_FILE = "schema_cache.json"

NOTIFICATION_SETTINGS_FILE = "notification_settings.json"
DEFAULT_SENDER_EMAIL = os.getenv("DEFAULT_NOTIFICATION_EMAIL", "harishkadhiravan.vtab@gmail.com")
NOTIFICATION_SETTINGS_FILE = "notification_settings.json"
DEFAULT_SENDER_EMAIL = os.getenv("DEFAULT_NOTIFICATION_EMAIL", "harishkadhiravan.vtab@gmail.com")
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")

# Defined by User Request (2025-12-16)
OFFICIAL_COLUMNS = [
    "Date", "Member ID key", "Attender Name", "Patient Name", "Gender", "Age", 
    "Patient Location", "Area", "Email Id", "Mobile Number", "Pain Point", 
    "Enquiry made for", "Service", "Hospital Location", "Source", "Crm Agent Name", 
    "Currently Assigned", "Lead Status", "Active/Inactive", "Reason For Rejection", 
    "Follow_1 Date", "Reminder Date_1", "Follow_2 Date", "Reminder Date_2", 
    "Follow_3 Date", "Reminder Date_3", "Follow_4 Date", "Ed Comments", "Timestamp", 
    "Patient Last Name", "Date of Birth", "Patient Blood", "Patient Marital", 
    "Nationality", "Religion", "Aadhaar", "ID Proof Type", "ID Proof Number", 
    "Door Number", "Street", "City", "District", "State", "Pin Code", 
    "Relational Name", "Relational Relationship", "Relational Mobile", 
    "Relational Mobile Alternative", "Emergency Address", "Patient Current Status", 
    "Patient Sugar level", "Patient Medical History", "Patient Allergy", "Room Type", 
    "Check In Date", "Check Out Date", "Attender Name_1", "Caretaker Name", 
    "Providing Services", "Room Rent", "Room Charges", "Total Amount", "Admission Date", 
    "Total Days Stayed", "Room Charge Total", "Bed Charge Total", "Nurse Payment Total", 
    "Hospital Payment Total", "Doctor Fee", "Service Charge", "Bill Grand Total", 
    "Bill Generated Date", 
    # Duplicate-like keys provided in request, keeping all to be safe or maybe aliases?
    "TotalDaysStayed", "RoomChargeTotal", "BedChargeTotal", "NursePaymentTotal", 
    "HospitalPaymentTotal", "DoctorFee", "ServiceCharge", "BillGrandTotal", "BillGeneratedDate"
]
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
EMAIL_SUBJECT = os.getenv("EMAIL_SUBJECT", "New CRM Lead Submission")
EMAIL_TRANSPORT = os.getenv("EMAIL_TRANSPORT", "smtp").strip().lower()

USE_GMAIL_API = False
if EMAIL_TRANSPORT == "gmail_api":
    try:
        from gmail_api_sender import send_email_via_gmail_api  # type: ignore
        USE_GMAIL_API = True
        print("[Email CONFIG] Gmail API transport enabled")
    except Exception as gmail_import_exc:  # pragma: no cover - import guard only
        print(f"[Email CONFIG] Gmail API not available ({gmail_import_exc}); falling back to SMTP")
        USE_GMAIL_API = False

# Optional: lock to a specific Google Sheet (recommended)
# - Set env var GOOGLE_SHEET_ID, or
# - Put the Sheet ID in a local file named "google_sheet_id.txt" in backend/
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID")
if not GOOGLE_SHEET_ID:
    try:
        with open("google_sheet_id.txt", "r", encoding="utf-8") as f:
            GOOGLE_SHEET_ID = f.read().strip() or None
    except FileNotFoundError:
        GOOGLE_SHEET_ID = None

ENQUIRIES_SHEET_NAME = "Enquiries"
PATIENT_ADMISSION_SHEET_NAME = "patient admission"

# --- NEW: Secondary Google Sheet for Patient Admission ---
PATIENT_ADMISSION_CREDENTIALS_FILE = "CRM-admission.json"

# Logic for Patient Admission Sheet ID (Env or File)
PATIENT_ADMISSION_SHEET_ID = os.getenv("PATIENT_ADMISSION_SHEET_ID")
if not PATIENT_ADMISSION_SHEET_ID:
    try:
        with open("patient_admission_sheet_id.txt", "r", encoding="utf-8") as f:
            PATIENT_ADMISSION_SHEET_ID = f.read().strip() or None
    except FileNotFoundError:
        PATIENT_ADMISSION_SHEET_ID = None
# -----------------------------------------------------


def normalize_field_name(name: str) -> str:
    """
    Normalize field name: lowercase, remove spaces, hyphens, underscores.
    Example: "First Name" -> "firstname"
    """
    if not name:
        return ""
    # Remove strict alphanumeric characters except common safe ones if needed, 
    # but requirement says: lowercase, remove spaces, hyphens, underscores.
    return str(name).strip().lower().replace(" ", "").replace("_", "").replace("-", "")



# Fields to exclude from the dynamic form (case-insensitive)
EXCLUDED_FIELD_NAMES = {
    "reason for rejection",
    "ed comments",
    "closed date",
}

# Keywords that, when present in a header, should exclude the field
EXCLUDED_KEYWORDS = [
    # Remove follow-up/reminder date columns like Follow1 Date, Reminder Date_1, etc.
    "follow",    # matches Follow1 Date, Follow_2 Date, Follow_3 Date, etc.
    "reminder",  # matches Reminder Date_1, Reminder Date_2, etc.
]

# Keywords for date fields that should remain blank unless explicitly provided
REMINDER_DATE_KEYWORD = "reminder"
FOLLOW_DATE_KEYWORD = "follow"

FOLLOW_DATE_KEYWORD = "follow"

# Cache for schemas: "enquiry" -> list, "admission" -> list
fields_cache: Dict[str, List[Dict[str, Any]]] = {
    "enquiry": [],
    "admission": []
}
notification_settings_cache: Optional[Dict[str, Any]] = None


def load_notification_settings() -> Dict[str, Any]:
    """Return cached notification settings from disk with defaults."""
    global notification_settings_cache
    if notification_settings_cache is not None:
        return notification_settings_cache

    settings = {
        "sender_email": DEFAULT_SENDER_EMAIL,
        "cc_emails": [],
    }

    if os.path.exists(NOTIFICATION_SETTINGS_FILE):
        try:
            with open(NOTIFICATION_SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                sender = str(data.get("sender_email", "")).strip()
                if sender:
                    settings["sender_email"] = sender
                raw_cc = data.get("cc_emails")
                if isinstance(raw_cc, list):
                    cleaned_cc = [str(item).strip() for item in raw_cc if str(item).strip()]
                    if cleaned_cc:
                        settings["cc_emails"] = cleaned_cc
        except Exception as exc:
            print(f"Warning: could not read {NOTIFICATION_SETTINGS_FILE}: {exc}")

    notification_settings_cache = settings
    return settings


def save_notification_settings(settings: Dict[str, Any]) -> None:
    """Persist notification settings and refresh cache."""
    global notification_settings_cache
    sender = str(settings.get("sender_email", DEFAULT_SENDER_EMAIL)).strip() or DEFAULT_SENDER_EMAIL
    raw_cc = settings.get("cc_emails") or []
    cleaned_cc = [str(item).strip() for item in raw_cc if str(item).strip()]

    notification_settings_cache = {
        "sender_email": sender,
        "cc_emails": cleaned_cc,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    try:
        with open(NOTIFICATION_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(notification_settings_cache, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"Warning: could not write {NOTIFICATION_SETTINGS_FILE}: {exc}")


class FormData(BaseModel):
    data: Dict[str, Any]


class FieldItem(BaseModel):
    name: str
    label: Optional[str] = None
    data_type: str
    options: Optional[List[str]] = None


class UpdateFieldsPayload(BaseModel):
    fields: List[FieldItem]


class NotificationSettings(BaseModel):
    sender_email: str
    cc_emails: Optional[List[str]] = None



class LoginRequest(BaseModel):
    User_name: str
    Password: str


class FilterCriteria(BaseModel):
    year: Optional[str] = None
    month: Optional[str] = None
    specificDate: Optional[str] = None
    startDate: Optional[str] = None
    endDate: Optional[str] = None

class DeletePreviewPayload(BaseModel):
    filters: FilterCriteria
    date_column: str
    preview_columns: Optional[List[str]] = None

class DeleteConfirmPayload(BaseModel):
    filters: FilterCriteria
    date_column: str



def ensure_notification_defaults() -> Dict[str, Any]:
    settings = load_notification_settings()
    if not settings.get("sender_email"):
        settings["sender_email"] = DEFAULT_SENDER_EMAIL
    if "cc_emails" not in settings or not isinstance(settings.get("cc_emails"), list):
        settings["cc_emails"] = []
    return settings


def build_email_message(sender: str, recipient: str, payload: Dict[str, Any], subject_override: Optional[str] = None) -> EmailMessage:
    message = EmailMessage()
    message["From"] = f"CRM Lead Form <{sender}>"
    message["To"] = recipient
    message["Reply-To"] = sender
    message["Return-Path"] = sender
    message["X-Mailer"] = "CRM Lead Application"
    message["X-Priority"] = "3"

    # CC
    settings = ensure_notification_defaults()
    cc_list = settings.get("cc_emails", [])
    if cc_list:
        message["Cc"] = ", ".join(cc_list)

    # Resolve common aliases from payload (case-insensitive)
    def get_value(aliases: List[str]) -> Optional[str]:
        if not payload:
            return None
        # Direct first
        for a in aliases:
            if a in payload:
                val = payload.get(a)
                if val is not None and str(val).strip():
                    return str(val).strip()
        # Case-insensitive
        lower_map = {str(k).strip().lower(): v for k, v in payload.items()}
        for a in aliases:
            v = lower_map.get(a.lower())
            if v is not None and str(v).strip():
                return str(v).strip()
        return None

    member_id = get_value([
        "Memberidkey", "Member ID", "member id", "Member_ID", "Memberid", "MID"
    ])
    attender_name = get_value(["Attender Name", "attender name", "Attender", "attender"]) or "Customer"

    # Subject with Member ID if available (or override)
    subject = subject_override or EMAIL_SUBJECT
    if member_id:
        if subject_override:
            subject = f"{subject_override}: {member_id}"
        else:
            subject = f"{EMAIL_SUBJECT} - {member_id}"
    message["Subject"] = subject

    # Body template
    lines: List[str] = []
    lines.append(f"Dear {attender_name}, your form has been submitted")
    # If this is a status-change notification, add explicit status line
    if (subject_override or "").strip().upper().startswith("LEAD STATUS CHANGED"):
        lead_status = get_value(["Lead Status", "lead status"]) or ""
        if lead_status:
            lines.append("")
            lines.append(f"Lead status changed to : {lead_status}")
    lines.append("")
    lines.append("confirm your details below:")
    for key, value in payload.items():
        lines.append(f"- {key}: {value}")
    message.set_content("\n".join(lines))
    return message


def send_notification_email(recipient: str, payload: Dict[str, Any], subject_override: Optional[str] = None) -> None:
    print(f"[Email DEBUG] Function called with recipient={recipient}")
    if not recipient:
        print("[Email DEBUG] No recipient, returning early")
        return

    settings = ensure_notification_defaults()
    sender = settings.get("sender_email", DEFAULT_SENDER_EMAIL)
    cc_list: List[str] = settings.get("cc_emails", [])
    print(f"[Email DEBUG] Sender={sender}, CC={cc_list}, transport={'GMAIL_API' if USE_GMAIL_API else 'SMTP'}")

    if not USE_GMAIL_API and (not SMTP_USERNAME or not SMTP_PASSWORD):
        print("Warning: SMTP credentials not configured; skipping email notification")
        return

    try:
        masked_pw_len = len(SMTP_PASSWORD) if SMTP_PASSWORD else 0
        print(f"[Email] Preparing to send | sender={sender} | recipient={recipient} | user={SMTP_USERNAME} | pw_len={masked_pw_len} | host={SMTP_HOST}:{SMTP_PORT}")
    except Exception:
        pass

    # Adjust sender alignment:
    # - For Gmail, align header/envelope From to SMTP username for deliverability
    # - For Brevo (smtp-relay.brevo.com / *@smtp-brevo.com), keep the configured sender
    host_l = str(SMTP_HOST or "").lower()
    user_l = str(SMTP_USERNAME or "").lower()
    is_brevo = ("brevo" in host_l) or user_l.endswith("@smtp-brevo.com")
    if not USE_GMAIL_API and not is_brevo:
        effective_sender = SMTP_USERNAME if SMTP_USERNAME else sender
        if effective_sender != sender:
            print(f"[Email DEBUG] Aligning header From with SMTP user: {sender} -> {effective_sender}")
            sender = effective_sender

    message = build_email_message(sender, recipient, payload, subject_override)

    # Optional simple mode to suppress CCs for stricter deliverability
    simple_mode = str(os.getenv("EMAIL_SIMPLE_MODE", "0")).strip() == "1"
    if simple_mode and cc_list:
        print("[Email DEBUG] EMAIL_SIMPLE_MODE=1 -> suppressing CCs for this send")
        cc_list = []

    # Ensure message headers match final CC list
    if cc_list:
        if "Cc" in message:
            message.replace_header("Cc", ", ".join(cc_list))
        else:
            message["Cc"] = ", ".join(cc_list)
    elif "Cc" in message:
        del message["Cc"]

    # Auto-BCC logic (skip Brevo technical usernames)
    bcc_list: List[str] = []
    if (USE_GMAIL_API or not is_brevo) and SMTP_USERNAME and SMTP_USERNAME not in ([recipient] + cc_list):
        bcc_list = [SMTP_USERNAME]
        print(f"[Email DEBUG] Auto-BCC enabled -> {bcc_list}")

    if USE_GMAIL_API:
        try:
            body_text = message.get_content()
        except Exception:
            body_text = str(message)
        try:
            print("[Email DEBUG] Sending via Gmail API...")
            send_email_via_gmail_api(
                sender_email=sender,
                recipient_email=recipient,
                subject=message["Subject"],
                body_text=body_text,
                cc_emails=cc_list or None,
                bcc_emails=bcc_list or None,
            )
            print(f"[Email SUCCESS] Gmail API accepted message for {recipient}")
        except Exception as exc:
            print(f"[Email ERROR] Gmail API send failed: {type(exc).__name__}: {exc}")
            import traceback
            traceback.print_exc()
        return

    try:
        print(f"[Email DEBUG] Connecting to {SMTP_HOST}:{SMTP_PORT}...")
        # Try SSL if port 465, otherwise TLS
        if SMTP_PORT == 465:
            print(f"[Email DEBUG] Using SMTP_SSL on port 465...")
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
        
        try:
            server.ehlo()
            if SMTP_PORT != 465:
                print(f"[Email DEBUG] Starting TLS...")
                server.starttls()
                server.ehlo()
            print(f"[Email DEBUG] Logging in as {SMTP_USERNAME}...")
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            print(f"[Email DEBUG] Login successful!")
            print(f"[Email DEBUG] Preparing to send...")
            print(f"[Email DEBUG] From: {sender}")
            print(f"[Email DEBUG] To: {recipient}")
            print(f"[Email DEBUG] CC: {cc_list}")
            print(f"[Email DEBUG] Subject: {message['Subject']}")
            all_recipients = [recipient] + cc_list + bcc_list
            # de-duplicate recipients while preserving order
            seen = set()
            deduped: List[str] = []
            for addr in all_recipients:
                if addr and addr not in seen:
                    deduped.append(addr)
                    seen.add(addr)
            all_recipients = deduped
            # Use sendmail instead of send_message for more control
            msg_string = message.as_string()
            refused = server.sendmail(sender, all_recipients, msg_string)
            if refused:
                print(f"[Email WARNING] Some recipients were refused: {refused}")
            else:
                print(f"[Email DEBUG] All recipients accepted by server")
            print(f"[Email SUCCESS] Email sent to {recipient} (cc: {', '.join(cc_list) if cc_list else 'none'})")
        finally:
            server.quit()
    except Exception as exc:
        print(f"[Email ERROR] Failed to send: {type(exc).__name__}: {exc}")
        import traceback
        traceback.print_exc()


def send_follow_email(recipient: str, member_entry: Dict[str, Any], query_trigger: str) -> None:
    """Send a follow-up reminder email for the specified member."""
    if not recipient:
        print("[Follow Email] No recipient email provided; skipping")
        return

    payload = dict(member_entry.get("fields", {}))
    payload.setdefault("Member ID", member_entry.get("id", ""))
    payload.setdefault("Member Name", member_entry.get("name", "Unknown"))
    payload["AI Follow-up Trigger"] = query_trigger

    subject = f"FOLLOW-UP REMINDER - {member_entry.get('id', 'Member')}"
    try:
        send_notification_email(recipient, payload, subject_override=subject)
        print(f"[Follow Email] Successfully queued follow-up email to {recipient}")
    except Exception as exc:
        print(f"[Follow Email] Failed to send email to {recipient}: {exc}")



def extract_recipient_email(payload: Dict[str, Any]) -> Optional[str]:
    if not payload:
        return None

    for key, value in payload.items():
        if value is None:
            continue
        if "email" in str(key).lower():
            candidate = str(value).strip()
            if candidate:
                return candidate

    lower_map: Dict[str, Any] = {}
    for key, value in payload.items():
        if value is None:
            continue
        lower_map[str(key).strip().lower()] = value

    for field in fields_cache.get("enquiry", []):
        name = field.get("name")
        if not name:
            continue
        normalized = str(name).strip().lower()
        if "email" not in normalized:
            continue
        candidate = lower_map.get(normalized)
        if candidate:
            text = str(candidate).strip()
            if text:
                return text
    return None


# --- CANONICAL MAPPING FOR SHEET1 ---
CANONICAL_MAP = {
    # Names
    "attender name": "Attender Name",
    "attender_name": "Attender Name",
    "attendername": "Attender Name",
    "Attender Name": "Attender Name",
    
    "patient name": "Patient Name",
    "patient_name": "Patient Name",
    "patientname": "Patient Name",
    "first name": "Patient Name",
    "firstname": "Patient Name",
    "Patient Name": "Patient Name",
    "full name": "Patient Name",
    "name": "Patient Name",
    "fullname": "Patient Name",
    
    # IDs
    "memberidkey": "Member ID Key",
    "member_id_key": "Member ID Key",
    "member id key": "Member ID Key",
    "member id": "Member ID Key",
    "memberid": "Member ID Key",
    "id": "Member ID Key",
    "Member ID Key": "Member ID Key",
    
    # Dates
    "date": "Date",
    "admissiondate": "Date", 
    "admission date": "Date",
    
    # Contact
    "phone": "Phone Number",
    "phonenumber": "Phone Number", 
    "mobile": "Phone Number",
    "mobilenumber": "Phone Number",
    "email": "Email ID",
    "emailid": "Email ID",
    "email id": "Email ID",
    "mail": "Email ID",
    
    # Address
    "address": "Address",
    "location": "Hospital Location", 
    "hospital location": "Hospital Location",

    # Medical
    "pain point": "Pain Point",
    "painpoint": "Pain Point",
    "pain_point": "Pain Point",
    "diagnosis": "Pain Point",
}

def normalize_for_sheet1(payload: dict):
    """
    Normalize payload keys to Canonical Sheet1 Headers.
    Prevents duplicate columns by converting variations to the single source of truth.
    """
    normalized = {}

    for key, value in payload.items():
        k_clean = str(key).strip().lower()
        # Remove underscores effectively for lookup?
        k_clean_us = k_clean.replace("_", "") # attender_name -> attendername
        
        # Try lookup
        canonical = CANONICAL_MAP.get(k_clean) or CANONICAL_MAP.get(k_clean_us) or CANONICAL_MAP.get(key)
        
        if canonical:
            normalized[canonical] = value
        else:
            # Keep original if no canonical map found
            normalized[key.strip()] = value

    return normalized

def normalize_payload(data: dict):
    """
    Remove suffixes like _2, _3 from keys and keep first non-empty value.
    Ensures data matches primary column names.
    """
    normalized = {}

    for key, value in data.items():
        # strict suffix check (digits or 'copy')
        # to avoid accidental splits like "Member_ID"
        base_key = key
        
        # Regex to remove _\d+$ or _copy$ (case insensitive)
        import re
        base_key_clean = re.sub(r'_(\d+|copy)$', '', base_key, flags=re.IGNORECASE).strip()
        
        if not base_key_clean:
             base_key_clean = base_key # fallback

        # Logic: "Keep first non-empty value only"
        if base_key_clean not in normalized or not normalized[base_key_clean]:
             normalized[base_key_clean] = value
             
    return normalized


def get_value(row, *keys):
    """
    Read value safely from row dealing with potential legacy variants/duplicates.
    """
    for k in keys:
        val = row.get(k)
        if val is not None and str(val).strip():
            return val
    return ""



def read_excel_headers() -> List[str]:
    """Read header row from the Excel file. Returns empty list if file not found."""
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"[read_excel_headers] Excel file not found: {EXCEL_FILE_PATH} - returning empty list")
        return []
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
        sheet = workbook[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME else workbook.active
        headers: List[str] = []
        for cell in sheet[1]:
            if not cell.value:
                continue
            header = str(cell.value).strip()
            header_lower = header.lower()
            if header_lower in EXCLUDED_FIELD_NAMES:
                continue
            if any(keyword in header_lower for keyword in EXCLUDED_KEYWORDS):
                continue
            headers.append(header)
        workbook.close()
        return headers
    except Exception as e:
        print(f"[read_excel_headers] Error reading Excel: {e}")
        return []


def read_csv_headers(csv_path: str) -> List[str]:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if any(str(x).strip() for x in row):
                return [h.strip() for h in row if str(h).strip()]
    return []


def read_excel_dropdown_options(workbook: openpyxl.Workbook) -> Dict[str, List[str]]:
    """Read dropdown options from the List box sheet"""
    dropdowns: Dict[str, List[str]] = {}
    
    try:
        if LIST_BOX_SHEET not in workbook.sheetnames:
            return dropdowns
        
        listbox_ws = workbook[LIST_BOX_SHEET]
        
        # Get headers from List box sheet (first row)
        headers = []
        for cell in listbox_ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append(None)
        
        # Read values for each column (exclude Date column as it's not a dropdown)
        for col_idx, header in enumerate(headers, start=1):
            if not header or header.lower() == 'date':
                continue
            
            values = set()
            # Read up to 1000 rows to get all unique values
            for row_idx in range(2, min(listbox_ws.max_row + 1, 1000)):
                val = listbox_ws.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip():
                    clean_val = str(val).strip()
                    # Skip datetime objects from Date column
                    if 'datetime' not in str(type(val)):
                        values.add(clean_val)
            
            if values:
                dropdowns[header] = sorted(list(values))
    
    except Exception as e:
        print(f"Warning: Could not read List box sheet: {e}")
    
    return dropdowns


def infer_field_type(field_name: str) -> str:
    """Infer input type based on field name"""
    field_lower = field_name.lower()
    
    if any(keyword in field_lower for keyword in ['date', 'dob', 'birth']):
        return 'date'
    elif any(keyword in field_lower for keyword in ['email', 'mail']):
        return 'email'
    elif any(keyword in field_lower for keyword in ['phone', 'mobile', 'contact']):
        return 'tel'
    elif any(keyword in field_lower for keyword in ['age', 'number', 'count', 'quantity']):
        return 'number'
    elif any(keyword in field_lower for keyword in ['description', 'comment', 'notes', 'address']):
        return 'textarea'
    else:
        return 'text'


def canonicalize_schema(headers: List[str]) -> List[Dict[str, Any]]:
    schema: List[Dict[str, Any]] = []
    for h in headers:
        dtype = infer_field_type(h)
        mapped_type = {
            'tel': 'phone',
            'textarea': 'text',
        }.get(dtype, dtype)
        schema.append({
            "name": h,
            "label": h,
            "data_type": mapped_type if mapped_type in {"text","number","date","email","phone","boolean","dropdown"} else "text",
            "options": []
        })
    return schema


def load_field_schema_from_disk(schema_type: str = "enquiry") -> Optional[List[Dict[str, Any]]]:
    filename = ADMISSION_SCHEMA_FILE if schema_type == "admission" else FIELD_SCHEMA_FILE
    # Fallback to cache file only for enquiry (legacy compatibility)
    paths = [filename]
    if schema_type == "enquiry":
        paths.insert(0, SCHEMA_CACHE_FILE)
        
    for path in paths:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        return data
            except Exception:
                continue
    return None


def save_field_schema_to_disk(schema: List[Dict[str, Any]], schema_type: str = "enquiry"):
    filename = ADMISSION_SCHEMA_FILE if schema_type == "admission" else FIELD_SCHEMA_FILE
    paths = [filename]
    if schema_type == "enquiry":
        paths.append(SCHEMA_CACHE_FILE)
        
    for path in paths:
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(schema, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Warning: could not write {path}: {e}")


def load_schema(schema_type: str = "enquiry") -> List[Dict[str, Any]]:
    """Load schema: local JSON -> CSV -> Excel."""
    global fields_cache
    
    # 1. Try Disk
    disk_schema = load_field_schema_from_disk(schema_type)
    if disk_schema:
        fields_cache[schema_type] = disk_schema
        return disk_schema
        
    # 2. Derive from Files (Only for Enquiry currently)
    if schema_type == "enquiry":
        if os.path.exists(CSV_FILE_PATH):
            headers = read_csv_headers(CSV_FILE_PATH)
            schema = canonicalize_schema(headers)
            fields_cache["enquiry"] = schema
            return schema
        headers = read_excel_headers()
        schema = canonicalize_schema(headers)
        fields_cache["enquiry"] = schema
        return schema
        
    return []


def build_schema_from_excel() -> List[Dict[str, Any]]:
    """Build schema from Excel file. Returns empty list if file not found."""
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"[build_schema_from_excel] Excel file not found: {EXCEL_FILE_PATH} - returning empty schema")
        return []
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True, data_only=True)
        ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME else wb.active
        headers: List[str] = []
        for cell in ws[1]:
            if cell.value and str(cell.value).strip():
                headers.append(str(cell.value).strip())
        dropdown_map = read_excel_dropdown_options(wb)
        # Build a case-insensitive view of dropdown_map
        dropdown_map_ci: Dict[str, List[str]] = {str(k).strip().lower(): v for k, v in dropdown_map.items()}
        schema: List[Dict[str, Any]] = []
        for h in headers:
            h_lower = h.strip().lower()
            dtype = 'dropdown' if h_lower in dropdown_map_ci else infer_field_type(h)
            mapped_type = {'tel': 'phone', 'textarea': 'text'}.get(dtype, dtype)
            schema.append({
                "name": h,
                "label": h,
                "data_type": mapped_type if mapped_type in {"text","number","date","email","phone","boolean","dropdown"} else "text",
                "options": dropdown_map_ci.get(h_lower, [])
            })
        wb.close()
        return schema
    except Exception as e:
        print(f"[build_schema_from_excel] Error reading Excel: {e}")
        return []


def upsert_to_sheet(sheet_name: str, data: Dict[str, Any], schema_type: str = "enquiry", strict_mode: bool = False) -> Dict[str, Any]:
    """
    Upsert data to Google Sheet.
    strict_mode=True: READ ONLY headers, RAW input, NO schema updates. (For CRM Lead Sheet1)
    strict_mode=False: DYNAMIC headers, USER_ENTERED input, Schema updates. (For Enquiries/others)
    """
    
    # --- FIX: NORMALIZE PAYLOAD FOR SHEET1/STRICT MODE ---
    if strict_mode or sheet_name in ["Sheet1", "CRM_Leads", GOOGLE_SHEET_NAME]:
        # We assume data is incoming raw.
        pass # We will handle canonicalization internally below

    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError("Google credentials file not found")

    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)

    # Open Sheet
    spreadsheet = None
    if GOOGLE_SHEET_ID:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
    else:
        try:
            spreadsheet = client.open(GOOGLE_SHEET_NAME)
        except gspread.SpreadsheetNotFound:
            spreadsheet = client.create(GOOGLE_SHEET_NAME)

    # Try case-insensitive math first
    sheet = None
    try:
        sheet = spreadsheet.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        all_sheets = spreadsheet.worksheets()
        for ws in all_sheets:
            if ws.title.lower() == sheet_name.lower():
                sheet = ws
                break
        if not sheet:
            sheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=20)

    # Get All Data
    all_values = sheet.get_all_values()
    
    if not all_values:
        headers = []
        rows = []
    else:
        headers = all_values[0]
        rows = all_values[1:]

    # --- PREPARE CANONICAL DATA LOOKUP ---
    # Map get_canonical_key(k) -> v
    # This ensures "Member ID", "memberid", "member_id" in payload all map to "memberidkey" bucket
    canonical_data_map = {}
    for k, v in data.items():
        c_k = get_canonical_key(k)
        if c_k:
            canonical_data_map[c_k] = v

    # Also keep original normalized map for fallback
    data_norm_simple = {normalize_field_name(k): v for k, v in data.items()}

    # Helper to find data value for a Sheet Header
    def get_data_value(header):
        # 1. Canonical Match (Robust)
        c_h = get_canonical_key(header)
        if c_h in canonical_data_map:
             return canonical_data_map[c_h]
        
        # 2. Exact/Simple Normalized Match (Fallback)
        if header in data:
            return data[header]
        h_norm_simple = normalize_field_name(header)
        if h_norm_simple in data_norm_simple:
            return data_norm_simple[h_norm_simple]
            
        return None

    # Find Member ID Column Index in Sheet
    # matches canonical keys for Member ID
    member_id_col_idx = -1
    canonical_member_id_key = "memberidkey" 

    for idx, h in enumerate(headers):
        if get_canonical_key(h) == canonical_member_id_key:
            member_id_col_idx = idx
            break
            
    # Find Member ID Value in Data
    member_id_val = canonical_data_map.get(canonical_member_id_key)
    if member_id_val:
        member_id_val = str(member_id_val).strip()
            
    # Find Existing Row
    row_index_to_update = -1
    existing_row_data = []
    
    if member_id_val and member_id_col_idx != -1:
        for idx, r in enumerate(rows):
            if len(r) > member_id_col_idx:
                if str(r[member_id_col_idx]).strip() == member_id_val:
                    row_index_to_update = idx + 2 # +2 for 1-based index
                    existing_row_data = r
                    if len(existing_row_data) < len(headers):
                        existing_row_data += [""] * (len(headers) - len(existing_row_data))
                    break

    # Prepare Final Row
    final_row = []
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if row_index_to_update != -1:
        # --- UPDATE MODE ---
        final_row = list(existing_row_data)
        for idx, h in enumerate(headers):
            new_val = get_data_value(h)
            if new_val is not None:
                final_row[idx] = str(new_val)
                
        action = "updated"
        range_to_write = f'A{row_index_to_update}'
        val_opt = 'RAW' if strict_mode else 'USER_ENTERED'
        sheet.update(range_name=range_to_write, values=[final_row], value_input_option=val_opt)
        
    else:
        # --- APPEND MODE ---
        # If headers are empty?
        if not headers and not strict_mode:
             # Initialize headers from keys
             headers = [k for k in data.keys()]
             sheet.update(range_name='1:1', values=[headers])

        final_row = [""] * len(headers)
        for idx, h in enumerate(headers):
             h_canon = get_canonical_key(h)
             if h_canon == 'timestamp':
                 final_row[idx] = current_time
             else:
                 val = get_data_value(h)
                 if val is not None:
                     final_row[idx] = str(val)
                     
        action = "appended"
        val_opt = 'RAW' if strict_mode else 'USER_ENTERED'
        sheet.append_row(final_row, value_input_option=val_opt)

    return {
        "status": "success",
        "action": action,
        "sheet_url": spreadsheet.url
    }


# --- UPDATED: Support Two Separate Google Sheets ---

def get_patient_admission_sheet_client():
    """
    Helper to get authenticated client for Patient Admission sheet.
    UPDATED: Supports BOTH architectures:
    - If CRM-admission.json exists: Use separate sheet (PATIENT_ADMISSION_SHEET_ID)
    - If not: Fall back to same sheet with different worksheet
    """
    # Try to use separate credentials first
    credentials_file = CREDENTIALS_FILE
    if os.path.exists(PATIENT_ADMISSION_CREDENTIALS_FILE):
        credentials_file = PATIENT_ADMISSION_CREDENTIALS_FILE
        print(f"[Admission] Using separate credentials: {PATIENT_ADMISSION_CREDENTIALS_FILE}")
    else:
        print(f"[Admission] Using shared credentials: {CREDENTIALS_FILE}")
    
    if not os.path.exists(credentials_file):
        raise FileNotFoundError(f"Credentials file not found: {credentials_file}")
    
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(credentials_file, scopes=scope)
    client = gspread.authorize(creds)
    
    # Determine which sheet to open
    spreadsheet = None
    
    # If using separate credentials, try to open separate sheet
    if credentials_file == PATIENT_ADMISSION_CREDENTIALS_FILE and PATIENT_ADMISSION_SHEET_ID:
        print(f"[Admission] Opening separate sheet: {PATIENT_ADMISSION_SHEET_ID}")
        spreadsheet = client.open_by_key(PATIENT_ADMISSION_SHEET_ID)
    elif PATIENT_ADMISSION_SHEET_ID:
        # Separate sheet ID provided but using shared credentials
        print(f"[Admission] Opening separate sheet with shared credentials: {PATIENT_ADMISSION_SHEET_ID}")
        spreadsheet = client.open_by_key(PATIENT_ADMISSION_SHEET_ID)
    elif GOOGLE_SHEET_ID:
        # Fall back to same sheet as CRM Lead
        print(f"[Admission] Using same sheet as CRM Lead: {GOOGLE_SHEET_ID}")
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
    else:
        # Last resort: open by name
        try:
            spreadsheet = client.open(GOOGLE_SHEET_NAME)
        except gspread.SpreadsheetNotFound:
            spreadsheet = client.create(GOOGLE_SHEET_NAME)
            
    return client, spreadsheet

# --- CANONICAL MAPPING FOR ADMISSION SHEET ---
ADMISSION_CANONICAL_MAP = {
    # Member ID variations
    "memberid": "memberidkey",
    "memberidkey": "memberidkey", 
    "memberkey": "memberidkey",
    "id": "memberidkey",
    "memberidkry": "memberidkey", # User specified
    "memberit": "memberidkey",   # User specified "MEMBER IT"
    "memberitkey": "memberidkey",
    "member_id": "memberidkey",
    "member_key": "memberidkey",
    "mid": "memberidkey",

    # Names
    "attendername": "attendername",
    "attender_name": "attendername",
    "patientname": "patientname",
    "patient_name": "patientname",
    "name": "patientname",
    "fullname": "patientname",

    # Contact
    "mobile": "mobilenumber",
    "phone": "mobilenumber",
    "phonenumber": "mobilenumber",
    "mobilenumber": "mobilenumber",
    "relationalmobile": "relationalmobile",
    "relational_mobile": "relationalmobile",

    # Location
    "hospital_location": "hospitallocation",
    "hospitallocation": "hospitallocation",
    "location": "hospitallocation",
    
    # Care Center
    "care_center": "carecenter",
    "carecenter": "carecenter",
    "center": "carecenter",
    
    # Dates
    "date": "date",
    "admissiondate": "date",
}

# Display names for Google Sheets columns (canonical_key -> Display Name)
ADMISSION_DISPLAY_NAMES = {
    "carecenter": "Care Center",
    "hospitallocation": "Hospital Location",
    "memberidkey": "Member ID Key",
    "attendername": "Attender Name",
    "patientname": "Patient Name",
    "mobilenumber": "Mobile Number",
    "relationalmobile": "Relational Mobile",
}


def get_canonical_key(key: str) -> str:
    """
    Normalize key to a canonical representation.
    1. Lowercase, strip, remove special chars ( - _).
    2. Check ADMISSION_CANONICAL_MAP.
    3. Return canonical or normalized string.
    """
    if not key:
        return ""
    # Normalize: lowercase, remove spaces, hyphens, underscores
    norm = str(key).strip().lower().replace(" ", "").replace("-", "").replace("_", "")
    
    # Check map
    return ADMISSION_CANONICAL_MAP.get(norm, norm)

def save_patient_admission_to_sheet(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Save mapped data to the Patient Admission sheet/worksheet.
    UPDATED: Handles both architectures:
    - Separate sheet: Saves to "Sheet1" in the admission sheet
    - Same sheet: Saves to "Patient Admission" worksheet
    Logic:
    1. Normalize Input Data Keys -> Canonical Keys.
    2. Read Sheet Headers and map to Canonical Keys.
    3. For each Sheet Header Column, find value from Canonical Data.
    4. Fill duplicate/variant columns with the same data (Merge behavior).
    """
    print(f"[Patient Admission] Incoming Data: {data}")

    # 1. Normalize Input Data
    # Map Canonical Key -> Value
    canonical_data = {}
    for k, v in data.items():
        c_key = get_canonical_key(k)
        # Verify valid value (non-empty preferred if duplicates exist in strict payload, but here last write wins is fine)
        if c_key:
             canonical_data[c_key] = v
    
    print(f"[Patient Admission] Normalized Data: {canonical_data.keys()}")

    client, spreadsheet = get_patient_admission_sheet_client()
    
    # Determine worksheet name based on whether we're using separate sheet or same sheet
    # If PATIENT_ADMISSION_SHEET_ID is set, we're using a separate sheet -> use "Sheet1"
    # Otherwise, we're using same sheet -> use "Patient Admission" worksheet
    if PATIENT_ADMISSION_SHEET_ID:
        sheet_name = "Sheet1"  # Separate sheet uses Sheet1
        print(f"[Patient Admission] Using separate sheet, worksheet: {sheet_name}")
    else:
        sheet_name = "Patient Admission"  # Same sheet uses dedicated worksheet
        print(f"[Patient Admission] Using same sheet, worksheet: {sheet_name}")
    
    sheet = None
    try:
        sheet = spreadsheet.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        sheet = spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=20)

    # 2. Get Sheet Headers
    all_values = sheet.get_all_values()
    headers = all_values[0] if all_values else []
    
    # Filter empty headers? GSpread might return empty strings for trailing cols.
    # We keep them to maintain index alignment, but usually we just append based on count.
    # Actually, simplistic append expects clean headers.
    headers = [str(h).strip() for h in headers if h] 

    print(f"[Patient Admission] Current Sheet Headers: {headers}")

    # --- DYNAMIC HEADER SYNC START ---
    # Check if we have data for a canonical key that has NO column in the sheet.
    # We need to map existing headers to canonical keys to see what's covered.
    covered_canonical_keys = set()
    for h in headers:
        covered_canonical_keys.add(get_canonical_key(h))
    
    new_headers = []
    # Find data keys not covered
    for c_key in canonical_data.keys():
        # strict check: don't add internal keys
        if len(c_key) > 50 or c_key.startswith('_'):
             continue
             
        if c_key not in covered_canonical_keys:
            # We need to add a column. 
            # Priority: 1) ADMISSION_DISPLAY_NAMES, 2) Original key from payload, 3) Canonical key
            original_display = c_key
            
            # First, check if we have a predefined display name
            if c_key in ADMISSION_DISPLAY_NAMES:
                original_display = ADMISSION_DISPLAY_NAMES[c_key]
            else:
                # Otherwise, use the first original key that mapped to this canonical key
                for k in data.keys():
                    if get_canonical_key(k) == c_key:
                        original_display = k
                        break
            
            new_headers.append(original_display)
            covered_canonical_keys.add(c_key) # mark processed

    if new_headers:
        print(f"[Patient Admission] Found new dynamic fields: {new_headers}. Updating Sheet Headers...")
        headers.extend(new_headers)
        try:
            sheet.update(range_name='1:1', values=[headers])
            print("[Patient Admission] Sheet Headers Updated.")
        except Exception as header_err:
             print(f"[Patient Admission] FAILED to update headers: {header_err}")
    # --- DYNAMIC HEADER SYNC END ---

    if not headers:
         raise Exception("Patient Admission Sheet has invalid or empty header columns.")

    # 3. Construct Row
    row = []
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for h in headers:
        c_key = get_canonical_key(h)
        val = ""
        
        # Auto-fill Timestamp
        if c_key == 'timestamp' and 'timestamp' not in canonical_data:
            val = current_time
        elif c_key in canonical_data:
            val = canonical_data[c_key]
        
        row.append(str(val) if val is not None else "")

    # 4. Append
    sheet.append_row(row, value_input_option="RAW")
    
    return {
        "status": "success",
        "action": "appended",
        "sheet_url": spreadsheet.url
    }

# Retain old wrapper for compatibility (redirects to UPSERT on Sheet1)
def upload_to_google_sheets(data: Dict[str, Any]):
    return upsert_to_sheet("Sheet1", data, "enquiry", strict_mode=True)


@app.post("/init_headers")
async def init_headers():
    """Overwrite row 1 headers in the target sheet with the current field list + Timestamp."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)

    # Ensure fields loaded
    if not fields_cache:
        load_schema()

    spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
    sheet = spreadsheet.sheet1

    headers = [field['name'] for field in fields_cache.get("enquiry", [])] + ['Timestamp']
    # Replace entire first row with headers
    sheet.update('1:1', [headers], value_input_option='USER_ENTERED')


    return {
        "status": "success",
        "message": "Headers initialized",
        "sheet_url": spreadsheet.url,
        "headers": headers,
        "worksheet_title": sheet.title,
    }


def get_sheet_data_as_df():
    """Helper to fetch all data from Google Sheet into a Pandas DataFrame."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=500, detail="Google credentials not found")
    
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    
    spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
    sheet = spreadsheet.sheet1
    
    data = sheet.get_all_values()
    if not data:
        return pd.DataFrame(), sheet
        
    headers = data[0]
    rows = data[1:]
    
    # Handle duplicate headers if any
    unique_headers = []
    seen = set()
    for h in headers:
        c = h
        i = 1
        while c in seen:
            c = f"{h}_{i}"
            i += 1
        seen.add(c)
        unique_headers.append(c)
            
    df = pd.DataFrame(rows, columns=unique_headers)
    
    # Store sheet object for deletion reference? No, stateless.
    return df, sheet


def apply_delete_filters(df: pd.DataFrame, filters: FilterCriteria, date_col: str):
    """
    Returns a boolean mask where True means 'to be deleted'.
    """
    if date_col not in df.columns:
        raise HTTPException(status_code=400, detail=f"Column '{date_col}' not found in sheet")

    # Convert date column to datetime
    # Robust parsing:
    # 1. Clean whitespace
    # 2. Try dayfirst=False (ISO/US)
    # 3. If mostly NaT, try dayfirst=True (EU/India)
    try:
        clean_col = df[date_col].astype(str).str.strip()
        
        # Try 1: Standard/ISO
        temp_dates = pd.to_datetime(clean_col, errors='coerce', dayfirst=False)
        
        # Check success rate
        valid_count = temp_dates.notna().sum()
        total_count = len(df)
        
        # If very few valid dates found (less than 20%), try dayfirst=True
        if total_count > 0 and (valid_count / total_count) < 0.2:
             # Try 2: DayFirst
             print(f"Initial date parsing yield low success ({valid_count}/{total_count}). Retrying with dayfirst=True")
             temp_dates_alt = pd.to_datetime(clean_col, errors='coerce', dayfirst=True)
             if temp_dates_alt.notna().sum() > valid_count:
                 temp_dates = temp_dates_alt
                 
    except Exception as e:
        print(f"Date parsing critical error: {e}")
        raise HTTPException(status_code=400, detail=f"Date parsing failed: {str(e)}")

    # Create mask (initially all False)
    mask = pd.Series([False] * len(df), index=df.index)
    
    # Priority 1: Date Range (Start/End)
    if filters.startDate or filters.endDate:
        print(f"[DEBUG] Using date range filter: startDate={filters.startDate}, endDate={filters.endDate}")
        range_mask = pd.Series([True] * len(df), index=df.index)
        
        if filters.startDate:
            try:
                start = pd.to_datetime(filters.startDate)
                range_mask = range_mask & (temp_dates >= start)
                print(f"[DEBUG] After start date filter: {range_mask.sum()} matches")
            except Exception as e:
                print(f"[ERROR] Failed to parse startDate: {e}")
        
        if filters.endDate:
            try:
                end = pd.to_datetime(filters.endDate)
                # Include the entire end date (up to 23:59:59)
                end_inclusive = end + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                range_mask = range_mask & (temp_dates <= end_inclusive)
                print(f"[DEBUG] After end date filter: {range_mask.sum()} matches")
            except Exception as e:
                print(f"[ERROR] Failed to parse endDate: {e}")
        
        mask = range_mask
    
    # Priority 2: Specific Date
    elif filters.specificDate:
        print(f"[DEBUG] Using specific date filter: {filters.specificDate}")
        try:
            target_date = pd.to_datetime(filters.specificDate)
            mask = (temp_dates.dt.normalize() == target_date.normalize())
            print(f"[DEBUG] Specific date matches: {mask.sum()}")
        except Exception as e:
            print(f"[ERROR] Failed to parse specificDate: {e}")
    
    # Priority 3: Year/Month Logic
    elif filters.year or filters.month:
        print(f"[DEBUG] Using year/month filter: year={filters.year}, month={filters.month}")
        year_mask = pd.Series([True] * len(df), index=df.index)
        month_mask = pd.Series([True] * len(df), index=df.index)
        
        if filters.year:
            try:
                y = int(filters.year)
                year_mask = (temp_dates.dt.year == y)
                print(f"[DEBUG] Year {y} matches: {year_mask.sum()}")
            except Exception as e:
                print(f"[ERROR] Failed to parse year: {e}")
        
        if filters.month:
            try:
                m = int(filters.month)
                month_mask = (temp_dates.dt.month == m)
                print(f"[DEBUG] Month {m} matches: {month_mask.sum()}")
            except Exception as e:
                print(f"[ERROR] Failed to parse month: {e}")

        mask = year_mask & month_mask
        print(f"[DEBUG] Combined year/month matches: {mask.sum()}")
    else:
        # No filters provided -> delete nothing
        print(f"[DEBUG] No filters provided, returning empty mask")
        return pd.Series([False] * len(df), index=df.index), temp_dates
        
    # Remove NaT from mask (don't delete rows with invalid dates even if they match 'None' logic?)
    mask = mask & temp_dates.notna()
    
    return mask, temp_dates


@app.post("/delete/preview")
async def preview_delete(payload: DeletePreviewPayload):
    try:
        print(f"\n[DELETE PREVIEW] Starting preview with filters: {payload.filters}")
        print(f"[DELETE PREVIEW] Date column: {payload.date_column}")
        
        df, _ = get_sheet_data_as_df()
        print(f"[DELETE PREVIEW] Loaded {len(df)} rows from sheet")
        
        if df.empty:
            print(f"[DELETE PREVIEW] Sheet is empty")
            return {"count": 0, "rows": [], "earliest": None, "latest": None, "headers": []}
            
        mask, date_series = apply_delete_filters(df, payload.filters, payload.date_column)
        
        print(f"[DELETE PREVIEW] Filter mask created, {mask.sum()} rows match filters")
        print(f"[DELETE PREVIEW] Sample parsed dates (first 5): {date_series.head().tolist()}")
        print(f"[DELETE PREVIEW] Date column raw values (first 5): {df[payload.date_column].head().tolist()}")
        
        # Filter the DF
        matches = df[mask]
        count = len(matches)
        
        print(f"[DELETE PREVIEW] Final match count: {count}")
        
        if count == 0:
            return {"count": 0, "rows": [], "earliest": None, "latest": None}
            
        # Get Earliest/Latest from the matched dates
        matched_dates = date_series[mask]
        earliest = matched_dates.min().strftime('%Y-%m-%d') if not matched_dates.empty else None
        latest = matched_dates.max().strftime('%Y-%m-%d') if not matched_dates.empty else None
        
        # Prepare Preview Rows
        # Limit to 50
        preview_page = matches.head(50)
        
        # Select columns if requested
        if payload.preview_columns:
            # Only keep columns that exist
            valid_cols = [c for c in payload.preview_columns if c in preview_page.columns]
            if not valid_cols:
                # Fallback to all if none match (shouldn't happen with correct inputs)
                valid_cols = preview_page.columns.tolist()
            preview_page = preview_page[valid_cols]
            
        return {
            "count": count,
            "headers": preview_page.columns.tolist(),
            "rows": preview_page.to_dict(orient='records'),
            "earliest": earliest,
            "latest": latest
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delete/confirm")
async def confirm_delete(payload: DeleteConfirmPayload):
    try:
        df, sheet = get_sheet_data_as_df()
        if df.empty:
            return {"status": "error", "message": "Sheet is empty"}
            
        mask, _ = apply_delete_filters(df, payload.filters, payload.date_column)
        matches_count = mask.sum()
        
        if matches_count == 0:
             return {"status": "success", "message": "No rows matched the criteria (0 deleted)."}
             
        # Invert mask to get rows to KEEP
        # We need to preserve the dataframe structure perfectly
        df_keep = df[~mask]
        
        # Rewrite Sheet
        # 1. Clear
        sheet.clear()
        
        # 2. Write Headers
        headers = df_keep.columns.tolist()
        
        # 3. Write Rows
        # Gspread append_rows or update. allow massive update.
        # convert df_keep to list of lists
        # Handle NaN/None -> empty string for sheets
        df_keep = df_keep.fillna("")
        values = [headers] + df_keep.values.tolist()
        
        sheet.update(range_name='A1', values=values)
        
        return {
            "status": "success", 
            "message": f"Successfully deleted {matches_count} rows.",
            "remaining_count": len(df_keep)
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@app.on_event("startup")
async def startup_event():
    """Load schema on startup"""
    try:
        load_schema("enquiry")
        load_schema("admission")
        print(f"Loaded schema. Enquiry fields: {len(fields_cache['enquiry'])}, Admission fields: {len(fields_cache['admission'])}")
        ensure_notification_defaults()
        # Debug: Check SMTP configuration
        smtp_user_set = "YES" if SMTP_USERNAME else "NO"
        smtp_pass_set = "YES" if SMTP_PASSWORD else "NO"
        print(f"[SMTP Config] Username set: {smtp_user_set}, Password set: {smtp_pass_set}")
        
        # Start Home Care billing scheduler
        if HOMECARE_SCHEDULER_AVAILABLE:
            try:
                start_billing_scheduler()
                print("[Home Care Scheduler] Started successfully")
            except Exception as e:
                print(f"[Home Care Scheduler] Failed to start: {e}")
        
        # Start Patient Admission billing scheduler
        if PATIENTADMISSION_SCHEDULER_AVAILABLE:
            try:
                start_pa_billing_scheduler()
                print("[Patient Admission Scheduler] Started successfully")
            except Exception as e:
                print(f"[Patient Admission Scheduler] Failed to start: {e}")
        
    except Exception as e:
        print(f"Warning: Could not load fields on startup: {str(e)}")
        print("Ensure CSV or Excel file is present in backend directory")


@app.get("/")
async def root():
    return {"message": "CRM Lead Form API", "status": "running"}


@app.get("/get_fields")
async def get_fields(type: str = "enquiry"):
    """Return the current field schema. type = 'enquiry' | 'admission'"""
    if type not in ["enquiry", "admission"]:
        type = "enquiry"
        
    current = fields_cache.get(type)
    if not current:
        try:
            load_schema(type)
            current = fields_cache.get(type)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    return {"fields": current or []}


@app.get("/test_email")
async def test_email_get(to: str):
    to_addr = str(to or "").strip()
    if not to_addr:
        raise HTTPException(status_code=400, detail="Missing 'to' query parameter")
    try:
        send_notification_email(to_addr, {"test": "ping"})
        return {"status": "ok", "to": to_addr}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"fields": fields_cache}


@app.post("/test_email")
async def test_email(payload: Dict[str, Any]):
    to_addr = str(payload.get("to", "")).strip()
    print(f"[TEST_EMAIL] Received request for {to_addr}")
    if not to_addr:
        raise HTTPException(status_code=400, detail="Missing 'to' in payload")
    try:
        print(f"[TEST_EMAIL] About to call send_notification_email...")
        send_notification_email(to_addr, {"test": "ping"})
        print(f"[TEST_EMAIL] send_notification_email returned")
        return {"status": "ok", "to": to_addr}
    except Exception as e:
        print(f"[TEST_EMAIL] Exception: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/notification_settings")
async def get_notification_settings():
    settings = ensure_notification_defaults()
    return {
        "sender_email": settings.get("sender_email", DEFAULT_SENDER_EMAIL),
        "cc_emails": settings.get("cc_emails", []),
    }


@app.post("/notification_settings")
async def update_notification_settings(payload: NotificationSettings):
    sender = str(payload.sender_email).strip()
    if not sender:
        raise HTTPException(status_code=400, detail="Sender email must not be empty")
    if "@" not in sender:
        raise HTTPException(status_code=400, detail="Sender email appears invalid")
    cc_emails: List[str] = []
    if payload.cc_emails:
        for item in payload.cc_emails:
            if not item:
                continue
            item_str = str(item).strip()
            if not item_str:
                continue
            if "@" not in item_str:
                raise HTTPException(status_code=400, detail=f"Invalid CC email: {item_str}")
            cc_emails.append(item_str)

    save_notification_settings({"sender_email": sender, "cc_emails": cc_emails})
    return {"status": "success", "sender_email": sender, "cc_emails": cc_emails}


@app.post("/upload_file")
async def upload_file(file: UploadFile = File(...)):
    """
    Handle file uploads.
    - Stores non-data files safely.
    - Processes data files (Excel/CSV) for schema changes.
    """
    try:
        # 1. Save the file
        file_path = save_upload(file, file.filename)
        
        # 2. Determine if it's a data file
        filename = file.filename.lower()
        is_data_file = filename.endswith(('.csv', '.xlsx', '.xls', '.xlsm'))
        
        if is_data_file:
            # Load current schema to compare (Enquiry default)
            existing_schema = fields_cache.get("enquiry")
            if not existing_schema:
                try:
                    load_schema("enquiry")
                    existing_schema = fields_cache.get("enquiry")
                except:
                    existing_schema = []
            
            result = process_data_file(file_path, existing_schema)
            # Add file path to result
            result['file_path'] = file_path
            return result
        else:
            # Non-data file
            return {
                "status": "success",
                "message": f"File '{file.filename}' stored successfully.",
                "file_path": file_path,
                "type": "non-data"
            }
            
    except Exception as e:
        print(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ConfirmUploadRequest(BaseModel):
    file_path: str


@app.post("/confirm_upload")
async def confirm_upload(payload: ConfirmUploadRequest):
    """
    Reads the previously uploaded file and APPENDS data to 'Sheet1' in Google Sheets.
    It maps columns by name to match the existing sheet headers.
    """
    file_path = payload.file_path
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials not found")

    try:
        # 1. Read Data
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(file_path)
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(file_path)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")

        df = df.fillna("")

        # 2. Connect to Sheets
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except gspread.WorksheetNotFound:
            sheet = spreadsheet.sheet1

        # 3. Get Existing Headers
        existing_values = sheet.get_all_values()
        if not existing_values:
            # Sheet is empty, write headers from file
            sheet_headers = df.columns.tolist() + ['Timestamp']
            sheet.append_row(sheet_headers, value_input_option='USER_ENTERED')
        else:
            sheet_headers = existing_values[0]

        # 3b. Detect and Add New Columns
        # Compare df columns to sheet_headers (case-insensitive check)
        sheet_headers_lower = {h.strip().lower() for h in sheet_headers}
        new_columns = []
        for col in df.columns:
            if col.strip().lower() not in sheet_headers_lower:
                new_columns.append(col)
        
        if new_columns:
            # Append new columns to header row in Sheet
            # We need to find the next available column index
            # But gspread's append_row doesn't work for partial row updates easily on existing rows
            # simplest way: update the first row with the extended list
            sheet_headers.extend(new_columns)
            sheet.update(range_name='1:1', values=[sheet_headers], value_input_option='USER_ENTERED')

        # 4. Map Data to Headers (now including new ones)
        # Create a map for case-insensitive matching of file columns
        file_cols_map = {c.strip().lower(): c for c in df.columns}
        
        data_to_append = []
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for _, row in df.iterrows():
            ordered_row = []
            for h in sheet_headers:
                h_lower = h.strip().lower()
                
                # Special handling for Timestamp
                if h_lower == 'timestamp':
                    ordered_row.append(current_time)
                    continue

                # Find matching column in file
                if h_lower in file_cols_map:
                    original_col_name = file_cols_map[h_lower]
                    # Handle NaN values to ensure empty string
                    val = row[original_col_name]
                    ordered_row.append(val if pd.notna(val) else "")
                else:
                    # Column exists in Sheet but not in File -> Empty
                    ordered_row.append("")
            
            data_to_append.append(ordered_row)

        if data_to_append:
            sheet.append_rows(data_to_append, value_input_option='USER_ENTERED')
            
        message = f"Successfully appended {len(data_to_append)} rows."
        if new_columns:
            message += f" Added {len(new_columns)} new columns: {', '.join(new_columns)}."

        return {
            "status": "success",
            "message": message,
            "sheet_url": spreadsheet.url
        }

    except Exception as e:
        print(f"Bulk update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ConfirmUploadRequest(BaseModel):
    file_path: str


@app.post("/confirm_upload")
async def confirm_upload(payload: ConfirmUploadRequest):
    """
    Reads the previously uploaded file and APPENDS data to 'Sheet1' in Google Sheets.
    It maps columns by name to match the existing sheet headers.
    """
    file_path = payload.file_path
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials not found")

    try:
        # 1. Read Data
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(file_path)
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(file_path)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")

        df = df.fillna("")

        # 2. Connect to Sheets
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except gspread.WorksheetNotFound:
            # Create if not exists? Or append to first sheet?
            sheet = spreadsheet.get_worksheet(0)
            
        # 3. Append to Sheet
        # Get existing headers
        existing_data = sheet.get_all_values()
        if not existing_data:
            # Sheet is empty, write fresh with headers
            sheet.update([df.columns.values.tolist()] + df.values.tolist())
        else:
            existing_headers = existing_data[0]
            
            # Helper for normalization
            def normalize_header(h):
                return str(h).lower().replace(" ", "").replace("_", "").strip()

            # Map normalized -> actual column name in DF
            df_norm_map = {normalize_header(c): c for c in df.columns}
            
            # Align new DF to existing headers
            # Add missing columns to DF using Fuzzy Match
            for h in existing_headers:
                h_norm = normalize_header(h)
                if h not in df.columns:
                    # Check fuzzy match
                    if h_norm in df_norm_map:
                        actual_col = df_norm_map[h_norm]
                        # Rename column to match existing header
                        df.rename(columns={actual_col: h}, inplace=True)
                        # Update map to reflect rename? No need, we found it.
                    else:
                        df[h] = ""
            
            # Reorder DF to match existing headers (discard extra new columns to follow strict schema if desired, 
            # OR logic to append new columns? For now, stick to existing schema as primary)
            df_to_append = df[existing_headers]
            
            sheet.append_rows(df_to_append.values.tolist())

        return {
            "status": "success",
            "message": f"Successfully appended {len(df)} rows to Google Sheet."
        }

    except Exception as e:
        print(f"Confirm update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/get_sheet_headers")
async def get_sheet_headers():
    """
    Fetch headers directly from the CRM Sheet (Sheet1) to populate column options for File Manager.
    """
    try:
        # Connect to Google Sheets
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except:
            sheet = spreadsheet.get_worksheet(0)
            
        # Get Header Row
        headers = sheet.row_values(1)
        
        fields = [{"name": h, "label": h, "type": "text"} for h in headers if h]
        
        return {"fields": fields}
        
    except Exception as e:
        print(f"Get sheet headers failed: {e}")
        # Fallback to empty
        return {"fields": []}


@app.get("/download_template")
async def download_template(format: str = "xlsx"):
    """
    Generate a dynamic template (CSV or Excel) based on the OFFICIAL_COLUMNS constant.
    This ensures exact format matching regardless of current sheet state.
    """
    try:
        # Use Fixed Schema
        headers = OFFICIAL_COLUMNS
        
        if format.lower() == 'csv':
            # Create CSV
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            output.seek(0)
            
            response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
            response.headers["Content-Disposition"] = "attachment; filename=contact_template.csv"
            return response
            
        else:
            # Create Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(headers)
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response.headers["Content-Disposition"] = "attachment; filename=contact_template.xlsx"
            return response

    except Exception as e:
            print(f"Download template failed: {e}")
            # Return error as text if failed
            return Response(content=f"Error generating template: {str(e)}", status_code=500)


# Patient search cache to prevent API rate limits
from datetime import datetime, timedelta
patient_search_cache = {
    "data": None,
    "timestamp": None,
    "ttl_minutes": 5  # Cache for 5 minutes
}

def get_cached_patients():
    """Get patients from cache if available and not expired"""
    global patient_search_cache
    
    if patient_search_cache["data"] is None:
        return None
    
    if patient_search_cache["timestamp"] is None:
        return None
    
    # Check if cache is expired
    cache_age = datetime.now() - patient_search_cache["timestamp"]
    if cache_age > timedelta(minutes=patient_search_cache["ttl_minutes"]):
        print("[Patient Search] Cache expired, will refresh")
        return None
    
    print(f"[Patient Search] Using cached data (age: {cache_age.seconds}s)")
    return patient_search_cache["data"]

def update_patient_cache(data):
    """Update the patient search cache"""
    global patient_search_cache
    patient_search_cache["data"] = data
    patient_search_cache["timestamp"] = datetime.now()
    print(f"[Patient Search] Cache updated with {len(data)} records")


@app.get("/api/patients/search")
async def search_patients(q: str = ""):
    """
    Search patients from CRM_Lead->Sheet1
    Returns Member ID Key and Patient Name for dropdown selection
    If query is empty, returns all patients (for dropdown population)
    """
    try:
        # Force reload - Updated to return all patients when query is empty
        # Try to get from cache first
        all_records = get_cached_patients()
        
        # If cache miss, fetch from Google Sheets
        if all_records is None:
            print("[Patient Search] Cache miss, fetching from Google Sheets")
            # Connect to Google Sheets
            scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
            client = gspread.authorize(creds)
            
            if not GOOGLE_SHEET_ID:
                raise HTTPException(status_code=500, detail="Google Sheet ID not configured")
            
            spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
            sheet = spreadsheet.worksheet("Sheet1")
            
            # Get all records
            all_records = sheet.get_all_records()
            
            # Update cache
            update_patient_cache(all_records)
        
        # Debug: Print available columns from first record
        if all_records:
            print(f"[Patient Search DEBUG] Total records: {len(all_records)}")
            print(f"[Patient Search DEBUG] Available columns: {list(all_records[0].keys())}")
            
            # Print first record completely to see the data structure
            if len(all_records) > 0:
                first_record = all_records[0]
                print(f"[Patient Search DEBUG] First record data:")
                for key, value in first_record.items():
                    if 'member' in key.lower() or 'patient' in key.lower() or 'name' in key.lower():
                        print(f"  - {key}: '{value}'")
        
        results = []
        
        # If query is empty, return all patients (limit to 100 for performance)
        if not q or len(q.strip()) < 1:
            print("[Patient Search] Empty query - returning all patients")
            for record in all_records[:100]:  # Limit to first 100 patients
                # Get patient name
                patient_name = str(record.get("Patient Name", "") or record.get("patient name", "")).strip()
                
                # Get Member ID Key
                member_id = ""
                for key in record.keys():
                    key_lower = str(key).lower().strip()
                    if "member" in key_lower and "id" in key_lower:
                        value = record.get(key, "")
                        if value and value is not None and str(value).strip() and str(value).strip().lower() != 'none':
                            member_id = str(value).strip()
                            break
                
                # Skip if both are empty
                if not patient_name and not member_id:
                    continue
                
                # Get additional fields for auto-population
                gender = str(record.get("GENDER", "") or record.get("Gender", "") or record.get("gender", "")).strip()
                age = str(record.get("AGE", "") or record.get("Age", "") or record.get("age", "")).strip()
                location = str(record.get("LOCATION", "") or record.get("Location", "") or record.get("location", "")).strip()
                
                results.append({
                    "member_id": member_id,
                    "patient_name": patient_name,
                    "gender": gender,
                    "age": age,
                    "location": location,
                    "display": f"{member_id} | {patient_name}" if member_id and patient_name else (member_id or patient_name)
                })
            
            print(f"[Patient Search DEBUG] Returning {len(results)} total patients")
            return {
                "status": "success",
                "patients": results
            }
        
        # If query is provided, search for matches
        query_lower = q.strip().lower()
        
        # Debug: Show what column names we're actually seeing
        if all_records:
            actual_columns = list(all_records[0].keys())
            member_id_columns = [col for col in actual_columns if 'member' in col.lower() and 'id' in col.lower()]
            print(f"[Patient Search DEBUG] Columns with 'member' and 'id': {member_id_columns}")
        
        for record in all_records:
            # Get patient name with various possible column name variations
            patient_name = str(record.get("Patient Name", "") or record.get("patient name", "")).strip()
            
            # Get Member ID Key - FIRST check the exact column names from the sheet
            member_id = ""
            
            # Strategy 1: Check all actual column names in the record
            for key in record.keys():
                key_lower = str(key).lower().strip()
                # Check if this column name contains both "member" and "id"
                if "member" in key_lower and "id" in key_lower:
                    value = record.get(key, "")
                    # Handle None, empty strings, and the string "None"
                    if value and value is not None and str(value).strip() and str(value).strip().lower() != 'none':
                        member_id = str(value).strip()
                        if len(results) < 3:  # Debug first few
                            print(f"[Patient Search DEBUG] Found Member ID in column '{key}' = '{member_id}'")
                        break
            
            # Strategy 2: Fallback to exact string matches if Strategy 1 didn't work
            if not member_id:
                for key_variant in [
                    "Member ID Key", "Member ID key", "member id key", 
                    "Member Id Key", "MEMBER ID KEY",
                    "Member_ID_Key", "member_id_key", "MemberIDKey", "Memberidkey",
                    "Member ID", "member id", "MemberID", "memberid",
                    "ID", "id"
                ]:
                    if key_variant in record and record[key_variant]:
                        member_id = str(record[key_variant]).strip()
                        break
            
            # Debug: Print what we found for first few matches
            if len(results) < 3 and (query_lower in patient_name.lower() or query_lower in member_id.lower()):
                print(f"[Patient Search DEBUG] Found match - Name: '{patient_name}', ID: '{member_id}'")
            
            # Skip if both are empty
            if not patient_name and not member_id:
                continue
            
            # Search in both patient name and member ID
            if (query_lower in patient_name.lower() or 
                query_lower in member_id.lower()):
                
                # Get additional fields for auto-population
                gender = str(record.get("GENDER", "") or record.get("Gender", "") or record.get("gender", "")).strip()
                age = str(record.get("AGE", "") or record.get("Age", "") or record.get("age", "")).strip()
                location = str(record.get("LOCATION", "") or record.get("Location", "") or record.get("location", "")).strip()
                
                results.append({
                    "member_id": member_id,
                    "patient_name": patient_name,
                    "gender": gender,
                    "age": age,
                    "location": location,
                    "display": f"{member_id} | {patient_name}" if member_id and patient_name else (member_id or patient_name)
                })
        
        # Limit to 50 results
        results = results[:50]
        
        print(f"[Patient Search DEBUG] Returning {len(results)} results")
        
        return {
            "status": "success",
            "patients": results
        }
        
    except Exception as e:
        print(f"Error searching patients: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to search patients: {str(e)}")


class DeletePreviewRequest(BaseModel):
    filters: Dict[str, Any]
    date_column: str
    preview_columns: List[str] = []

@app.post("/delete/preview")
async def delete_preview(payload: DeletePreviewRequest):
    """
    Preview rows that would be deleted based on filters.
    """
    try:
        # Connect & Read Sheet
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except:
            sheet = spreadsheet.get_worksheet(0)
            
        all_values = sheet.get_all_values()
        if not all_values:
            return {"count": 0, "rows": [], "headers": []}
            
        headers = all_values[0]
        rows = all_values[1:]
        
        df = pd.DataFrame(rows, columns=headers)
        
        # Apply Filters
        date_col = payload.date_column
        filters = payload.filters
        
        # Case-insensitive check
        cols_lower = {c.lower(): c for c in df.columns}
        if date_col.lower() not in cols_lower:
             raise HTTPException(status_code=400, detail=f"Column '{date_col}' not found in sheet. Available: {list(df.columns)}")
             
        actual_date_col = cols_lower[date_col.lower()]
             
        # Convert date column to datetime for filtering
        # Handle various formats?
        # Flexible date parsing to handle mixed formats (DD/MM/YYYY and YYYY-MM-DD)
        # First pass: try dayfirst=True (handles 11/12/2025 correctly)
        df['temp_date'] = pd.to_datetime(df[actual_date_col], dayfirst=True, errors='coerce')
        
        # Second pass: for any NaT (failed parse), try default ISO parsing (handles 2025-12-11)
        mask_nat = df['temp_date'].isna()
        if mask_nat.any():
            df.loc[mask_nat, 'temp_date'] = pd.to_datetime(df.loc[mask_nat, actual_date_col], errors='coerce')
        # Fallback for YYYY-MM-DD if dayfirst failed? Pandas usually handles mixed well but let's assume standard.
        
        mask = pd.Series([True] * len(df))
        
        if filters.get('specificDate'):
            target = pd.to_datetime(filters['specificDate']).date()
            mask = mask & (df['temp_date'].dt.date == target)
            
        if filters.get('startDate'):
            start = pd.to_datetime(filters['startDate']).date()
            mask = mask & (df['temp_date'].dt.date >= start)
            
        if filters.get('endDate'):
            end = pd.to_datetime(filters['endDate']).date()
            mask = mask & (df['temp_date'].dt.date <= end)
            
        if filters.get('month'):
            m = int(filters['month'])
            mask = mask & (df['temp_date'].dt.month == m)
            
        if filters.get('year'):
            y = int(filters['year'])
            mask = mask & (df['temp_date'].dt.year == y)
            
        filtered_df = df[mask]
        
        # Prepare result
        result_headers = payload.preview_columns if payload.preview_columns else headers[:5]
        # Validate result headers exist
        result_headers = [h for h in result_headers if h in df.columns]
        
        result_rows = filtered_df[result_headers].head(10).to_dict('records')
        
        # Get date range of matching rows for metadata
        earliest = filtered_df['temp_date'].min()
        latest = filtered_df['temp_date'].max()
        
        return {
            "count": len(filtered_df),
            "rows": result_rows,
            "headers": result_headers,
            "earliest": earliest.strftime('%Y-%m-%d') if pd.notnull(earliest) else None,
            "latest": latest.strftime('%Y-%m-%d') if pd.notnull(latest) else None,
            "debug_info": {
                "active_sheet_id": GOOGLE_SHEET_ID,
                "date_col_used": actual_date_col,
                "raw_sample": df[actual_date_col].head(5).tolist(),
                "parsed_sample": df['temp_date'].head(5).astype(str).tolist(),
                "nat_count": int(df['temp_date'].isna().sum()),
                "total_rows": len(df),
                "filter_start": str(filters.get('startDate')),
                "filter_end": str(filters.get('endDate'))
            }
        }
        
    except Exception as e:
        print(f"Delete preview failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delete/confirm")
async def delete_confirm(payload: DeletePreviewRequest):
    """
    Execute deletion. Matches logic of preview but writes back the INVERSE set.
    """
    try:
        # Connect & Read Sheet
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except:
            sheet = spreadsheet.get_worksheet(0)
            
        all_values = sheet.get_all_values()
        if not all_values:
            return {"message": "Sheet is empty."}
            
        headers = all_values[0]
        rows = all_values[1:]
        
        df = pd.DataFrame(rows, columns=headers)
        
        # Apply Filters to find rows to DELETE
        date_col = payload.date_column
        filters = payload.filters
        
        # Case-insensitive check
        cols_lower = {c.lower(): c for c in df.columns}
        if date_col.lower() not in cols_lower:
             raise HTTPException(status_code=400, detail=f"Column '{date_col}' not found. Available: {list(df.columns)}")
             
        actual_date_col = cols_lower[date_col.lower()]
             
        # Flexible date parsing to handle mixed formats (DD/MM/YYYY and YYYY-MM-DD)
        # First pass: try dayfirst=True (handles 11/12/2025 correctly)
        df['temp_date'] = pd.to_datetime(df[actual_date_col], dayfirst=True, errors='coerce')
        
        # Second pass: for any NaT (failed parse), try default ISO parsing (handles 2025-12-11)
        mask_nat = df['temp_date'].isna()
        if mask_nat.any():
            df.loc[mask_nat, 'temp_date'] = pd.to_datetime(df.loc[mask_nat, actual_date_col], errors='coerce')
        
        mask = pd.Series([True] * len(df))
        
        if filters.get('specificDate'):
            target = pd.to_datetime(filters['specificDate']).date()
            mask = mask & (df['temp_date'].dt.date == target)
        if filters.get('startDate'):
            start = pd.to_datetime(filters['startDate']).date()
            mask = mask & (df['temp_date'].dt.date >= start)
        if filters.get('endDate'):
            end = pd.to_datetime(filters['endDate']).date()
            mask = mask & (df['temp_date'].dt.date <= end)
        if filters.get('month'):
            m = int(filters['month'])
            mask = mask & (df['temp_date'].dt.month == m)
        if filters.get('year'):
            y = int(filters['year'])
            mask = mask & (df['temp_date'].dt.year == y)
            
        # Rows to KEEP are those where mask is False (or we select ~mask, but beware of NAs in date conversion?)
        # Logic: mask selects rows to DELETE.
        # So we keep rows where mask IS FALSE.
        
        rows_to_delete_count = mask.sum()
        
        if rows_to_delete_count == 0:
            return {"message": "No rows matched the criteria. Nothing deleted."}
            
        df_keep = df[~mask]
        
        # Remove temp col
        if 'temp_date' in df_keep.columns:
            df_keep = df_keep.drop(columns=['temp_date'])
            
        # Write back
        # Clear sheet
        sheet.clear()
        
        # Update with headers + kept rows
        # Replace NaN with ""
        df_keep = df_keep.fillna("")
        
        updated_data = [df_keep.columns.values.tolist()] + df_keep.values.tolist()
        sheet.update(updated_data)
        
        return {"message": f"Deleted {rows_to_delete_count} rows successfully."}
        
    except Exception as e:
        print(f"Delete confirm failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sync_fields")
async def sync_fields():
    """Read Excel, rebuild schema, persist locally, update CSV and Google Sheet (headers + dropdowns)."""
    global fields_cache
    try:
        schema = build_schema_from_excel()
        save_field_schema_to_disk(schema)
        fields_cache["enquiry"] = schema
        # Update local CSV header row
        try:
            header_row = [f["name"] for f in schema]
            rows: List[List[str]] = []
            if os.path.exists(CSV_FILE_PATH):
                with open(CSV_FILE_PATH, "r", encoding="utf-8-sig", newline="") as rf:
                    reader = csv.reader(rf)
                    existing = list(reader)
                rows = existing[1:] if len(existing) > 0 else []
            with open(CSV_FILE_PATH, "w", encoding="utf-8", newline="") as wf:
                writer = csv.writer(wf)
                writer.writerow(header_row)
                for r in rows:
                    writer.writerow(r)
        except Exception as e:
            print(f"Warning: failed updating local CSV: {e}")

        # Update Google Sheet
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        headers = [f["name"] for f in schema] + ["Timestamp"]
        
        # 1. Sync Sheet1 (Master)
        try:
            sheet1 = spreadsheet.sheet1
            sheet1.update('1:1', [headers], value_input_option='USER_ENTERED')
        except Exception as e:
            print(f"Warning: failed syncing Sheet1 headers: {e}")

        # 2. Sync Enquiries Sheet
        try:
            try:
                sheet_enq = spreadsheet.worksheet(ENQUIRIES_SHEET_NAME)
            except gspread.WorksheetNotFound:
                sheet_enq = spreadsheet.add_worksheet(title=ENQUIRIES_SHEET_NAME, rows=1000, cols=20)
            sheet_enq.update('1:1', [headers], value_input_option='USER_ENTERED')
        except Exception as e:
            print(f"Warning: failed syncing Enquiries sheet headers: {e}")

        sync_dropdown_options_to_sheet(client, spreadsheet, schema)
        return {"status": "success", "fields": schema, "sheet_url": spreadsheet.url}
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {e}")


class UpdateSingleFieldPayload(BaseModel):
    original_name: str
    name: Optional[str] = None
    label: Optional[str] = None
    data_type: Optional[str] = None

class BillingInputs(BaseModel):
    room_charge: float
    bed_charge: float
    nurse_payment: float
    additional_nurse_payment: float = 0
    other_charges_amenities: float = 0
    hospital_payment: float
    doctor_fee: float
    service_charge: float
    discount: float = 0

class BillingExportRequest(BaseModel):
    member_id: str
    patient_data: Dict[str, Any]
    billing_inputs: BillingInputs
    total_amount: float

class BillingSaveRequest(BaseModel):
    member_id: str
    billing_data: BillingInputs
    patient_data: Dict[str, Any]
    total_days: int
    grand_total: float

@app.post("/billing-summary/save")
async def save_billing_summary(payload: BillingSaveRequest):
    """
    Save calculated billing summary to Sheet1 (Master).
    Updates existing row for the Member ID with billing details.
    Adds billing columns if they do not exist.
    """
    # Normalize patient_data to avoid duplicate column issues
    if payload.patient_data:
        payload.patient_data = normalize_payload(payload.patient_data)

    print(f"[Billing Save] Saving summary for {payload.member_id}")
    
    try:
        if not os.path.exists(CREDENTIALS_FILE):
             raise HTTPException(status_code=404, detail="Google credentials not found")

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        spreadsheet = ensure_google_sheet(client)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except:
            sheet = spreadsheet.sheet1
            
        # 1. Define Billing Columns
        billing_columns = [
            "TotalDaysStayed",
            "RoomChargeTotal",
            "BedChargeTotal", 
            "NursePaymentTotal",
            "AdditionalNursePaymentTotal",
            "OtherChargesTotal",
            "HospitalPaymentTotal", 
            "DoctorFee", 
            "ServiceCharge",
            "Discount",
            "BillGrandTotal", 
            "BillGeneratedDate"
        ]
        
        # 2. Check & Update Headers
        existing_values = sheet.get_all_values()
        if not existing_values:
            headers = []
        else:
            headers = existing_values[0]
            
        headers_lower = [h.strip().lower() for h in headers]
        new_cols = []
        
        for col in billing_columns:
            if col.lower() not in headers_lower:
                new_cols.append(col)
                
        if new_cols:
            print(f"[Billing Save] Adding new columns: {new_cols}")
            headers.extend(new_cols)
            sheet.update(range_name='1:1', values=[headers], value_input_option='USER_ENTERED')
            # Refresh headers after update
            headers_lower = [h.strip().lower() for h in headers]

        # 3. Find Row by Member ID
        member_id_col_idx = -1
        target_id = payload.member_id.strip().lower().replace(" ", "").replace("_", "").replace("-", "")
        
        # Find ID column index
        id_keys = ["memberidkey", "memberid", "member_id", "mid", "patientid", "id"]
        for idx, h in enumerate(headers):
            h_norm = normalize_field_name(h)
            if h_norm in id_keys:
                member_id_col_idx = idx
                break
                
        if member_id_col_idx == -1:
            raise HTTPException(status_code=400, detail="Member ID column not found in Sheet1")
            
        # Find the row
        row_to_update = -1
        # Fetch ID column values
        id_col_values = sheet.col_values(member_id_col_idx + 1)
        
        # Skip header
        for i, val in enumerate(id_col_values):
            if i == 0: continue
            if str(val).strip().lower().replace(" ", "").replace("_", "").replace("-", "") == target_id:
                row_to_update = i + 1
                break
                
        if row_to_update == -1:
             raise HTTPException(status_code=404, detail=f"Member ID {payload.member_id} not found in Sheet1")

        # 4. Prepare Values
        current_date = datetime.now().strftime('%Y-%m-%d')
        b_data = payload.billing_data
        
        # Calculate totals
        # Note: Frontend sends totals usually, but let's trust frontend inputs * days or just inputs?
        # User prompt Request 2: "Auto calculate totals... GrandTotal = sum of all above"
        # The payload has 'billing_data' which seems to be the RATES/Inputs.
        # But wait, Request 4 says save "RoomChargeTotal", not "RoomRate".
        # So we need to calculate the totals here or accept them from frontend.
        # The prompt says: "Auto-update... Allow user to modify... Grand total must recalculate"
        # The frontend state has `billingInputs` (rates/fees).
        # We should calculate the Totals based on `total_days`.
        
        days = payload.total_days if payload.total_days > 0 else 1
        
        # Values to save
        vals_map = {
            "TotalDaysStayed": days,
            "RoomChargeTotal": b_data.room_charge * days,
            "BedChargeTotal": b_data.bed_charge * days,
            "NursePaymentTotal": b_data.nurse_payment * days,
            "AdditionalNursePaymentTotal": b_data.additional_nurse_payment * days,
            "OtherChargesTotal": b_data.other_charges_amenities * days,
            "HospitalPaymentTotal": b_data.hospital_payment * days,
            "DoctorFee": b_data.doctor_fee,
            "ServiceCharge": b_data.service_charge,
            "Discount": b_data.discount,
            "BillGrandTotal": payload.grand_total,
            "BillGeneratedDate": current_date
        }
        
        # 5. Update Cells
        # We need to find the column index for each field and update that specific cell
        # Batch update is better
        
        cells_to_update = []
        for col_name, val in vals_map.items():
            # Find col index
            try:
                # header list is 0-indexed, gspread cols are 1-indexed
                col_idx = -1
                for idx, h in enumerate(headers):
                    if normalize_header(h) == normalize_header(col_name):
                        col_idx = idx + 1
                        break
                
                if col_idx != -1:
                    cells_to_update.append({
                        'range': gspread.utils.rowcol_to_a1(row_to_update, col_idx),
                        'values': [[val]]
                    })
            except Exception as loop_e:
                print(f"[Billing Save] Error mapping col {col_name}: {loop_e}")

        if cells_to_update:
            # Gspread batch_update requires a different format or multiple calls
            # simple update for each is fine for low volume, or batch_update
            # batch_update accepts a list of range/values objects? No, that's API dict.
            # Client.batch_update is for spreadsheet. sheet.batch_update takes data list.
            
            batch_data = []
            for item in cells_to_update:
                batch_data.append(item)
                
            sheet.batch_update(batch_data, value_input_option='USER_ENTERED')
            
        print(f"[Billing Save] Success for row {row_to_update}")
        return {"status": "success", "message": "Billing details saved to Sheet1", "row": row_to_update}

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
async def get_admission_details(member_id: str = Query(..., alias="member_id")):
    """
    Fetch patient admission details from the Patient Admission Google Sheet or Sheet1 as fallback.
    """
    print(f"[Admission Search] Searching for Member ID: {member_id}")
    
    # helper to normalize keys for comparison
    def norm_key(k):
        return str(k).strip().lower().replace(" ", "").replace("_", "").replace("-", "")

    target_id = norm_key(member_id)
    
    # 1. Try Patient Admission Sheet First (Secondary)
    try:
        client, spreadsheet = get_patient_admission_sheet_client()
        # Try finding the right worksheet - usually Sheet1 or by name
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except:
            sheet = spreadsheet.sheet1
            
        records = sheet.get_all_records()
        print(f"[Admission Search] Found {len(records)} records in Secondary Sheet")
        
        for row in records:
            # Find ID in this row
            # Filter keys that look like member id
            for k, v in row.items():
                if norm_key(k) in ["memberidkey", "memberid", "member_id", "mid", "patientid", "id"]:
                    if norm_key(v) == target_id:
                        print(f"[Admission Search] Match found in Secondary Sheet")
                        return row
    except Exception as e:
        print(f"[Admission Search] Error checking Secondary Sheet: {e}")

    # 2. Fallback to Master Sheet (Sheet1) if not found or error
    try:
        df, _ = get_sheet_data_as_df()
        if not df.empty:
            # Normalize column names for search
            # We need to find the column that holds member ID
            id_col = None
            for col in df.columns:
                if norm_key(col) in ["memberidkey", "memberid", "member_id", "mid", "patientid", "id"]:
                    id_col = col
                    break
            
            if id_col:
                # Search in DF
                # Use string comparison
                match = df[df[id_col].astype(str).map(norm_key) == target_id]
                if not match.empty:
                    # Return first match as dict
                    # Replace NaN with None or empty string
                    data = match.iloc[0].where(pd.notnull(match.iloc[0]), "").to_dict()
                    print(f"[Admission Search] Match found in Master Sheet")
                    return data
    except Exception as e:
         print(f"[Admission Search] Error checking Master Sheet: {e}")

    raise HTTPException(status_code=404, detail="Member ID not found in admission records")


@app.post("/billing-summary/export")
async def export_billing_summary(payload: BillingExportRequest):
    """
    Generate an Excel file with Patient Details and Billing Summary.
    """
    print(f"[Billing Export] Generating summary for {payload.member_id}")
    
    try:
        # Create a BytesIO buffer
        output = io.BytesIO()
        
        # We will create a structured Excel file
        # Use pandas ExcelWriter
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # 1. Patient Details Sheet
            # Convert patient_data dict to DataFrame (vertical layout)
            p_data = {
                "Field": list(payload.patient_data.keys()),
                "Value": list(payload.patient_data.values())
            }
            df_patient = pd.DataFrame(p_data)
            df_patient.to_excel(writer, sheet_name='Billing Summary', startrow=0, startcol=0, index=False)
            
            # Access the workbook and sheet
            workbook = writer.book
            worksheet = writer.sheets['Billing Summary']
            
            # Add Styling / Spacing
            # We'll write the Calculation part below the patient details
            
            start_row = len(df_patient) + 4
            
            # Header for Billing
            worksheet.cell(row=start_row, column=1, value="BILLING CALCULATION")
            worksheet.cell(row=start_row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
            
            # Billing Details
            b_inputs = payload.billing_inputs
            billing_rows = [
                ["Item", "Amount / Value"],
                ["Room Charge", b_inputs.room_charge],
                ["Bed Charge", b_inputs.bed_charge],
                ["Nurse Payment", b_inputs.nurse_payment],
                ["Additional Nurse Payment", b_inputs.additional_nurse_payment],
                ["Other Charges (Amenities)", b_inputs.other_charges_amenities],
                ["Hospital Payment", b_inputs.hospital_payment],
                ["Doctor Fee", b_inputs.doctor_fee],
                ["Service Charge", b_inputs.service_charge],
                ["Discount", b_inputs.discount],
                ["TOTAL AMOUNT", payload.total_amount]
            ]
            
            for i, row in enumerate(billing_rows):
                worksheet.cell(row=start_row + 2 + i, column=1, value=row[0])
                worksheet.cell(row=start_row + 2 + i, column=2, value=row[1])
                
            # Bold Total
            last_row = start_row + 2 + len(billing_rows) - 1
            worksheet.cell(row=last_row, column=1).font = openpyxl.styles.Font(bold=True)
            worksheet.cell(row=last_row, column=2).font = openpyxl.styles.Font(bold=True)
            
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="Billing_Summary_{payload.member_id}.xlsx"'
        }
        
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        print(f"[Billing Export] Error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/update_fields")
async def update_fields(payload: UpdateFieldsPayload, type: str = "enquiry"):
    """
    Bulk update fields schema.
    Replaces the entire schema list for the given type.
    """
    global fields_cache
    if type not in ["enquiry", "admission"]:
        type = "enquiry"
    
    # Convert Pydantic models to dicts
    new_schema = [field.dict() for field in payload.fields]
    
    # Update Cache
    fields_cache[type] = new_schema
    
    # Save to Disk
    try:
        save_field_schema_to_disk(new_schema, type)
        print(f"Schema saved for {type}: {len(new_schema)} fields")
        return {"status": "success", "message": "Schema saved successfully", "count": len(new_schema)}
    except Exception as e:
        print(f"Failed to save schema: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    options: Optional[List[str]] = None
    type: str = "enquiry"  # 'enquiry' or 'admission'


@app.post("/update_field")
async def update_field(payload: UpdateSingleFieldPayload):
    """Modify one field; persist, update CSV and Google Sheet."""
    global fields_cache
    schema_type = payload.type
    if schema_type not in ["enquiry", "admission"]:
        schema_type = "enquiry"

    current_schema = fields_cache.get(schema_type, [])
    if not current_schema:
        load_schema(schema_type)
        current_schema = fields_cache.get(schema_type, [])
        
    idx = next((i for i, f in enumerate(current_schema) if f.get('name') == payload.original_name), -1)
    
    # If not found, maybe append? For now, error
    if idx == -1:
        # Check if it's a new field addition? 
        raise HTTPException(status_code=404, detail=f"Field not found: {payload.original_name}")
        
    updated = dict(current_schema[idx])
    if payload.name is not None:
        updated['name'] = payload.name.strip()
    if payload.label is not None:
        updated['label'] = payload.label.strip()
    if payload.data_type is not None:
        updated['data_type'] = payload.data_type
        if payload.data_type != 'dropdown':
            updated['options'] = []
    if payload.options is not None:
        updated['options'] = [str(o).strip() for o in payload.options]
        
    current_schema[idx] = updated
    fields_cache[schema_type] = current_schema
    save_field_schema_to_disk(current_schema, schema_type)
    
    # Sync to Sheet Headers? 
    # For now, we only update local definition. 
    # To sync rename to sheet, we'd need more complex logic. 
    # Assuming just schema definition update for now.


@app.post("/search")
def search_data(payload: dict = Body(...)):
    date_filter = (payload.get("date") or "").strip().lower()
    name_filter = (payload.get("name") or "").strip().lower()
    member_filter = (payload.get("memberId") or "").strip().lower()

    try:
        # Re-initialize client
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        sheet = None
        error_log = []
        
        # PRIORITY 1: Try configured sheet ID first (most reliable)
        if GOOGLE_SHEET_ID:
            try:
                sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Sheet1")
                print(f"[Search] Opened sheet by ID: {GOOGLE_SHEET_ID}")
            except Exception as e:
                error_log.append(f"ByID({GOOGLE_SHEET_ID}): {str(e)}")
        
        # PRIORITY 2: Try common sheet names
        if not sheet:
            for sheet_name in ["CRM Leads", "CRM_Leads", "Lead CRM"]:
                try:
                    sheet = client.open(sheet_name).worksheet("Sheet1")
                    print(f"[Search] Opened sheet by name: {sheet_name}")
                    break
                except Exception as e:
                    error_log.append(f"{sheet_name}: {str(e)}")

        if not sheet:
             # Final attempt: try list all spreadsheets to see what's available (debugging)
             # But for now just fail with log
             raise Exception(f"Could not open sheet. Attempts: {'; '.join(error_log)}")

        # FIX: get_all_records() fails if headers are duplicate. Use get_all_values() instead.
        all_values = sheet.get_all_values()
        
        if not all_values:
            return {"status": "success", "data": []}
            
        raw_headers = all_values[0]
        # Create unique headers to handle duplicates (e.g. "Pain Point", "Pain Point")
        headers = []
        counts = {}
        for h in raw_headers:
            h_str = str(h).strip()
            if h_str in counts:
                counts[h_str] += 1
                headers.append(f"{h_str}_{counts[h_str]}")
            else:
                counts[h_str] = 1
                headers.append(h_str)
        
        # Parse rows
        rows = []
        for r_vals in all_values[1:]:
            # Ensure row length matches headers
            if len(r_vals) < len(headers):
                r_vals += [""] * (len(headers) - len(r_vals))
            elif len(r_vals) > len(headers):
                r_vals = r_vals[:len(headers)]
            
            row_dict = dict(zip(headers, r_vals))
            rows.append(row_dict)

        def match(row):
            # Safe read using get_value for potential duplicates (backward compatibility)
            d = str(get_value(row, "Date", "Date_2") or "").lower()
            
            # Check various name fields
            n = str(
                get_value(row, "Patient Name", "Name", "Full Name", "Patient Name_2", "Name_2", "Full Name_2")
            ).lower()
            
            # Check Member ID
            mid = str(get_value(row, "Member ID Key", "Member ID", "MemberID", "Member ID Key_2") or "").lower()

            return (
                (date_filter in d if date_filter else True)
                and (name_filter in n if name_filter else True)
                and (member_filter in mid if member_filter else True)
            )

        filtered = [r for r in rows if match(r)]

        # Limit results if too large? The user requested "ALL rows" for empty filters.
        # But for huge sheets this might be slow. The previous logic was "limit=50" for the old GET API,
        # but the new requirement said "If filters are empty -> return ALL rows".
        # So I will return valid filtered list.
        # However, to avoid breaking frontend with massive JSON, maybe I should check size?
        # User said "View Full Sheet button... load all records". I will return all.

        return {"status": "success", "data": filtered}

    except Exception as e:
        print("SEARCH ERROR DETAILED:", str(e))
        return {"status": "error", "message": f"Search failed: {str(e)}"}

@app.get("/download_template")
async def download_template(format: str = Query(default="xlsx", regex="^(csv|xlsx)$")):
    """Generate and return a sample template (CSV or Excel) based on active Sheet1 headers."""
    headers = []
    
    # Priority 1: Fetch Live Headers from Google Sheet
    try:
        if GOOGLE_SHEET_ID or os.path.exists(CREDENTIALS_FILE):
             client, spreadsheet = get_google_sheet_client()
             sheet = spreadsheet.sheet1
             # Get first row as headers
             live_headers = sheet.row_values(1)
             if live_headers:
                 headers = live_headers
    except Exception as e:
        print(f"Template generation: Failed to fetch live headers ({e}). Using fallback schema.")

    # Priority 2: Fallback to Local Schema
    if not headers:
        current_schema = fields_cache.get("enquiry", [])
        if not current_schema:
            load_schema("enquiry")
            current_schema = fields_cache.get("enquiry", [])
        headers = [f['name'] for f in current_schema]
    
    # Prepare sample data
    sample_data = [
        {
            'date': '03/12/2025',
            'memberidkey': 'MID-2025-12-03-001',
            'attender name': 'John Smith',
            'patient name': 'John Patient',
            'gender': 'Male',
            'age': '45',
            'patient location': 'New York',
            'area': 'Manhattan',
            'email id': 'john@example.com',
            'mobile number': '1234567890'
        },
        {
            'date': '04/12/2025',
            'memberidkey': 'MID-2025-12-04-002',
            'attender name': 'Jane Doe',
            'patient name': 'Jane Patient',
            'gender': 'Female',
            'age': '32',
            'patient location': 'London',
            'area': 'Westminster',
            'email id': 'jane@example.com',
            'mobile number': '0987654321'
        },
        {
            'date': '05/12/2025',
            'memberidkey': 'MID-2025-12-05-003',
            'attender name': 'Robert Johnson',
            'patient name': 'Robert Patient',
            'gender': 'Male',
            'age': '58',
            'patient location': 'Paris',
            'area': 'Marais',
            'email id': 'robert@example.com',
            'mobile number': '1122334455'
        }
    ]
    
    # Build rows
    rows_data = []
    current_schema = fields_cache.get("enquiry", [])
    schema_map = {f['name'].lower().strip(): f for f in current_schema} if current_schema else {}
    
    for sample in sample_data:
        example_row = []
        
        for h in headers:
            h_lower = str(h).lower().strip()
            
            # Check if we have sample data for this field
            if h_lower in sample:
                example_row.append(sample[h_lower])
            else:
                # Fallback to smart defaults
                field_def = schema_map.get(h_lower)
                dtype = field_def.get('data_type', 'text') if field_def else 'text'
                
                if 'date' in h_lower or dtype == 'date':
                    example_row.append(datetime.now().strftime('%d/%m/%Y'))
                elif 'email' in h_lower or dtype == 'email':
                    example_row.append('example@domain.com')
                elif 'phone' in h_lower or 'mobile' in h_lower or dtype == 'phone':
                    example_row.append('9876543210')
                elif dtype == 'number' or 'price' in h_lower or 'rent' in h_lower or 'age' in h_lower:
                    example_row.append('100')
                elif dtype == 'dropdown' and field_def and field_def.get('options'):
                    example_row.append(field_def['options'][0] if field_def['options'] else '')
                else:
                    example_row.append('')
                    
        rows_data.append(example_row)
    
    # Generate file based on format
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows_data:
            writer.writerow(row)
        output.seek(0)
        
        response = StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv"
        )
        response.headers["Content-Disposition"] = "attachment; filename=patient_data_template.csv"
        return response
    else:  # xlsx
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patient Data"
        
        # Write headers
        ws.append(headers)
        
        # Write data rows
        for row in rows_data:
            ws.append(row)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = "attachment; filename=patient_data_template.xlsx"
        return response


@app.post("/admission/register")
async def register_admission(payload: Dict[str, Any] = Body(...)):
    """
    Handle Admission Registration.
    1. Normalize fields to specific user-defined schema keys (patientblood, etc).
    2. Merge 'First Name'/'Patient Name' -> 'patient_name'.
    3. Save to Sheet1 (Master) and Patient Admission sheet.
    """
    print(f"[Admission Debug] Received payload keys: {list(payload.keys())}")
    
    # NORMALIZE RAW DATA
    raw_input = payload.get('data', {})
    raw_data = normalize_payload(raw_input)
    
    print(f"[Admission Debug] Raw data size: {len(raw_data)}")
    
    if not raw_data:
        print("[Admission Error] No data provided in payload")
        raise HTTPException(status_code=400, detail="No data provided")

    # --- 1. Strict Mapping Setup ---
    # Map normalized input keys to User's Preferred Schema Keys
    key_map = {
        # Personal
        "firstname": "patient_name",
        "patientname": "patient_name", # Merge target
        "lastname": "patientlastname",
        "patientlastname": "patientlastname",
        "gender": "gender", # Schema has 'gender'
        "bloodgroup": "patientblood",
        "blood": "patientblood",
        "maritalstatus": "patientmarital",
        
        # Emergency / Relational
        "emergencyname": "relationalname",
        "emergencyrelation": "relationalrelationship",
        "emergencyphone": "relationalmobile",
        "emergencyaltcontact": "relationalmobilealternative",
        
        # Medical
        "conditions": "patientcurrentstatus",
        "medications": "patientmedicalhistory",
        "allergies": "patientallergy",
        "surgeries": "patientsugarlevel", # Frontend mapping confirmed
        
        # Contact
        "phone": "mobilenumber",
        "mobile": "mobilenumber",
        "mobilenumber": "mobilenumber",
        "email": "emailid",
        "emailid": "emailid",
        
        # Address
        "door_num": "doornumber",
        "doornum": "doornumber",
        "street": "street",
        "city": "city",
        "district": "district",
        "state": "state",
        "pincode": "pincode",
        
        # Service / Charges
        "admissiondays": "no.ofdays",
        "noofdays": "no.ofdays",
        "miscellaneouscharges": "miscellaneous",
        "miscellaneous": "miscellaneous",
        
        # IDs
        "memberid": "MemberidKey",
        "memberidkey": "MemberidKey",
        "id": "MemberidKey",
        "memberidkey": "MemberidKey"
    }

    normalized_data = {}
    
    # Pre-scan for name fields to merge
    first_name_val = ""
    patient_name_val = ""
    
    # Process all inputs
    for k, v in raw_data.items():
        val_str = str(v).strip() if v is not None else ""
        raw_norm = normalize_field_name(k)
        
        # Name Logic
        if raw_norm == "firstname":
            first_name_val = val_str
        elif raw_norm == "patientname":
            patient_name_val = val_str
            
        # Target Key Lookup
        target_key = key_map.get(raw_norm, raw_norm) # Default to normalized self
        
        # Store
        if val_str:
            normalized_data[target_key] = val_str

    # Name Merge Execution
    final_name = patient_name_val if patient_name_val else first_name_val
    normalized_data["patient_name"] = final_name
    
    # Ensure MemberidKey aliases
    if "MemberidKey" in normalized_data:
        normalized_data["memberid"] = normalized_data["MemberidKey"]
        normalized_data["memberidkey"] = normalized_data["MemberidKey"] # User requested 'memberidkey'
        
    # Ensure Member ID exists (Fallbacks)
    if "MemberidKey" not in normalized_data:
        # Try to find it in raw_data under any casing
        for k, v in raw_data.items():
            if normalize_field_name(k) in ["memberid", "memberidkey", "id", "memberidkey", "mid"]:
                val = str(v).strip()
                normalized_data["MemberidKey"] = val
                normalized_data["memberid"] = val
                normalized_data["memberidkey"] = val
                break

    print(f"[Admission Debug] Normalized data to save: {normalized_data}")

    try:
        # --- 2. Dual Write ---
        
        # 1. Upsert to Master Sheet (Sheet1)
        print(f"[Admission Debug] Upserting to {GOOGLE_SHEET_NAME}...")
        res1 = upsert_to_sheet(GOOGLE_SHEET_NAME, normalized_data, "admission", strict_mode=True) 
        print(f"[Admission Debug] Master sheet result: {res1.get('status')}")
        
        # 2. Upsert to Patient Admission Sheet
        print(f"[Admission Debug] Upserting to Secondary Sheet...")
        res2 = save_patient_admission_to_sheet(normalized_data)
        print(f"[Admission Debug] Admission sheet result: {res2.get('status')}")
        
        # 3. Notification Logic
        p_name = normalized_data.get("patient_name", "Patient")
        m_id = normalized_data.get("MemberidKey", "Unknown")
        adm_date = normalized_data.get("admissiondate", datetime.now().strftime('%Y-%m-%d'))
        p_email = normalized_data.get("emailid") # Mapped key
        
        if p_email and '@' in p_email:
            subject = "Admission Confirmation"
            message_body = f"""Dear {p_name},

Your admission on {adm_date} has been successfully recorded.
Your Patient ID is {m_id}.

If you need assistance, please contact us.
"""
            try:
                msg = EmailMessage()
                msg.set_content(message_body)
                msg["Subject"] = subject
                msg["From"] = ensure_notification_defaults().get('sender_email', 'noreply@crm.com')
                msg["To"] = p_email
                with smtplib.SMTP(os.getenv("SMTP_SERVER", "smtp.gmail.com"), int(os.getenv("SMTP_PORT", 587))) as server:
                    server.starttls()
                    server.login(os.getenv("SMTP_USERNAME"), os.getenv("SMTP_PASSWORD"))
                    server.send_message(msg)
            except:
                pass

        return {
            "status": "success",
            "message": "Admission registered successfully",
            "sheet_url": res2.get('sheet_url')
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[Admission Error] Registration failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Cleanup orphan block

# --- NEW: Secondary Google Sheet Endpoint ---

@app.post("/patient-admission/save")
async def save_patient_admission(payload: Dict[str, Any] = Body(...)):
    """
    Save Patient Admission data to the SECONDARY Google Sheet.
    Supports both single object (legacy) and "rows" array (multi-patient).
    Wrapper with FULL error handling.
    """
    try:
        rows_to_process = []
        
        # Check for multi-row payload
        if "rows" in payload and isinstance(payload["rows"], list):
            rows_to_process = payload["rows"]
        else:
            # Fallback to single object logic
            raw_data = payload.get('data', {})
            if not raw_data:
                raw_data = payload
            rows_to_process = [raw_data]

        print(f"[Patient Admission Save] Processing {len(rows_to_process)} rows...")
        
        results = []
        for i, row in enumerate(rows_to_process):
            # Basic validation/cleaning could happen here if needed
            print(f"[Patient Admission Save] Saving Row {i+1}/{len(rows_to_process)}")
            try:
                # 1. Save to Secondary (Admission) Sheet
                res = save_patient_admission_to_sheet(row)
                
                # 2. Sync to Primary (Lead) Sheet1 (Update details using Member ID Key)
                try:
                    res_lead = upsert_to_sheet("Sheet1", row, "enquiry", strict_mode=True)
                    print(f"[Patient Admission Save] Synced to Lead Sheet1: {res_lead.get('status')}")
                    # Merge lead result info? mostly just need to know it worked.
                    res['lead_sync'] = "success"
                except Exception as lead_err:
                    print(f"[Patient Admission Save] Failed to sync to Lead Sheet1: {lead_err}")
                    res['lead_sync'] = f"failed: {str(lead_err)}"

                results.append({"status": "success", "row_index": i, "details": res})
            except Exception as e:
                print(f"[Patient Admission Save] Error on Row {i}: {e}")
                results.append({"status": "error", "row_index": i, "error": str(e)})

        # Check if all failed
        if all(r['status'] == 'error' for r in results) and results:
             return {"status": "error", "message": "All rows failed to save", "results": results}

        return {
            "status": "success", 
            "message": f"Processed {len(rows_to_process)} rows.",
            "results": results
        }
        
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print("PATIENT ADMISSION ERROR:", str(e))
        print(f"Trackback: {tb}")
        # Return error as JSON, don't crash with 500
        return {"status": "error", "message": str(e)}

@app.get("/patient-admission/view")
async def view_patient_admissions():
    """
    Retrieve all rows from the Patient Admission sheet/worksheet.
    Returns JSON structure with status, total count, and data list.
    """
    try:
        client, spreadsheet = get_patient_admission_sheet_client()
        
        # Use same logic as save function to determine worksheet name
        if PATIENT_ADMISSION_SHEET_ID:
            sheet_name = "Sheet1"  # Separate sheet uses Sheet1
        else:
            sheet_name = "Patient Admission"  # Same sheet uses dedicated worksheet
            
        print(f"[View Admissions] Attempting to read from worksheet: {sheet_name}")
        
        try:
            sheet = spreadsheet.worksheet(sheet_name)
        except gspread.WorksheetNotFound:
            print(f"[View Admissions] Worksheet '{sheet_name}' not found, returning empty")
            return {"status": "success", "total": 0, "data": []}
            
        # Get all records
        records = sheet.get_all_records()
        print(f"[View Admissions] Found {len(records)} records")
        
        return {
            "status": "success", 
            "total": len(records), 
            "data": records
        }
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Error in view_patient_admissions: {tb}")
        raise HTTPException(status_code=500, detail=f"{str(e)} | Trackback: {tb}")



@app.post("/submit")
async def submit_form(form_data: FormData, background_tasks: BackgroundTasks):
    """Submit form data to Google Sheets"""
    try:
        print(f"[Submit] Received form submission with {len(form_data.data)} fields")
        # Ensure defaults: auto Member ID and today's date for empty date fields
        enriched = dict(form_data.data)

        # Auto Member ID if a matching field exists and is empty
        member_id_field = None
        # Use enquiry schema for this check
        current_schema = fields_cache.get("enquiry", [])
        for field in current_schema:
            name_lower = field['name'].lower()
            if ("member" in name_lower) and ("id" in name_lower or "key" in name_lower):
                member_id_field = field['name']
                break
        
        if member_id_field and not enriched.get(member_id_field):
            # Always set server-side to keep it immutable from client edits
            # Format: MID-yyyy-mm-dd-numericalid
            now = datetime.now()
            date_part = now.strftime('%Y-%m-%d')
            # Generate unique numerical ID using timestamp + random number
            numerical_id = int(now.timestamp() * 1000) % 100000 + random.randint(1000, 9999)
            generated = f"MID-{date_part}-{numerical_id}"
            enriched[member_id_field] = generated

        # Fill empty date fields with today's date (YYYY-MM-DD)
        today = datetime.now().strftime('%Y-%m-%d')
        for field in current_schema:
            if field.get('data_type') == 'date' or field.get('type') == 'date':
                name = field['name']
                name_lower = str(name).lower()
                if REMINDER_DATE_KEYWORD in name_lower or FOLLOW_DATE_KEYWORD in name_lower:
                    continue
                if not str(enriched.get(name, "")).strip():
                    enriched[name] = today

        # 1. Upsert to Master (Sheet1)
        res1 = upsert_to_sheet("Sheet1", enriched, "enquiry", strict_mode=True)
        
        # 2. Upsert to Enquiry Sheet
        res2 = upsert_to_sheet(ENQUIRIES_SHEET_NAME, enriched, "enquiry")

        recipient_email = extract_recipient_email(enriched)
        print(f"[Email] Recipient detected from payload: {recipient_email}")
        if recipient_email:
            background_tasks.add_task(send_notification_email, recipient_email, dict(enriched))

        return {
            "status": "success", 
            "message": "Enquiry submitted successfully", 
            "sheet_url": res1.get("sheet_url")
        }
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    excel_exists = os.path.exists(EXCEL_FILE_PATH)
    creds_exist = os.path.exists(CREDENTIALS_FILE)
    
    print(f"[Health] fields_cache type: {type(fields_cache)}, len: {len(fields_cache)}")
    return {
        "status": "healthy",
        "excel_file": excel_exists,
        "google_credentials": creds_exist,
        "fields_loaded": len(fields_cache)
    }


@app.get("/debug/latest")
async def debug_latest_rows():
    """Return the target spreadsheet URL, first worksheet title, and last 5 rows."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)

    spreadsheet = None
    if GOOGLE_SHEET_ID:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
    else:
        try:
            spreadsheet = client.open(GOOGLE_SHEET_NAME)
        except gspread.SpreadsheetNotFound:
            raise HTTPException(status_code=404, detail="Spreadsheet not found")

    sheet = spreadsheet.sheet1
    values = sheet.get_all_values()
    last_rows = values[-5:] if len(values) > 5 else values

    return {
        "sheet_url": spreadsheet.url,
        "sheet_title": spreadsheet.title,
        "worksheet_title": sheet.title,
        "rows_total": len(values),
        "last_rows": last_rows,
    }


def ensure_google_sheet(client: gspread.Client) -> gspread.Spreadsheet:
    if GOOGLE_SHEET_ID:
        return client.open_by_key(GOOGLE_SHEET_ID)
    try:
        return client.open(GOOGLE_SHEET_NAME)
    except gspread.SpreadsheetNotFound:
        return client.create(GOOGLE_SHEET_NAME)


def get_google_sheet_client():
    """Helper to get authenticated gspread client and spreadsheet."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")
    
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    return client, spreadsheet



def sync_dropdown_options_to_sheet(client: gspread.Client, spreadsheet: gspread.Spreadsheet, schema: List[Dict[str, Any]]):
    try:
        try:
            ws = spreadsheet.worksheet("DropdownOptions")
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title="DropdownOptions", rows=100, cols=2)
        rows = [["Field Name", "Options"]]
        for f in schema:
            if (f.get("data_type") == "dropdown") and f.get("options"):
                rows.append([f.get("name"), ", ".join([str(o) for o in f.get("options")])])
        ws.clear()
        if rows:
            ws.update("A1", rows)
    except Exception as e:
        print(f"Warning: failed syncing dropdown options: {e}")


class FieldItem(BaseModel):
    name: str
    label: Optional[str] = None
    data_type: str
    options: Optional[List[str]] = None


class UpdateFieldsPayload(BaseModel):
    fields: List[FieldItem]


@app.post("/update_fields")
async def update_fields(payload: UpdateFieldsPayload):
    global fields_cache
    allowed_types = {"text","number","date","email","phone","boolean","dropdown"}
    schema: List[Dict[str, Any]] = []
    for f in payload.fields:
        if f.data_type not in allowed_types:
            raise HTTPException(status_code=400, detail=f"Invalid data_type for field '{f.name}': {f.data_type}")
        schema.append({
            "name": f.name.strip(),
            "label": (f.label or f.name).strip(),
            "data_type": f.data_type,
            "options": [str(o).strip() for o in (f.options or [])]
        })

    # Persist schema JSON
    with open(FIELD_SCHEMA_FILE, "w", encoding="utf-8") as f:
        json.dump(schema, f, ensure_ascii=False, indent=2)
    fields_cache["enquiry"] = schema

    # Update local CSV header row
    try:
        header_row = [f["name"] for f in schema]
        rows: List[List[str]] = []
        if os.path.exists(CSV_FILE_PATH):
            with open(CSV_FILE_PATH, "r", encoding="utf-8-sig", newline="") as rf:
                reader = csv.reader(rf)
                existing = list(reader)
            rows = existing[1:] if len(existing) > 0 else []
        with open(CSV_FILE_PATH, "w", encoding="utf-8", newline="") as wf:
            writer = csv.writer(wf)
            writer.writerow(header_row)
            for r in rows:
                writer.writerow(r)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update local CSV: {e}")

    # Update Google Sheet header row + dropdowns
    try:
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        headers = [f["name"] for f in schema] + ["Timestamp"]
        
        # 1. Update Sheet1
        try:
            sheet1 = spreadsheet.sheet1
            sheet1.update('1:1', [headers], value_input_option='USER_ENTERED')
        except Exception as e:
            print(f"Warning: failed updating Sheet1 headers: {e}")

        # 2. Update Enquiries Sheet
        try:
            try:
                sheet_enq = spreadsheet.worksheet(ENQUIRIES_SHEET_NAME)
            except gspread.WorksheetNotFound:
                sheet_enq = spreadsheet.add_worksheet(title=ENQUIRIES_SHEET_NAME, rows=1000, cols=20)
            sheet_enq.update('1:1', [headers], value_input_option='USER_ENTERED')
        except Exception as e:
            print(f"Warning: failed updating Enquiries headers: {e}")

        sync_dropdown_options_to_sheet(client, spreadsheet, schema)
        return {"status": "success", "message": "Fields updated", "sheet_url": spreadsheet.url}
    except FileNotFoundError:
        return {"status": "success", "message": "Fields updated (local). Google credentials not found to sync."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to sync Google Sheet: {e}")


@app.get("/preview_data")
async def preview_data():
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    sheet = spreadsheet.sheet1
    values = sheet.get_all_values()
    sample = values[:6] if len(values) >= 6 else values
    return {"rows": sample, "sheet_url": spreadsheet.url}





@app.put("/update_record")
async def update_record(member_id: Optional[str] = None, payload: Optional[Dict[str, Any]] = Body(default=None)):
    """Update a record in Google Sheets by member ID.
    Accepts member_id either as a query parameter or inside the JSON body.
    JSON body can be either {"member_id": "...", "data": {...}} or {"data": {...}}.
    If member_id is not explicitly provided, it will be inferred from data keys containing member/id.
    """
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    sheet = spreadsheet.sheet1
    values = sheet.get_all_values()

    if len(values) < 2:  # No data rows
        raise HTTPException(status_code=404, detail="No records found to update")

    headers = values[0]
    data_rows = values[1:]

    # Find member ID column
    member_id_col = None
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in ['member', 'id', 'key']) and ('member' in header_lower or 'id' in header_lower):
            member_id_col = i
            break

    if member_id_col is None:
        raise HTTPException(status_code=400, detail="Member ID column not found")

    # Read payload safely
    body_member_id: Optional[str] = None
    body_data: Dict[str, Any] = {}
    if isinstance(payload, dict):
        # FastAPI will parse JSON into dict when type is Dict[str, Any]
        body_member_id = payload.get("member_id")
        body_data = payload.get("data") or {}
    # Final member id resolution order: query -> body.member_id -> infer from body.data
    resolved_member_id: Optional[str] = (member_id or body_member_id)
    if not resolved_member_id:
        # Try to infer from data keys
        for k, v in body_data.items():
            k_lower = str(k).strip().lower()
            if ("member" in k_lower) and ("id" in k_lower or "key" in k_lower):
                resolved_member_id = str(v).strip()
                break

    if not resolved_member_id:
        raise HTTPException(status_code=400, detail="member_id is required (query param or in JSON body)")

    # Find the row with matching member ID
    target_row_idx = None
    for i, row in enumerate(data_rows):
        if member_id_col < len(row):
            row_member_id = str(row[member_id_col]).strip()
            if row_member_id == resolved_member_id.strip():
                target_row_idx = i + 2  # +2 because gspread is 1-indexed and we skip header
                break

    if target_row_idx is None:
        raise HTTPException(status_code=404, detail=f"Record with Member ID '{member_id}' not found")

    # Build case-insensitive source map from data
    source_ci = {str(k).strip().lower(): v for k, v in (body_data or {}).items()}

    # Helper: find a key in source that matches header name, case-insensitive,
    # with a fallback for Member ID key variations
    def resolve_value(header_name: str) -> Any:
        h_lower = header_name.strip().lower()
        if h_lower in source_ci:
            return source_ci[h_lower]

        # Fuzzy match for member id/key variations
        if ("member" in h_lower) and ("id" in h_lower or "key" in h_lower):
            for k_lower, v in source_ci.items():
                if ("member" in k_lower) and ("id" in k_lower or "key" in k_lower):
                    return v
        return ""

    # Build updated row aligned to existing headers
    # Preserve existing values if not provided in payload
    original_row = data_rows[target_row_idx - 2] if (target_row_idx - 2) < len(data_rows) else []
    updated_row: List[Any] = []
    for col_idx, h in enumerate(headers):
        new_val = resolve_value(h)
        if (new_val is None) or (str(new_val).strip() == ""):
            # fallback to original cell value if available
            if col_idx < len(original_row):
                updated_row.append(original_row[col_idx])
            else:
                updated_row.append("")
        else:
            updated_row.append(new_val)

    # Update the timestamp if it exists
    try:
        ts_idx = [h.strip().lower() for h in headers].index('timestamp')
        updated_row[ts_idx] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    except ValueError:
        pass

    # Update the specific row
    sheet.update(f'{target_row_idx}:{target_row_idx}', [updated_row], value_input_option='USER_ENTERED')

    # Determine if lead status changed and get email to notify
    try:
        headers_lower = [h.strip().lower() for h in headers]
        status_idx = headers_lower.index('lead status') if 'lead status' in headers_lower else None
        email_idx = None
        for i, h in enumerate(headers_lower):
            if 'email' in h:
                email_idx = i
                break
        member_id_idx = None
        for i, h in enumerate(headers_lower):
            if ('member' in h) and ('id' in h or 'key' in h):
                member_id_idx = i
                break

        status_changed = False
        recipient_email = None
        member_id_value = None
        if status_idx is not None and (target_row_idx - 2) < len(data_rows):
            prev_status = data_rows[target_row_idx - 2][status_idx] if status_idx < len(data_rows[target_row_idx - 2]) else ''
            new_status = updated_row[status_idx] if status_idx < len(updated_row) else prev_status
            status_changed = (str(prev_status).strip() != str(new_status).strip())
        if email_idx is not None and email_idx < len(updated_row):
            recipient_email = str(updated_row[email_idx]).strip()
        if member_id_idx is not None and member_id_idx < len(updated_row):
            member_id_value = str(updated_row[member_id_idx]).strip()

        if status_changed and recipient_email:
            payload_map = {h: (updated_row[i] if i < len(updated_row) else '') for i, h in enumerate(headers)}
            send_notification_email(recipient_email, payload_map, subject_override='LEAD STATUS CHANGED')
    except Exception as _e:
        print(f"[UpdateRecord] Notification check failed: {_e}")

    return {
        "status": "success",
        "message": f"Record with Member ID '{resolved_member_id}' updated successfully",
    }


# Helper global for login cache
login_cache: Dict[str, Any] = {"users": {}, "last_fetched": 0}
LOGIN_CACHE_DURATION = 300  # 5 minutes

@app.post("/login")
async def login(payload: LoginRequest):
    """Validate credentials against Google Sheet worksheet 'login details'.
    Expected columns: User_name, Password (case-insensitive).
    Fallback: accepts admin/admin for development if Google Sheet access fails.
    """
    import time
    start_time = time.time()
    print(f"[Login] Request received for user: {payload.User_name}")

    in_user = str(payload.User_name or '').strip()
    in_pass = str(payload.Password or '').strip()
    
    if not in_user or not in_pass:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    # Development fallback credentials
    DEV_CREDENTIALS = {
        "admin": "admin",
        "user": "user123",
        "test": "test123"
    }
    
    # Check dev credentials first for quick access
    if in_user in DEV_CREDENTIALS and DEV_CREDENTIALS[in_user] == in_pass:
        print(f"[Login] Dev credentials accepted. Time: {time.time() - start_time:.2f}s")
        return {"status": "ok"}

    # Check cache first
    current_time = time.time()
    if login_cache["users"] and (current_time - login_cache["last_fetched"] < LOGIN_CACHE_DURATION):
        cached_pass = login_cache["users"].get(in_user)
        if cached_pass and cached_pass == in_pass:
            print(f"[Login] Cached credentials accepted. Time: {time.time() - start_time:.2f}s")
            return {"status": "ok"}
    
    # Try Google Sheet authentication
    if not os.path.exists(CREDENTIALS_FILE):
        print("[Login] Google credentials file not found, using dev fallback only")
        raise HTTPException(status_code=401, detail="Invalid username or password")

    try:
        print(f"[Login] Connecting to Google Sheets... Time: {time.time() - start_time:.2f}s")
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)

        # Case-insensitive lookup for worksheet named 'Login Details'
        print(f"[Login] Finding worksheet... Time: {time.time() - start_time:.2f}s")
        try:
            ws = None
            # Fast check first
            try:
                ws = spreadsheet.worksheet("Login Details")
            except gspread.WorksheetNotFound:
                # Slow check
                for w in spreadsheet.worksheets():
                    if str(w.title).strip().lower() == 'login details':
                        ws = w
                        break
            
            if ws is None:
                print("[Login] 'Login Details' worksheet not found, using dev fallback only")
                raise HTTPException(status_code=401, detail="Invalid username or password")
        except HTTPException:
            raise
        except Exception as e:
            print(f"[Login] Error accessing worksheets: {e}")
            raise HTTPException(status_code=401, detail="Invalid username or password")

        print(f"[Login] Fetching values... Time: {time.time() - start_time:.2f}s")
        values = ws.get_all_values() or []
        if len(values) < 2:
            print("[Login] No user data in Login Details worksheet")
            raise HTTPException(status_code=401, detail="Invalid username or password")

        headers = values[0]
        rows = values[1:]
        # map header names to indices (case-insensitive)
        header_map = {str(h).strip().lower(): i for i, h in enumerate(headers)}
        user_col = (
            header_map.get('user_name')
            or header_map.get('user name')
            or header_map.get('username')
            or header_map.get('user-name')
        )
        pass_col = header_map.get('password')
        if user_col is None or pass_col is None:
            print("[Login] Required columns not found in Login Details worksheet")
            raise HTTPException(status_code=401, detail="Invalid username or password")

        # Refresh cache
        new_cache = {}
        found = False
        for r in rows:
            u = str(r[user_col]).strip() if user_col < len(r) else ''
            p = str(r[pass_col]).strip() if pass_col < len(r) else ''
            if u:
                new_cache[u] = p
            if u == in_user and p == in_pass:
                found = True

        # Update global cache
        login_cache["users"] = new_cache
        login_cache["last_fetched"] = current_time
        
        if found:
            print(f"[Login] Google Sheet credentials accepted and cached. Time: {time.time() - start_time:.2f}s")
            return {"status": "ok"}

        print(f"[Login] Credential mismatch for user '{in_user}'. Time: {time.time() - start_time:.2f}s")
        raise HTTPException(status_code=401, detail="Invalid username or password")
    except HTTPException:
        raise
    except Exception as exc:
        print(f"[Login] Exception during Google Sheet auth: {exc}")
        # If google auth fails, fallback to checking cache one last time if we have old data? 
        # No, simpler to just fail safe.
        raise HTTPException(status_code=401, detail="Invalid username or password")


@app.get("/list_sheets")
async def list_sheets():
    """Return the list of worksheet names in the Google Spreadsheet and whether 'login details' exists."""
    # Trigger reload
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=500, detail="Google credentials file not found")
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    names = [ws.title for ws in spreadsheet.worksheets()]
    return {
        "sheets": names,
        "has_login_details": any(str(n).strip().lower() == 'login details' for n in names)
    }



# --- Bed Management API ---

class BedAllocationRequest(BaseModel):
    patient_name: str
    member_id: Optional[str] = None
    gender: str
    room_no: str
    bed_index: int
    room_type: str
    admission_date: str
    discharge_date: Optional[str] = None
    pain_point: Optional[str] = None

class ComplaintRequest(BaseModel):
    patient_name: str
    room_no: str
    complaint_type: str
    description: Optional[str] = None

class FeedbackRequest(BaseModel):
    patient_name: str
    rating_comfort: int
    rating_cleanliness: int
    rating_staff: int
    comments: Optional[str] = None

def ensure_bed_sheets_google(spreadsheet: gspread.Spreadsheet):
    """Ensure Admission Details, Complaints, and Feedback sheets exist in Google Sheet."""

    def seed_bed_sheet(ws):
        """Ensure headers and default bed rows exist."""
        headers = [
            "Room No", "Room Type", "Bed Count", "Bed Index",
            "Patient Name", "Member ID", "Gender",
            "Admission Date", "Discharge Date", "Status",
            "Pain Point", "Complaints"
        ]

        existing_rows = ws.get_all_values()

        if not existing_rows:
            ws.append_row(headers)
            existing_rows = [headers]
        else:
            existing_header = [h.strip().lower() for h in existing_rows[0]]
            expected_header = [h.strip().lower() for h in headers]
            if existing_header != expected_header:
                # Reset header row to expected format
                ws.update('A1:L1', [headers])
                existing_rows = [headers]

        # If there are no bed entries yet, seed with default rooms
        if len(existing_rows) <= 1:
            rows_to_add = []
            for i in range(101, 106):
                rows_to_add.append([str(i), "Single", 1, 0, "", "", "", "", "", "Available", "", ""])
            for i in range(201, 206):
                rows_to_add.append([str(i), "Twin", 2, 0, "", "", "", "", "", "Available", "", ""])
                rows_to_add.append([str(i), "Twin", 2, 1, "", "", "", "", "", "Available", "", ""])
            ws.append_rows(rows_to_add)

    # --- Admission Details ---
    try:
        ws_admission = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws_admission = spreadsheet.add_worksheet(title=ADMISSION_SHEET_NAME, rows=200, cols=20)
    seed_bed_sheet(ws_admission)

    # --- Complaints ---
    try:
        spreadsheet.worksheet(COMPLAINT_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws_complaints = spreadsheet.add_worksheet(title=COMPLAINT_SHEET_NAME, rows=100, cols=10)
        ws_complaints.append_row(["Date", "Room No", "Patient Name", "Type", "Description", "Resolved"])

    # --- Feedback ---
    try:
        spreadsheet.worksheet(FEEDBACK_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws_feedback = spreadsheet.add_worksheet(title=FEEDBACK_SHEET_NAME, rows=100, cols=10)
        ws_feedback.append_row(["Date", "Patient Name", "Comfort", "Cleanliness", "Staff", "Comments"])


@app.get("/api/beds")
async def get_beds():
    """Get all bed status from Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ensure_bed_sheets_google(spreadsheet)
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        rows = ws.get_all_values()
        if len(rows) < 2:
            return {"beds": []}
            
        headers = [str(h).strip().lower() for h in rows[0]]
        beds = []
        
        # Parse rows
        for idx, row in enumerate(rows[1:], start=2): # Start=2 to match sheet row number
            bed_data = {}
            for h_idx, h in enumerate(headers):
                if h_idx < len(row):
                    bed_data[h] = row[h_idx]
            
            clean_bed = {
                "room_no": bed_data.get("room no"),
                "room_type": bed_data.get("room type"),
                "bed_index": bed_data.get("bed index"),
                "patient_name": bed_data.get("patient name"),
                "member_id": bed_data.get("member id"),
                "gender": bed_data.get("gender"),
                "status": bed_data.get("status", "Available"),
                "admission_date": bed_data.get("admission date"),
                "discharge_date": bed_data.get("discharge date"),
                "pain_point": bed_data.get("pain point"),
                "row_idx": idx 
            }
            # Only add valid rows (skip empty ones if any)
            if clean_bed["room_no"]:
                beds.append(clean_bed)
                
        return {"beds": beds}
    except Exception as e:
        print(f"Error fetching beds: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/beds/allocate")
async def allocate_bed(payload: BedAllocationRequest):
    """Allocate a bed in Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        # Get all rows to find the target bed
        rows = ws.get_all_values()
        headers = [str(h).strip().lower() for h in rows[0]]
        
        # Find indices
        try:
            room_no_idx = headers.index("room no")
            bed_index_idx = headers.index("bed index")
            status_idx = headers.index("status")
            patient_idx = headers.index("patient name")
            mid_idx = headers.index("member id")
            gender_idx = headers.index("gender")
            adm_idx = headers.index("admission date")
            dis_idx = headers.index("discharge date")
            pain_idx = headers.index("pain point")
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"Sheet headers missing: {e}")

        target_row_num = -1
        
        # Search for the bed row (skip header)
        for i, row in enumerate(rows[1:], start=2):
            # Check range to avoid index errors
            if len(row) <= max(room_no_idx, bed_index_idx):
                continue
                
            r_no = str(row[room_no_idx]).strip()
            b_idx = str(row[bed_index_idx]).strip()
            
            if r_no == str(payload.room_no) and b_idx == str(payload.bed_index):
                target_row_num = i
                # Check status
                if len(row) > status_idx and row[status_idx] == "Occupied":
                    raise HTTPException(status_code=400, detail="Bed already occupied")
                break
        
        if target_row_num == -1:
             raise HTTPException(status_code=404, detail="Bed not found in system")

        # Update the row
        # gspread uses 1-based indexing for columns
        updates = [
            {"range": f"{chr(65+patient_idx)}{target_row_num}", "values": [[payload.patient_name]]},
            {"range": f"{chr(65+mid_idx)}{target_row_num}", "values": [[payload.member_id or ""]]},
            {"range": f"{chr(65+gender_idx)}{target_row_num}", "values": [[payload.gender]]},
            {"range": f"{chr(65+adm_idx)}{target_row_num}", "values": [[payload.admission_date]]},
            {"range": f"{chr(65+dis_idx)}{target_row_num}", "values": [[payload.discharge_date or ""]]},
            {"range": f"{chr(65+status_idx)}{target_row_num}", "values": [["Occupied"]]},
            {"range": f"{chr(65+pain_idx)}{target_row_num}", "values": [[payload.pain_point or ""]]}
        ]
        
        # Batch update is more efficient (or multiple update_cells)
        # For simplicity with gspread, we can just update the row range 
        # But constructing the full row might be tricky if we don't have it all.
        # Let's use update_acell or update matching the column index.
        # Note: headers.index returns 0-based. gspread col is i+1.
        
        ws.update_cell(target_row_num, patient_idx + 1, payload.patient_name)
        ws.update_cell(target_row_num, mid_idx + 1, payload.member_id or "")
        ws.update_cell(target_row_num, gender_idx + 1, payload.gender)
        ws.update_cell(target_row_num, adm_idx + 1, payload.admission_date)
        ws.update_cell(target_row_num, dis_idx + 1, payload.discharge_date or "")
        ws.update_cell(target_row_num, status_idx + 1, "Occupied")
        ws.update_cell(target_row_num, pain_idx + 1, payload.pain_point or "")
        
        return {"status": "success", "message": "Bed allocated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error allocating bed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/complaints")
async def log_complaint(payload: ComplaintRequest):
    """Log a complaint to Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        try:
            ws = spreadsheet.worksheet(COMPLAINT_SHEET_NAME)
        except gspread.WorksheetNotFound:
            # Fallback if ensure wasn't called (though get_beds calls it)
            ensure_bed_sheets_google(spreadsheet)
            ws = spreadsheet.worksheet(COMPLAINT_SHEET_NAME)
            
        ws.append_row([
            datetime.now().strftime('%Y-%m-%d'),
            payload.room_no,
            payload.patient_name,
            payload.complaint_type,
            payload.description,
            "No"
        ])
        
        return {"status": "success", "message": "Complaint logged"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/feedback")
async def submit_feedback(payload: FeedbackRequest):
    """Submit feedback to Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        try:
            ws = spreadsheet.worksheet(FEEDBACK_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ensure_bed_sheets_google(spreadsheet)
            ws = spreadsheet.worksheet(FEEDBACK_SHEET_NAME)
            
        ws.append_row([
            datetime.now().strftime('%Y-%m-%d'),
            payload.patient_name,
            payload.rating_comfort,
            payload.rating_cleanliness,
            payload.rating_staff,
            payload.comments
        ])
        
        return {"status": "success", "message": "Feedback submitted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/beds/discharge")
async def discharge_bed(room_no: str, bed_index: int):
    """Discharge a patient and release the bed. Patient data remains in main sheet."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        # Get all rows
        rows = ws.get_all_values()
        headers = [str(h).strip().lower() for h in rows[0]]
        
        # Find indices
        try:
            room_no_idx = headers.index("room no")
            bed_index_idx = headers.index("bed index")
            status_idx = headers.index("status")
            patient_idx = headers.index("patient name")
            mid_idx = headers.index("member id")
            gender_idx = headers.index("gender")
            adm_idx = headers.index("admission date")
            dis_idx = headers.index("discharge date")
            pain_idx = headers.index("pain point")
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"Sheet headers missing: {e}")

        target_row_num = -1
        
        # Find the bed row
        for i, row in enumerate(rows[1:], start=2):
            if len(row) <= max(room_no_idx, bed_index_idx):
                continue
                
            r_no = str(row[room_no_idx]).strip()
            b_idx = str(row[bed_index_idx]).strip()
            
            if r_no == str(room_no) and b_idx == str(bed_index):
                target_row_num = i
                # Check if occupied
                if len(row) > status_idx and row[status_idx] != "Occupied":
                    raise HTTPException(status_code=400, detail="Bed is not occupied")
                break
        
        if target_row_num == -1:
            raise HTTPException(status_code=404, detail="Bed not found")

        # Clear bed data using batch update for better performance
        # Convert column indices to A1 notation
        def col_to_letter(col_idx):
            """Convert 0-based column index to Excel column letter"""
            col_idx += 1  # Convert to 1-based
            result = ""
            while col_idx > 0:
                col_idx -= 1
                result = chr(65 + (col_idx % 26)) + result
                col_idx //= 26
            return result
        
        # Prepare batch update
        updates = [
            {
                'range': f'{col_to_letter(patient_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(mid_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(gender_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(adm_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(dis_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(status_idx)}{target_row_num}',
                'values': [['Available']]
            },
            {
                'range': f'{col_to_letter(pain_idx)}{target_row_num}',
                'values': [['']]
            }
        ]
        
        ws.batch_update(updates)
        
        return {"status": "success", "message": "Patient discharged successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error discharging bed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/beds/update-discharge")
async def update_discharge_date(room_no: str, bed_index: int, discharge_date: str):
    """Update discharge date for an occupied bed. Does NOT discharge the patient."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        # Get all rows
        rows = ws.get_all_values()
        headers = [str(h).strip().lower() for h in rows[0]]
        
        # Find indices
        try:
            room_no_idx = headers.index("room no")
            bed_index_idx = headers.index("bed index")
            status_idx = headers.index("status")
            dis_idx = headers.index("discharge date")
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"Sheet headers missing: {e}")

        target_row_num = -1
        
        # Find the bed row
        for i, row in enumerate(rows[1:], start=2):
            if len(row) <= max(room_no_idx, bed_index_idx):
                continue
                
            r_no = str(row[room_no_idx]).strip()
            b_idx = str(row[bed_index_idx]).strip()
            
            if r_no == str(room_no) and b_idx == str(bed_index):
                target_row_num = i
                # Check if occupied
                if len(row) > status_idx and row[status_idx] != "Occupied":
                    raise HTTPException(status_code=400, detail="Bed is not occupied")
                break
        
        if target_row_num == -1:
            raise HTTPException(status_code=404, detail="Bed not found")

        # ONLY update discharge date - do NOT change status or clear patient data
        ws.update_cell(target_row_num, dis_idx + 1, discharge_date or "")
        
        return {"status": "success", "message": "Discharge date updated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating discharge date: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class AIChatRequest(BaseModel):
    query: str
    filter: Optional[str] = None


# AI Configuration
AI_PROVIDER = os.getenv('AI_PROVIDER', 'groq')
AI_ENABLED = os.getenv('AI_ENABLED', 'True').lower() == 'true'

# Groq Configuration (Fast & Reliable)
GROQ_API_KEY = os.getenv('GROQ_API_KEY', '')
GROQ_MODEL = os.getenv('GROQ_MODEL', 'llama-3.1-70b-versatile')
GROQ_API_BASE_URL = 'https://api.groq.com/openai/v1'

# Hugging Face Configuration (Backup)
HF_TOKEN = os.getenv('HF_TOKEN', '')
HF_MODEL = os.getenv('HF_MODEL', 'mistralai/Mistral-7B-Instruct-v0.2')
HF_API_BASE_URL = os.getenv('HF_API_BASE_URL', 'https://router.huggingface.co/v1')
HF_ENABLED = os.getenv('HF_ENABLED', 'False').lower() == 'true'


async def query_groq_ai(user_query: str, crm_data_summary: str, member_ids: List[str]) -> Optional[str]:
    """
    Query Groq API with CRM context.
    
    Args:
        user_query: The user's question
        crm_data_summary: Summary of relevant CRM data
        member_ids: List of relevant member IDs
        
    Returns:
        AI-generated response or None if API fails
    """
    if not GROQ_API_KEY:
        return None
    
    try:
        # Build the prompt with CRM context
        system_prompt = """You are an AI CRM assistant for an elderly care service. 
You help staff quickly understand customer data and follow-up requirements.
Be concise, professional, and helpful.

IMPORTANT: Member names are shown as "AttenderName/PatientName" when both exist. Search BOTH names when looking for a person.

When users ask for "names", provide the patient/customer names.
When users ask for "member IDs" or "IDs", provide the member IDs.
When users ask about locations (e.g., city, area, where), provide the location for each relevant member.
When users ask for phone numbers (e.g., phone, mobile, contact number), provide the phone number for each relevant member.
Otherwise, provide both name and ID in format: "Name (ID)" and include location and phone if it adds clarity."""
        
        user_prompt = f"""CRM Data Context:
{crm_data_summary}

User Question: {user_query}

Provide a clear, concise answer based on the CRM data above.
If the user asks for names, list the names from the member details.
If the user asks for IDs, list the member IDs.
If the user asks about location, provide the member locations alongside their names or IDs.
If the user asks for phone numbers, provide the phone/mobile number alongside the name or ID.
Otherwise, provide both name and ID, including location and phone if helpful."""
        
        # Prepare the API request
        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": GROQ_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "max_tokens": 500,
            "temperature": 0.7,
            "top_p": 0.9
        }
        
        # Make async request to Groq
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{GROQ_API_BASE_URL}/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code == 200:
                result = response.json()
                ai_response = result.get('choices', [{}])[0].get('message', {}).get('content', '')
                if ai_response:
                    print(f"✅ Groq AI Response received: {ai_response[:100]}...")
                    return ai_response.strip()
                else:
                    print("⚠️ Groq returned empty response")
                    return None
            else:
                print(f"❌ Groq API error: {response.status_code} - {response.text}")
                return None
                
    except Exception as e:
        print(f"Error calling Groq API: {e}")
        return None


async def query_huggingface_ai(user_query: str, crm_data_summary: str, member_ids: List[str]) -> Optional[str]:
    """
    Query Hugging Face Inference API with CRM context.
    
    Args:
        user_query: The user's question
        crm_data_summary: Summary of relevant CRM data
        member_ids: List of relevant member IDs
        
    Returns:
        AI-generated response or None if API fails
    """
    if not HF_TOKEN or not HF_ENABLED:
        return None
    
    try:
        # Build the prompt with CRM context
        system_prompt = """You are an AI CRM assistant for an elderly care service. 
You help staff quickly understand customer data and follow-up requirements.
Be concise, professional, and helpful.

IMPORTANT: Member names are shown as "AttenderName/PatientName" when both exist. Search BOTH names when looking for a person.

When users ask for "names", provide the patient/customer names.
When users ask for "member IDs" or "IDs", provide the member IDs.
When users ask about locations (e.g., city, area, where), provide the location for each relevant member.
When users ask for phone numbers (e.g., phone, mobile, contact number), provide the phone number for each relevant member.
Otherwise, provide both name and ID in format: "Name (ID)" and include location and phone if it adds clarity."""
        
        user_prompt = f"""CRM Data Context:
{crm_data_summary}

User Question: {user_query}

Provide a clear, concise answer based on the CRM data above.
If the user asks for names, list the names from the member details.
If the user asks for IDs, list the member IDs.
If the user asks about location, provide the member locations alongside their names or IDs.
If the user asks for phone numbers, provide the phone/mobile number alongside the name or ID.
Otherwise, provide both name and ID, including location and phone if helpful."""
        
        # Prepare the API request
        headers = {
            "Authorization": f"Bearer {HF_TOKEN}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": HF_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "max_tokens": 500,
            "temperature": 0.7,
            "top_p": 0.9
        }
        
        # Make async request to Hugging Face (OpenAI-compatible format)
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{HF_API_BASE_URL}/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code == 200:
                result = response.json()
                # OpenAI-compatible format
                ai_response = result.get('choices', [{}])[0].get('message', {}).get('content', '')
                if ai_response:
                    print(f"✅ AI Response received: {ai_response[:100]}...")
                    return ai_response.strip()
                else:
                    print("⚠️ AI returned empty response")
                    return None
            elif response.status_code == 503:
                # Model is loading
                print(f"⚠️ Model is loading, using fallback...")
                return None
            else:
                print(f"❌ Hugging Face API error: {response.status_code} - {response.text}")
                return None
                
    except Exception as e:
        print(f"Error calling Hugging Face API: {e}")
        return None


def generate_fallback_response(query_lower: str, filter_type: str, filtered_rows: List, 
                              member_ids: List[str], status_counts: Dict = None, 
                              member_data: List[Dict] = None) -> str:
    """
    Generate rule-based response as fallback when AI is unavailable.
    
    Args:
        query_lower: Lowercase user query
        filter_type: Applied filter (today, this_week, overdue)
        filtered_rows: Filtered CRM data rows
        member_ids: List of member IDs
        status_counts: Dictionary of status counts
        member_data: List of {id, name, location} dictionaries
        
    Returns:
        Formatted response string
    """
    # Check if user is asking for names, locations, or phone numbers
    name_keywords = ['name', 'who', 'patient', 'customer']
    location_keywords = ['location', 'where', 'city', 'area', 'place', 'address', 'region']
    phone_keywords = ['phone', 'mobile', 'contact', 'number', 'contact number', 'mobile number']
    asking_for_names = any(keyword in query_lower for keyword in name_keywords)
    asking_for_location = any(keyword in query_lower for keyword in location_keywords)
    asking_for_phone = any(keyword in query_lower for keyword in phone_keywords)

    # Format member info based on what user is asking
    def format_members(limit=10):
        if not member_data:
            return ', '.join(member_ids[:limit])

        formatted = []
        for item in member_data[:limit]:
            name = item.get('name', 'Unknown')
            mid = item.get('id', '')
            location = item.get('location')
            phone = item.get('phone')
            has_location = location and location.lower() != 'unknown location'
            has_phone = bool(str(phone).strip())

            # Determine display based on query intent
            if asking_for_phone and not asking_for_location and not asking_for_names:
                # Only phone requested
                if has_phone and name != 'Unknown':
                    formatted.append(f"{name} - {phone}")
                elif has_phone:
                    formatted.append(f"{mid} - {phone}")
                else:
                    formatted.append(f"{name if name != 'Unknown' else mid} - phone not available")
            elif asking_for_location and not asking_for_names:
                if has_location and name != 'Unknown':
                    formatted.append(f"{name} - {location}")
                elif has_location:
                    formatted.append(f"{mid} - {location}")
                else:
                    formatted.append(f"{name if name != 'Unknown' else mid} - location not available")
            elif asking_for_names and not asking_for_location:
                if name != 'Unknown':
                    if asking_for_phone and has_phone and has_location:
                        formatted.append(f"{name} ({location} · {phone})")
                    elif asking_for_phone and has_phone:
                        formatted.append(f"{name} ({phone})")
                    elif has_location:
                        formatted.append(f"{name} ({location})")
                    else:
                        formatted.append(name)
                else:
                    if asking_for_phone and has_phone and has_location:
                        formatted.append(f"{mid} ({location} · {phone})")
                    elif asking_for_phone and has_phone:
                        formatted.append(f"{mid} ({phone})")
                    elif has_location:
                        formatted.append(f"{mid} ({location})")
                    else:
                        formatted.append(mid)
            else:
                base = f"{mid} ({name})"
                if has_location:
                    base += f" - {location}"
                if asking_for_phone and has_phone:
                    base += f" - {phone}"
                formatted.append(base)

        if not formatted:
            # Fallback if member_data was empty or formatting failed
            base_list = member_names if asking_for_names else member_ids
            return ', '.join(base_list[:limit])

        return ', '.join(formatted)

    if 'follow' in query_lower or 'followup' in query_lower or 'follow-up' in query_lower:
        members_str = format_members(10)
        if filter_type == 'today':
            return f"Found {len(filtered_rows)} member(s) requiring follow-up today: {members_str}"
        elif filter_type == 'this_week' or filter_type == 'this week':
            return f"Found {len(filtered_rows)} member(s) requiring follow-up this week: {members_str}"
        elif filter_type == 'overdue':
            return f"Found {len(filtered_rows)} overdue follow-up(s): {members_str}"
        else:
            return f"Found {len(filtered_rows)} member(s) with follow-ups: {members_str}"
    
    elif 'status' in query_lower and status_counts:
        status_summary = ', '.join([f"{k}: {v}" for k, v in status_counts.items()])
        members_str = format_members(10)
        return f"Lead status summary: {status_summary}. Members: {members_str}"
    
    elif 'count' in query_lower or 'how many' in query_lower:
        response = f"Found {len(filtered_rows)} member(s) matching your criteria."
        if member_data or member_ids:
            members_str = format_members(10)
            response += f" Members: {members_str}"
        return response
    
    else:
        if len(filtered_rows) == 0:
            return "No members found matching your criteria."
        else:
            members_str = format_members(10)
            response = f"Found {len(filtered_rows)} member(s): {members_str}"
            if len(member_ids) > 10:
                response += f" and {len(member_ids) - 10} more..."
            return response


# AI CRM Chat Models
class AIChatRequest(BaseModel):
    query: str
    filter: Optional[str] = None

def get_all_google_sheets_data():
    """Get ALL data from Google Sheets for AI Analytics - no field filtering"""
    if not os.path.exists(CREDENTIALS_FILE):
        raise Exception("Google credentials file not found")
    
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    sheet = spreadsheet.sheet1
    
    # Get ALL values from the sheet
    values = sheet.get_all_values()
    
    if len(values) < 2:
        return {"headers": [], "data": [], "total_rows": 0}
    
        return "No data available"
    
    # Take a sample of rows for AI Analytics
    sample_rows = data_rows[:max_rows]
    
    # Create a structured text representation
    lines = []
    lines.append(f"CRM Data Summary ({len(sample_rows)} of {len(data_rows)} total records):")
    lines.append("=" * 50)
    lines.append("")
    
    # Add headers
    lines.append("Available Columns:")
    for i, header in enumerate(headers):
        lines.append(f"{i+1}. {header}")
    lines.append("")
    
    # Add sample data in a readable format
    lines.append("Sample Records:")
    for row_idx, row in enumerate(sample_rows):
        lines.append(f"\nRecord {row_idx + 1}:")
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and value.strip():  # Only show non-empty values
                lines.append(f"  {headers[col_idx]}: {value}")
        
        # Only show first 10 records in detail to avoid overwhelming the AI
        if row_idx >= 9:
            lines.append(f"\n... and {len(sample_rows) - 10} more records")
            break
    
    return "\n".join(lines)

@app.post("/api/ai-crm/chat")
async def ai_crm_chat(request: AIChatRequest, background_tasks: BackgroundTasks):
    """
    AI CRM Chat endpoint - queries Google Sheets data based on user input.
    Supports filters: today, this_week, overdue
    """
    try:
        if not os.path.exists(CREDENTIALS_FILE):
            return {
                "response": "Google Sheets is not connected. Please configure credentials.",
                "member_ids": [],
                "connected": False
            }
        
        # Connect to Google Sheets
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        
        # Try to find the correct worksheet
        try:
            worksheet = spreadsheet.worksheet(GOOGLE_SHEET_NAME)
        except gspread.exceptions.WorksheetNotFound:
            # Try alternative names
            worksheets = spreadsheet.worksheets()
            worksheet_names = [ws.title for ws in worksheets]
            
            # Try to find a worksheet with "CRM" or "Lead" in the name
            crm_sheet = None
            for ws in worksheets:
                ws_name_lower = ws.title.lower()
                if 'crm' in ws_name_lower or 'lead' in ws_name_lower:
                    crm_sheet = ws
                    break
            
            if crm_sheet:
                worksheet = crm_sheet
            else:
                # Use the first worksheet if no CRM sheet found
                worksheet = worksheets[0] if worksheets else None
                
            if not worksheet:
                return {
                    "response": f"Could not find CRM worksheet. Available sheets: {', '.join(worksheet_names)}",
                    "member_ids": [],
                    "connected": True
                }
        
        # Get all data
        all_rows = worksheet.get_all_values()
        if len(all_rows) < 2:
            return {
                "response": "No data found in the CRM sheet.",
                "member_ids": [],
                "connected": True
            }
        
        original_headers = [str(h).strip() for h in all_rows[0]]
        headers = [h.lower() for h in original_headers]
        data_rows = all_rows[1:]
        
        # Find relevant column indices
        member_id_col = next((i for i, h in enumerate(headers) if 'member' in h and 'id' in h), None)
        date_col = next((i for i, h in enumerate(headers) if h == 'date'), None)
        follow1_col = next((i for i, h in enumerate(headers) if 'follow1 date' in h or 'follow_1 date' in h), None)
        follow2_col = next((i for i, h in enumerate(headers) if 'follow_2 date' in h or 'follow2 date' in h), None)
        follow3_col = next((i for i, h in enumerate(headers) if 'follow_3 date' in h or 'follow3 date' in h), None)
        lead_status_col = next((i for i, h in enumerate(headers) if 'lead status' in h), None)
        patient_name_col = next((i for i, h in enumerate(headers) if 'patient name' in h), None)
        attender_name_col = next((i for i, h in enumerate(headers) if 'attender name' in h), None)
        patient_location_col = next((i for i, h in enumerate(headers) if 'patient location' in h), None)
        location_col = next((i for i, h in enumerate(headers) if h == 'location'), None)
        area_col = next((i for i, h in enumerate(headers) if 'area' in h), None)
        # Phone/Mobile/Contact number column
        email_col = next((i for i, h in enumerate(headers) if 'email' in h), None)
        mobile_col = next((i for i, h in enumerate(headers) if any(k in h for k in ['mobile', 'phone', 'contact'])), None)
        
        # Get today's date
        from datetime import datetime, timedelta
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        
        def parse_date(date_str):
            """Parse date string in various formats"""
            if not date_str or str(date_str).strip() == '':
                return None
            try:
                # Try DD/MM/YYYY format
                if '/' in str(date_str):
                    parts = str(date_str).split('/')
                    if len(parts) == 3:
                        return datetime.strptime(date_str, '%d/%m/%Y').date()
                # Try YYYY-MM-DD format
                elif '-' in str(date_str):
                    return datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                pass
            return None
        
        def smart_filter_rows(all_rows, query_text, headers_list):
            """
            Smart retrieval: Scan ALL rows and filter based on query keywords.
            For field-specific queries (age, gender, email, etc.), return ALL rows.
            """
            query_lower = query_text.lower()
            
            # Field-specific query keywords - return ALL rows for these
            field_keywords = ['age', 'gender', 'email', 'phone', 'mobile', 'service', 'status', 
                            'location', 'area', 'source', 'pain', 'date', 'assigned', 'agent']
            
            # If query is asking about a field value, return ALL rows
            if any(keyword in query_lower for keyword in field_keywords):
                return all_rows
            
            relevant_rows = []
            
            # Extract potential search terms from query
            search_terms = []
            # Remove common words
            stop_words = {'the', 'a', 'an', 'is', 'are', 'was', 'were', 'of', 'for', 'in', 'on', 'at', 'to', 'from', 'by', 'with', 'what', 'who', 'where', 'when', 'how', 'show', 'give', 'get', 'find', 'tell', 'me', 'my', 'i', 'you', 'your', 'member', 'members', 'patient', 'patients'}
            words = query_lower.split()
            search_terms = [w for w in words if w not in stop_words and len(w) > 2]
            
            # Scan ALL rows
            for row in all_rows:
                if len(row) == 0:
                    continue
                
                # If no specific search terms, include all rows
                if not search_terms:
                    relevant_rows.append(row)
                    continue
                
                # Check if any search term matches any cell in the row
                row_text = ' '.join([str(cell).lower() for cell in row])
                if any(term in row_text for term in search_terms):
                    relevant_rows.append(row)
            
            return relevant_rows if relevant_rows else all_rows  # Return all if no matches
        
        # Smart filtering: Scan ALL rows first, then apply query-based filtering
        query_lower = request.query.lower()
        filter_type = (request.filter or '').lower()
        
        # Step 1: Smart filter based on query (scans ALL rows)
        query_filtered_rows = smart_filter_rows(data_rows, request.query, headers)
        
        # Step 2: Apply date-based filter to query-filtered rows
        filtered_rows = []
        for row in query_filtered_rows:
            if len(row) == 0:
                continue
            
            # Apply filter
            if filter_type == 'today':
                # Check if any follow-up date is today
                follow_dates = []
                if follow1_col is not None and follow1_col < len(row):
                    follow_dates.append(parse_date(row[follow1_col]))
                if follow2_col is not None and follow2_col < len(row):
                    follow_dates.append(parse_date(row[follow2_col]))
                if follow3_col is not None and follow3_col < len(row):
                    follow_dates.append(parse_date(row[follow3_col]))
                
                if not any(d == today for d in follow_dates if d):
                    continue
                    
            elif filter_type == 'this_week' or filter_type == 'this week':
                # Check if any follow-up date is within this week
                follow_dates = []
                if follow1_col is not None and follow1_col < len(row):
                    follow_dates.append(parse_date(row[follow1_col]))
                if follow2_col is not None and follow2_col < len(row):
                    follow_dates.append(parse_date(row[follow2_col]))
                if follow3_col is not None and follow3_col < len(row):
                    follow_dates.append(parse_date(row[follow3_col]))
                
                if not any(week_ago <= d <= today for d in follow_dates if d):
                    continue
                    
            elif filter_type == 'overdue':
                # Check if any follow-up date is in the past
                follow_dates = []
                if follow1_col is not None and follow1_col < len(row):
                    follow_dates.append(parse_date(row[follow1_col]))
                if follow2_col is not None and follow2_col < len(row):
                    follow_dates.append(parse_date(row[follow2_col]))
                if follow3_col is not None and follow3_col < len(row):
                    follow_dates.append(parse_date(row[follow3_col]))
                
                if not any(d < today for d in follow_dates if d):
                    continue
            
            filtered_rows.append(row)
        
        # Extract member IDs and names from filtered rows
        member_ids = []
        member_names = []
        member_data = []  # Store {id, name, location, phone, fields} dictionaries

        if member_id_col is not None:
            for row in filtered_rows:
                if member_id_col < len(row) and row[member_id_col]:
                    mid = str(row[member_id_col]).strip()
                    member_ids.append(mid)
                    
                    # Try to get patient/customer name (check both patient and attender)
                    name = ""
                    if patient_name_col is not None and patient_name_col < len(row):
                        name = str(row[patient_name_col]).strip()
                    
                    # Also check attender name
                    attender_name = ""
                    if attender_name_col is not None and attender_name_col < len(row):
                        attender_name = str(row[attender_name_col]).strip()
                    
                    # Prefer patient name, but use attender if patient is empty
                    if not name and attender_name:
                        name = attender_name
                    
                    # If still no name, try other name columns
                    if not name:
                        for i, h in enumerate(headers):
                            if 'name' in h and i < len(row):
                                potential_name = str(row[i]).strip()
                                if potential_name and potential_name != mid:
                                    name = potential_name
                                    break
                    
                    name = name if name else "Unknown"

                    # Determine best available location information
                    location = ""
                    if patient_location_col is not None and patient_location_col < len(row):
                        location = str(row[patient_location_col]).strip()
                    if (not location) and location_col is not None and location_col < len(row):
                        location = str(row[location_col]).strip()
                    if (not location) and area_col is not None and area_col < len(row):
                        location = str(row[area_col]).strip()
                    if not location:
                        # Search for other location-related headers
                        for i, h in enumerate(headers):
                            if any(keyword in h for keyword in ['city', 'town', 'district']) and i < len(row):
                                potential_loc = str(row[i]).strip()
                                if potential_loc:
                                    location = potential_loc
                                    break
                    location = location if location else "Unknown location"

                    # Extract phone/mobile number if present
                    phone = ""
                    if mobile_col is not None and mobile_col < len(row):
                        phone = str(row[mobile_col]).strip()

                    email_value = ""
                    if email_col is not None and email_col < len(row):
                        email_value = str(row[email_col]).strip()

                    # Build a field map for this row (header -> value)
                    field_map = {}
                    for ci, ch in enumerate(original_headers):
                        if ci < len(row):
                            val = str(row[ci]).strip()
                            if val:
                                field_map[ch] = val

                    member_names.append(name)
                    # Store both patient and attender names for AI search
                    member_data.append({
                        "id": mid, 
                        "name": name, 
                        "patient_name": name if patient_name_col is not None and patient_name_col < len(row) else "",
                        "attender_name": attender_name,
                        "location": location, 
                        "phone": phone, 
                        "fields": field_map,
                        "email": email_value,
                    })
        
        # Collect status counts for context
        status_counts = {}
        if lead_status_col is not None:
            for row in filtered_rows:
                if lead_status_col < len(row):
                    status = str(row[lead_status_col]).strip() or 'Unknown'
                    status_counts[status] = status_counts.get(status, 0) + 1
        
        # Detect special commands (e.g., send mail)
        response_text = None
        send_mail_triggered = False
        query_lower = request.query.lower()

        if "send mail" in query_lower and "follow" in query_lower:
            send_mail_triggered = True
            target_member: Optional[Dict[str, Any]] = None
            target_member_id = None

            id_match = re.search(r"(mid-[\w-]+)", request.query, flags=re.IGNORECASE)
            if id_match:
                target_member_id = id_match.group(1).upper()

            if target_member_id:
                target_member = next((m for m in member_data if m.get("id", "").upper() == target_member_id), None)
            else:
                for item in member_data:
                    name_candidates = [item.get('name'), item.get('patient_name'), item.get('attender_name')]
                    for candidate in name_candidates:
                        if candidate and candidate.lower() in query_lower:
                            target_member = item
                            break
                    if target_member:
                        break

            if target_member:
                recipient_email = target_member.get("email")
                if not recipient_email and target_member.get("fields"):
                    for k, v in target_member["fields"].items():
                        if 'email' in k.lower() and str(v).strip():
                            recipient_email = str(v).strip()
                            break

                if recipient_email:
                    background_tasks.add_task(send_follow_email, recipient_email, target_member, request.query)
                    response_text = (
                        f"Scheduled follow-up email to {recipient_email} for "
                        f"{target_member.get('name', 'the member')} ({target_member.get('id')})."
                    )
                else:
                    response_text = (
                        f"I found {target_member.get('id')} but there is no email address on file."
                    )
            else:
                response_text = "I could not identify which member to email. Please include the member name or ID."

        # Build CRM data summary for AI with names
        crm_summary_parts = [
            f"Total records: {len(filtered_rows)}",
            f"Filter applied: {filter_type or 'none'}",
            f"Members found: {len(member_ids)}"
        ]
        # Include ALL available fields (schema) to allow AI to use any column dynamically
        if original_headers:
            hdr_text = ", ".join(original_headers)
            crm_summary_parts.append(f"Available fields: {hdr_text}")
        
        # Add member details (ID + Name) for AI context - ALL members for real-time accuracy
        if member_data:
            member_details = []
            for item in member_data:  # Include ALL members, no limit
                # Show both patient and attender names if different - COMPACT format
                att = item.get('attender_name', '')
                pat = item.get('patient_name', '')
                
                if att and pat and att != pat:
                    # Both names present and different
                    name_part = f"{att}/{pat}"
                elif att:
                    name_part = att
                elif pat:
                    name_part = pat
                else:
                    name_part = item['name']
                
                # Compact format: ID(Name-Location-Phone)
                parts = [item['id'], f"({name_part}"]
                if item.get('location') and item['location'] != 'Unknown location':
                    parts[-1] += f"-{item['location']}"
                if item.get('phone'):
                    parts[-1] += f"-{item['phone']}"
                parts[-1] += ")"
                member_details.append("".join(parts))
            crm_summary_parts.append(f"Members: {', '.join(member_details)}")
            
            # Add field details for queries asking about specific fields
            query_lower_check = request.query.lower()
            field_query_keywords = ['age', 'gender', 'email', 'service', 'pain', 'source', 'agent', 'assigned']
            if any(kw in query_lower_check for kw in field_query_keywords):
                # Include relevant field data for ALL members
                field_lines = []
                for item in member_data[:50]:  # Limit to 50 for token efficiency
                    fm = item.get('fields') or {}
                    # Extract fields mentioned in query
                    relevant_fields = {}
                    for field_name, field_value in fm.items():
                        field_lower = field_name.lower()
                        if any(kw in field_lower for kw in field_query_keywords):
                            relevant_fields[field_name] = field_value
                    if relevant_fields:
                        field_str = "; ".join([f"{k}={v}" for k, v in relevant_fields.items()])
                        field_lines.append(f"{item['id']}: {field_str}")
                if field_lines:
                    crm_summary_parts.append(f"Field details: {' | '.join(field_lines)}")
        
        if status_counts:
            status_summary = ', '.join([f"{k}: {v}" for k, v in status_counts.items()])
            crm_summary_parts.append(f"Lead statuses: {status_summary}")
        
        if filter_type == 'today':
            crm_summary_parts.append("These members need follow-up TODAY")
        elif filter_type == 'this_week' or filter_type == 'this week':
            crm_summary_parts.append("These members need follow-up THIS WEEK")
        elif filter_type == 'overdue':
            crm_summary_parts.append("These follow-ups are OVERDUE")
        
        crm_data_summary = "\n".join(crm_summary_parts)
        
        # Try to get AI-powered response first (only if no special command handled it)
        ai_response_text = None
        if response_text is None and AI_ENABLED:
            print(f"🤖 Calling {AI_PROVIDER.upper()} AI for query: {request.query}")

            if AI_PROVIDER == 'groq' and GROQ_API_KEY:
                ai_response_text = await query_groq_ai(
                    user_query=request.query,
                    crm_data_summary=crm_data_summary,
                    member_ids=member_ids
                )
            elif HF_ENABLED and HF_TOKEN:
                ai_response_text = await query_huggingface_ai(
                    user_query=request.query,
                    crm_data_summary=crm_data_summary,
                    member_ids=member_ids
                )

            if ai_response_text:
                print(f"✅ Using {AI_PROVIDER.upper()} AI response")
                response_text = ai_response_text
            else:
                print("⚠️ AI failed, using fallback")

        # Fall back to rule-based response if AI fails or is disabled
        if response_text is None:
            response_text = generate_fallback_response(
                query_lower=query_lower,
                filter_type=filter_type,
                filtered_rows=filtered_rows,
                member_ids=member_ids,
                status_counts=status_counts,
                member_data=member_data
            )
        
        return {
            "response": response_text,
            "member_ids": member_ids,  # Return all IDs
            "connected": True,
            "count": len(filtered_rows)
        }
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"AI Chat error: {e}")
        print(f"Full traceback:\n{error_details}")
        return {
            "response": f"Sorry, I encountered an error: {str(e)}. Please check if Google Sheets is properly configured.",
            "member_ids": [],
            "connected": False,
            "error_details": str(e)
        }



@app.get("/api/settings/charges")
async def get_charge_settings():
    return load_settings()


@app.post("/api/settings/charges")
async def update_charge_settings(settings: Settings):
    save_settings(settings)
    return {"status": "success", "settings": settings}


# ============ Dropdown Management API Endpoints ============

@app.get("/api/dropdown-options")
async def api_get_all_dropdown_options():
    """Get all dropdown fields and their options from DropdownOption sheet."""
    try:
        options = get_all_dropdown_options()
        return {"status": "success", "dropdown_options": options}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching dropdown options: {str(e)}")


@app.get("/api/dropdown-options/{field_name}")
async def api_get_dropdown_options(field_name: str):
    """Get options for a specific dropdown field."""
    try:
        options = get_dropdown_options_for_field(field_name)
        return {"status": "success", "field_name": field_name, "options": options}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching options for {field_name}: {str(e)}")


@app.post("/api/dropdown-options/{field_name}")
async def api_add_dropdown_option(field_name: str, option: str = Body(..., embed=True)):
    """Add a new option to a dropdown field."""
    try:
        if not option or not option.strip():
            raise HTTPException(status_code=400, detail="Option cannot be empty")
        
        result = add_dropdown_option(field_name, option.strip())
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error adding option: {str(e)}")


@app.delete("/api/dropdown-options/{field_name}/{option}")
async def api_delete_dropdown_option(field_name: str, option: str):
    """Remove an option from a dropdown field."""
    try:
        result = delete_dropdown_option(field_name, option)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting option: {str(e)}")


@app.post("/api/sync-schema-sheet")
async def api_sync_schema_sheet():
    """Sync schema fields with DropdownOption sheet columns bidirectionally."""
    try:
        # Sync schema to sheet (add new columns for new schema fields)
        sync_schema_to_dropdown_sheet()
        
        # Sync sheet to schema (update schema with sheet options)
        sync_dropdown_options_to_schema()
        
        return {
            "status": "success",
            "message": "Schema and DropdownOption sheet synced successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing: {str(e)}")


@app.get("/api/settings/charges")
async def get_charge_settings():
    return load_settings()

@app.post("/api/settings/charges")
async def update_charge_settings(settings: Settings):
    save_settings(settings)
    return {"status": "success", "settings": settings}

@app.get("/admission-details")
async def get_admission_details(member_id: str = Query(...)):
    """Fetch details for a specific patient by Member ID from Patient Admission sheet."""
    try:
        # Reuse the view logic to get all data, then filter
        # Optimization: In real DB we'd query directly, but for Sheet we fetch all or search.
        # Let's fetch all for now as sheet size is manageable, or search if we had a search helper.
        # We have 'view_patient_admissions'. Let's call it logic directly.
        
        client, spreadsheet = get_patient_admission_sheet_client()
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except gspread.WorksheetNotFound:
             raise HTTPException(status_code=404, detail="Patient Admission Sheet1 not found")
             
        records = sheet.get_all_records()
        
        # Filter
        member_id_clean = member_id.strip().lower()
        
        target_record = None
        for r in records:
            # Check common ID keys
            curr_id = str(r.get("Member ID Key") or r.get("memberidkey") or r.get("Member ID") or "").strip().lower()
            if curr_id == member_id_clean:
                target_record = r
                break
        
        if target_record:
            return target_record
        else:
            raise HTTPException(status_code=404, detail="Patient not found in Admission records")

    except Exception as e:
        print(f"Error fetching admission details: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class DischargePayload(BaseModel):
    patient_data: Dict[str, Any]
    billing_data: Dict[str, Any]
    totals: Dict[str, Any]
    calculated_days: int

@app.post("/generate-discharge-summary")
async def generate_discharge_summary(payload: dict):
    try:
        patient = payload.get("patient_data", {})
        totals = payload.get("totals", {})
        billing_data = payload.get("billing_data", {})
        days = payload.get("calculated_days", 1)

        if not patient or not totals:
            raise Exception("Missing patient or billing data")

        buffer = io.BytesIO()

        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.colors import HexColor, black, white, lightgrey
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader

        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Colors
        primary_green = HexColor("#2E7D32")
        dark_gray = HexColor("#333333")
        light_gray = HexColor("#666666")
        border_gray = HexColor("#E0E0E0")
        bg_light = HexColor("#F5F5F5")
        section_bg = HexColor("#F8F9FA")
        
        # Logo path
        logo_path = os.path.join(os.path.dirname(__file__), "Gw- Logo new (2) (1).png")
        
        # Page settings
        header_height = 100
        footer_height = 50
        margin_left = 40
        margin_right = 40
        content_width = width - margin_left - margin_right
        
        # Track current page
        page_num = [1]
        
        # Helper function to get patient value with multiple key attempts
        def get_patient_val(keys):
            if isinstance(keys, str):
                keys = [keys]
            for key in keys:
                if key in patient and patient[key]:
                    return str(patient[key])
                lower_key = key.lower()
                for pk in patient:
                    if pk.lower() == lower_key and patient[pk]:
                        return str(patient[pk])
                    if pk.lower().replace(" ", "").replace("_", "") == lower_key.replace(" ", "").replace("_", "") and patient[pk]:
                        return str(patient[pk])
            return "-"
        
        def draw_header():
            """Draw header on each page"""
            # Header background
            c.setFillColor(white)
            c.rect(0, height - header_height, width, header_height, fill=1, stroke=0)
            
            # Logo
            if os.path.exists(logo_path):
                try:
                    logo = ImageReader(logo_path)
                    c.drawImage(logo, 30, height - 85, width=70, height=70, preserveAspectRatio=True, mask='auto')
                except Exception as logo_err:
                    print(f"Logo error: {logo_err}")
            
            # Hospital name and details
            c.setFillColor(dark_gray)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(110, height - 35, "GRAND WORLD ELDER CARE")
            
            c.setFont("Helvetica", 8)
            c.setFillColor(light_gray)
            c.drawString(110, height - 48, "Assisted Living  |  Clinics  |  Home Nursing")
            c.drawString(110, height - 60, "Contact: +91-XXXXXXXXXX  |  Email: info@grandworld.com")
            c.drawString(110, height - 72, "Address: Chennai, Tamil Nadu, India")
            
            # Document title - right aligned
            c.setFillColor(primary_green)
            c.setFont("Helvetica-Bold", 14)
            c.drawRightString(width - 40, height - 35, "DISCHARGE SUMMARY")
            
            # Date - right aligned below title
            c.setFont("Helvetica", 9)
            c.setFillColor(light_gray)
            current_date = datetime.now().strftime("%d %B %Y")
            c.drawRightString(width - 40, height - 50, f"Date: {current_date}")
            
            # Header bottom border
            c.setStrokeColor(primary_green)
            c.setLineWidth(2)
            c.line(30, height - header_height, width - 30, height - header_height)
        
        def draw_footer():
            """Draw footer on each page"""
            c.setStrokeColor(primary_green)
            c.setLineWidth(1)
            c.line(30, footer_height, width - 30, footer_height)
            
            c.setFont("Helvetica", 7)
            c.setFillColor(light_gray)
            c.drawCentredString(width / 2, footer_height - 15, "This is a computer-generated document. For any queries, please contact the hospital administration.")
            c.drawCentredString(width / 2, footer_height - 27, "Thank you for choosing Grand World Elder Care. Wishing you good health!")
            
            # Page number
            c.drawRightString(width - 40, footer_height - 15, f"Page {page_num[0]}")
        
        def check_page_break(y_pos, needed_space=100):
            """Check if we need a new page and create one if necessary"""
            if y_pos < footer_height + needed_space:
                draw_footer()
                c.showPage()
                page_num[0] += 1
                draw_header()
                return height - header_height - 25
            return y_pos
        
        def draw_section_header(y_pos, title):
            """Draw a section header with consistent styling"""
            y_pos = check_page_break(y_pos, 80)
            c.setFillColor(primary_green)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(margin_left, y_pos, title)
            y_pos -= 5
            c.setStrokeColor(primary_green)
            c.setLineWidth(1)
            c.line(margin_left, y_pos, margin_left + 180, y_pos)
            return y_pos - 18
        
        def draw_field(x, y_pos, label, value, label_width=95):
            """Draw a field with label and value"""
            c.setFont("Helvetica-Bold", 8)
            c.setFillColor(dark_gray)
            c.drawString(x, y_pos, f"{label}:")
            c.setFont("Helvetica", 8)
            c.setFillColor(light_gray)
            # Truncate long values
            val_str = str(value) if value and value != "-" else "-"
            if len(val_str) > 30:
                val_str = val_str[:27] + "..."
            c.drawString(x + label_width, y_pos, val_str)
        
        def draw_field_full_width(y_pos, label, value):
            """Draw a field that spans full width for long text"""
            y_pos = check_page_break(y_pos, 30)
            c.setFont("Helvetica-Bold", 8)
            c.setFillColor(dark_gray)
            c.drawString(margin_left, y_pos, f"{label}:")
            c.setFont("Helvetica", 8)
            c.setFillColor(light_gray)
            val_str = str(value) if value and value != "-" else "-"
            # Word wrap for long text
            if len(val_str) > 80:
                words = val_str.split()
                lines = []
                current_line = ""
                for word in words:
                    if len(current_line + " " + word) < 80:
                        current_line = current_line + " " + word if current_line else word
                    else:
                        lines.append(current_line)
                        current_line = word
                if current_line:
                    lines.append(current_line)
                y_pos -= 12
                for line in lines[:3]:  # Max 3 lines
                    c.drawString(margin_left + 10, y_pos, line.strip())
                    y_pos -= 12
            else:
                c.drawString(margin_left + 100, y_pos, val_str)
                y_pos -= 15
            return y_pos
        
        # ==================== START DRAWING ====================
        draw_header()
        y = height - header_height - 25
        
        left_col = margin_left
        right_col = margin_left + 270
        
        # ==================== 1. PATIENT INFORMATION ====================
        y = draw_section_header(y, "PATIENT INFORMATION")
        
        # Row 1
        draw_field(left_col, y, "Member ID", get_patient_val(["memberidkey", "member_id_key", "memberid", "id"]))
        draw_field(right_col, y, "Registration Date", get_patient_val(["date", "registration_date", "reg_date"]))
        y -= 15
        
        # Row 2
        draw_field(left_col, y, "Patient Name", get_patient_val(["patientname", "patient_name", "name", "firstname"]))
        draw_field(right_col, y, "Last Name", get_patient_val(["patientlastname", "patient_last_name", "lastname"]))
        y -= 15
        
        # Row 3
        draw_field(left_col, y, "Gender", get_patient_val(["gender", "sex"]))
        draw_field(right_col, y, "Date of Birth", get_patient_val(["dateofbirth", "date_of_birth", "dob"]))
        y -= 15
        
        # Row 4
        draw_field(left_col, y, "Age", get_patient_val(["age"]))
        draw_field(right_col, y, "Blood Group", get_patient_val(["patientblood", "patient_blood", "bloodgroup", "blood_group", "blood"]))
        y -= 15
        
        # Row 5
        draw_field(left_col, y, "Marital Status", get_patient_val(["patientmaritalstatus", "patient_marital_status", "maritalstatus"]))
        draw_field(right_col, y, "Nationality", get_patient_val(["nationality"]))
        y -= 15
        
        # Row 6
        draw_field(left_col, y, "Religion", get_patient_val(["religion"]))
        draw_field(right_col, y, "Aadhaar No", get_patient_val(["aadhaar", "aadhar", "aadhaar_no"]))
        y -= 15
        
        # Row 7
        draw_field(left_col, y, "ID Proof Type", get_patient_val(["idprooftype", "id_proof_type"]))
        draw_field(right_col, y, "ID Proof Number", get_patient_val(["idproofnumber", "id_proof_number"]))
        y -= 25
        
        # ==================== 2. CONTACT INFORMATION ====================
        y = draw_section_header(y, "CONTACT INFORMATION")
        
        # Row 1
        draw_field(left_col, y, "Mobile Number", get_patient_val(["mobilenumber", "mobile_number", "mobile", "phone", "contact"]))
        draw_field(right_col, y, "Email ID", get_patient_val(["emailid", "email_id", "email"]))
        y -= 15
        
        # Row 2
        draw_field(left_col, y, "Door Number", get_patient_val(["doornumber", "door_number"]))
        draw_field(right_col, y, "Street", get_patient_val(["street"]))
        y -= 15
        
        # Row 3
        draw_field(left_col, y, "City", get_patient_val(["city", "area"]))
        draw_field(right_col, y, "District", get_patient_val(["district", "patientlocation", "patient_location"]))
        y -= 15
        
        # Row 4
        draw_field(left_col, y, "State", get_patient_val(["state"]))
        draw_field(right_col, y, "Pin Code", get_patient_val(["pincode", "pin_code"]))
        y -= 25
        
        # ==================== 3. EMERGENCY CONTACT ====================
        y = draw_section_header(y, "EMERGENCY CONTACT DETAILS")
        
        # Row 1
        draw_field(left_col, y, "Contact Name", get_patient_val(["relationalname", "relational_name", "attendername", "attender_name", "emergencyname"]))
        draw_field(right_col, y, "Relationship", get_patient_val(["relationalrelationship", "relational_relationship", "relationship"]))
        y -= 15
        
        # Row 2
        draw_field(left_col, y, "Contact Mobile", get_patient_val(["relationalmobile", "relational_mobile", "emergencymobile"]))
        draw_field(right_col, y, "Alt. Mobile", get_patient_val(["relationalmobilealternative", "relational_mobile_alternative", "altmobile"]))
        y -= 15
        
        # Emergency Address
        y = draw_field_full_width(y, "Emergency Address", get_patient_val(["emergencyaddress", "emergency_address"]))
        y -= 10
        
        # ==================== 4. MEDICAL HISTORY ====================
        y = draw_section_header(y, "MEDICAL HISTORY")
        
        # Row 1
        draw_field(left_col, y, "Current Status", get_patient_val(["patientcurrentstatus", "patient_current_status", "currentstatus"]))
        draw_field(right_col, y, "Sugar Level", get_patient_val(["patientsugarlevel", "patient_sugar_level", "sugarlevel"]))
        y -= 15
        
        # Row 2
        draw_field(left_col, y, "Pain Point", get_patient_val(["painpoint", "pain_point"]))
        draw_field(right_col, y, "Allergies", get_patient_val(["patientallergy", "patient_allergy", "allergy", "allergies"]))
        y -= 15
        
        # Medical History (full width)
        y = draw_field_full_width(y, "Medical History", get_patient_val(["patientmedicalhistory", "patient_medical_history", "medicalhistory"]))
        y -= 10
        
        # ==================== 5. SERVICE DETAILS ====================
        y = draw_section_header(y, "SERVICE DETAILS")
        
        # Row 1
        draw_field(left_col, y, "Service Type", get_patient_val(["service", "servicetype", "service_type"]))
        draw_field(right_col, y, "Enquiry For", get_patient_val(["enquirymadefor", "enquiry_made_for", "enquiry"]))
        y -= 15
        
        # Row 2
        draw_field(left_col, y, "Services Provided", get_patient_val(["providingservices", "providing_services", "serviceprovided"]))
        draw_field(right_col, y, "Hospital Location", get_patient_val(["hospitallocation", "hospital_location"]))
        y -= 15
        
        # Row 3
        draw_field(left_col, y, "Caretaker Name", get_patient_val(["caretakername", "caretaker_name"]))
        draw_field(right_col, y, "Source", get_patient_val(["source"]))
        y -= 25
        
        # ==================== 6. ADMISSION DETAILS ====================
        y = draw_section_header(y, "ADMISSION DETAILS")
        
        # Row 1
        draw_field(left_col, y, "Check-In Date", get_patient_val(["checkindate", "check_in_date", "admissiondate", "admission_date"]))
        draw_field(right_col, y, "Check-Out Date", get_patient_val(["checkoutdate", "check_out_date", "dischargedate", "discharge_date"]))
        y -= 15
        
        # Row 2
        draw_field(left_col, y, "Room Type", get_patient_val(["roomtype", "room_type", "room"]))
        draw_field(right_col, y, "Room Rent", get_patient_val(["roomrent", "room_rent"]))
        y -= 15
        
        # Row 3
        draw_field(left_col, y, "Bed No", get_patient_val(["bedno", "bed_no", "bed"]))
        draw_field(right_col, y, "Total Stay", f"{days} Day(s)")
        y -= 15
        
        # Row 4
        draw_field(left_col, y, "Attender Name", get_patient_val(["attendername", "attender_name"]))
        draw_field(right_col, y, "Lead Status", get_patient_val(["leadstatus", "lead_status", "status"]))
        y -= 25

        # ==================== 7. BILLING SUMMARY ====================
        y = check_page_break(y, 220)  # Need space for billing table
        y = draw_section_header(y, "BILLING SUMMARY")

        # Table settings
        table_left = margin_left
        table_right = width - margin_right
        table_width = table_right - table_left
        row_height = 20

        # Table header
        c.setFillColor(primary_green)
        c.rect(table_left, y - row_height + 5, table_width, row_height, fill=1, stroke=0)

        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(table_left + 10, y - 10, "Description")
        c.drawString(table_left + 220, y - 10, "Rate/Day (₹)")
        c.drawString(table_left + 320, y - 10, "Days")
        c.drawRightString(table_right - 10, y - 10, "Amount (₹)")
        y -= row_height

        def draw_table_row(y_pos, desc, rate, days_count, amount, is_fixed=False, alt_bg=False):
            if alt_bg:
                c.setFillColor(bg_light)
            else:
                c.setFillColor(white)
            c.rect(table_left, y_pos - row_height + 5, table_width, row_height, fill=1, stroke=0)

            c.setFillColor(dark_gray)
            c.setFont("Helvetica", 8)
            c.drawString(table_left + 10, y_pos - 10, desc)

            if is_fixed:
                c.drawString(table_left + 220, y_pos - 10, "-")
                c.drawString(table_left + 320, y_pos - 10, "-")
            else:
                c.drawString(table_left + 220, y_pos - 10, f"{rate:,.0f}" if rate else "0")
                c.drawString(table_left + 320, y_pos - 10, str(days_count))

            c.drawRightString(table_right - 10, y_pos - 10, f"{amount:,.0f}" if amount else "0")
            return y_pos - row_height

        # Daily charges
        room_rate = billing_data.get("room_charge", 0)
        bed_rate = billing_data.get("bed_charge", 0)
        nurse_rate = billing_data.get("nurse_payment", 0)
        additional_nurse_rate = billing_data.get("additional_nurse_payment", 0)
        other_charges_rate = billing_data.get("other_charges_amenities", 0)
        hospital_rate = billing_data.get("hospital_payment", 0)

        y = draw_table_row(y, "Room Charge", room_rate, days, totals.get("room", 0), alt_bg=True)
        y = draw_table_row(y, "Bed Charge", bed_rate, days, totals.get("bed", 0), alt_bg=False)
        y = draw_table_row(y, "Nursing Fee", nurse_rate, days, totals.get("nurse", 0), alt_bg=True)
        y = draw_table_row(y, "Additional Nursing Fee", additional_nurse_rate, days, totals.get("additional_nurse", 0), alt_bg=False)
        y = draw_table_row(y, "Other Charges (Amenities)", other_charges_rate, days, totals.get("other_charges", 0), alt_bg=True)
        y = draw_table_row(y, "Hospital Fee", hospital_rate, days, totals.get("hospital", 0), alt_bg=False)

        # Fixed charges
        y = draw_table_row(y, "Doctor Fee", 0, 0, totals.get("doctor", 0), is_fixed=True, alt_bg=True)
        y = draw_table_row(y, "Service Charge", 0, 0, totals.get("service", 0), is_fixed=True, alt_bg=False)
        
        # Discount (subtract from total)
        discount_amount = totals.get("discount", 0)
        if discount_amount > 0:
            y = draw_table_row(y, "Discount", 0, 0, -discount_amount, is_fixed=True, alt_bg=True)

        # Grand total row with extra spacing below
        y -= 3
        c.setFillColor(primary_green)
        c.rect(table_left, y - row_height + 5, table_width, row_height, fill=1, stroke=0)

        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(table_left + 10, y - 11, "GRAND TOTAL")
        grand_total = totals.get("grand", 0)
        c.drawRightString(table_right - 10, y - 11, f"₹ {grand_total:,.0f}")
        y -= row_height + 90  # extra gap after grand total

        # Amount in words
        y = check_page_break(y, 160)

        def number_to_words(num):
            ones = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
                    'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
                    'Seventeen', 'Eighteen', 'Nineteen']
            tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

            if num == 0:
                return 'Zero'
            num = int(num)
            if num < 20:
                return ones[num]
            elif num < 100:
                return tens[num // 10] + ('' if num % 10 == 0 else ' ' + ones[num % 10])
            elif num < 1000:
                return ones[num // 100] + ' Hundred' + ('' if num % 100 == 0 else ' and ' + number_to_words(num % 100))
            elif num < 100000:
                return number_to_words(num // 1000) + ' Thousand' + ('' if num % 1000 == 0 else ' ' + number_to_words(num % 1000))
            elif num < 10000000:
                return number_to_words(num // 100000) + ' Lakh' + ('' if num % 100000 == 0 else ' ' + number_to_words(num % 100000))
            else:
                return number_to_words(num // 10000000) + ' Crore' + ('' if num % 10000000 == 0 else ' ' + number_to_words(num % 10000000))

        amount_words = number_to_words(grand_total) + " Rupees Only"
        c.setFillColor(dark_gray)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(margin_left, y, "Amount in Words:")
        c.setFont("Helvetica-Oblique", 8)
        c.drawString(margin_left + 90, y, amount_words)
        y -= 70  # extra gap before signatures

        # Signatures section
        y = check_page_break(y, 120)
        c.setStrokeColor(border_gray)
        c.setLineWidth(0.5)
        c.line(margin_left, y + 10, width - margin_right, y + 10)

        sig_y = y - 25
        c.setFont("Helvetica", 8)
        c.setFillColor(light_gray)
        c.drawString(60, sig_y + 35, "Patient/Attender Signature")
        c.setStrokeColor(dark_gray)
        c.setLineWidth(0.5)
        c.line(60, sig_y + 30, 180, sig_y + 30)

        c.drawString(380, sig_y + 35, "Authorized Signature")
        c.line(380, sig_y + 30, 500, sig_y + 30)

        c.setFont("Helvetica", 7)
        c.setFillColor(light_gray)
        c.drawCentredString(440, sig_y, "(Hospital Stamp)")

        # Draw footer on last page
        draw_footer()

        c.showPage()
        c.save()

        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": "attachment; filename=Discharge_Summary.pdf"
            }
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/search_data")
async def search_data(query: Optional[str] = Query(None, min_length=2), limit: int = 50):
    """Search for patient data for auto-fill. If query is empty, returns all data up to limit."""
    try:
        # Use gspread directly to get raw data
        if not os.path.exists(CREDENTIALS_FILE):
             # Try to start without creds (maybe public?) No, strict requirement here.
             raise HTTPException(status_code=404, detail="Google credentials file not found")

        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        sheet = spreadsheet.sheet1
        values = sheet.get_all_values()

        if not values:
            return {"results": [], "headers": [], "rows": []}

        headers = values[0]
        rows = values[1:]
        
        results = []
        
        # If no query, return top N rows
        if not query:
            # Convert all rows to dicts
            for row in rows[:limit]:
                 # Handle rows shorter than headers
                r_len = len(row)
                row_dict = {h: (row[i] if i < r_len else "") for i, h in enumerate(headers)}
                results.append(row_dict)
            return {"results": results, "headers": headers, "rows": rows[:limit]} # Return raw rows to ensure frontend compatibility

        q_lower = query.lower()
        
        for row in rows:
            # Handle rows shorter than headers
            r_len = len(row)
            row_dict = {h: (row[i] if i < r_len else "") for i, h in enumerate(headers)}
            
            # Helper to safely get value
            def match(val):
                return str(val).lower().find(q_lower) >= 0

            # Find specific columns if possible or just search all values
            # Heuristic search
            found = False
            
            # Simple broad search across all values in the row
            # But prompt logic suggested checking specific fields?
            # Stick to broad search for robustness:
            for v in row:
                 if match(v):
                     found = True
                     break
            
            if found:
                results.append(row_dict)
                if len(results) >= limit: # Limit results
                    break
                    
        return {"results": results, "headers": headers, "rows": [list(r.values()) for r in results]} # Maintain compatibility
        
    except Exception as e:
        print(f"Search failed: {e}")
        return {"results": [], "headers": [], "rows": []}


# ============== AI Chat Query Endpoint ==============
class ChatQueryRequest(BaseModel):
    query: str
    filter: Optional[str] = "today"

@app.post("/chat_query")
async def chat_query(request: ChatQueryRequest):
    """
    AI Chat endpoint for the CRM assistant.
    Processes natural language queries about patient data, follow-ups, etc.
    """
    try:
        query_lower = request.query.lower().strip()
        filter_type = request.filter or "today"
        
        # Get data from Google Sheets for context
        if not os.path.exists(CREDENTIALS_FILE):
            return {"answer": "I'm having trouble accessing the database. Please check the credentials configuration."}
        
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        sheet = spreadsheet.sheet1
        values = sheet.get_all_values()
        
        if not values or len(values) < 2:
            return {"answer": "No patient data found in the system yet."}
        
        headers = values[0]
        rows = values[1:]
        
        # Convert to list of dicts
        records = []
        for row in rows:
            record = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
            records.append(record)
        
        today = datetime.now().date()
        
        # Filter records based on filter_type
        def parse_date(date_str):
            if not date_str:
                return None
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                try:
                    return datetime.strptime(date_str.strip(), fmt).date()
                except:
                    continue
            return None
        
        # Find date column
        date_col = None
        for h in headers:
            if "date" in h.lower() and "reminder" not in h.lower() and "follow" not in h.lower():
                date_col = h
                break
        if not date_col and headers:
            date_col = headers[0]  # fallback to first column
        
        filtered_records = []
        for rec in records:
            rec_date = parse_date(rec.get(date_col, ""))
            if filter_type == "today":
                if rec_date == today:
                    filtered_records.append(rec)
            elif filter_type == "this_week":
                if rec_date and (today - rec_date).days <= 7 and rec_date <= today:
                    filtered_records.append(rec)
            elif filter_type == "overdue":
                # Check follow-up dates
                for h in headers:
                    if "follow" in h.lower() or "reminder" in h.lower():
                        follow_date = parse_date(rec.get(h, ""))
                        if follow_date and follow_date < today:
                            filtered_records.append(rec)
                            break
            else:  # "all" or empty
                filtered_records.append(rec)
        
        # Process common queries
        if "follow" in query_lower or "today" in query_lower:
            count = len(filtered_records)
            if count == 0:
                return {"answer": f"No follow-ups scheduled for {filter_type.replace('_', ' ')}."}
            
            # Get patient names if available
            name_col = None
            for h in headers:
                if "patient" in h.lower() and "name" in h.lower():
                    name_col = h
                    break
            
            if name_col and count <= 5:
                names = [rec.get(name_col, "Unknown") for rec in filtered_records[:5]]
                return {"answer": f"You have {count} follow-up(s) for {filter_type.replace('_', ' ')}:\n" + "\n".join(f"• {n}" for n in names)}
            else:
                return {"answer": f"You have {count} follow-up(s) scheduled for {filter_type.replace('_', ' ')}."}
        
        elif "count" in query_lower or "how many" in query_lower:
            return {"answer": f"There are {len(filtered_records)} records for {filter_type.replace('_', ' ')}."}
        
        elif "patient" in query_lower:
            count = len(filtered_records)
            return {"answer": f"Found {count} patient record(s) matching your criteria."}
        
        elif "status" in query_lower:
            # Try to find status column
            status_col = None
            for h in headers:
                if "status" in h.lower():
                    status_col = h
                    break
            
            if status_col:
                statuses = {}
                for rec in filtered_records:
                    s = rec.get(status_col, "Unknown") or "Unknown"
                    statuses[s] = statuses.get(s, 0) + 1
                
                status_summary = "\n".join(f"• {k}: {v}" for k, v in statuses.items())
                return {"answer": f"Status breakdown:\n{status_summary}"}
            else:
                return {"answer": "No status information available."}
        
        elif "help" in query_lower:
            return {"answer": "I can help you with:\n• Follow-ups today/this week\n• Patient counts\n• Status summaries\n• Overdue reminders\n\nTry asking: 'Follow ups today?' or 'How many patients this week?'"}
        
        else:
            # Generic response
            count = len(filtered_records)
            return {"answer": f"I found {count} records for {filter_type.replace('_', ' ')}. Try asking about follow-ups, patient counts, or status summaries."}
        
    except Exception as e:
        print(f"Chat query error: {e}")
        import traceback
        traceback.print_exc()
        return {"answer": "I encountered an error processing your request. Please try again."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
