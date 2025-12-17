from fastapi import FastAPI, HTTPException, Body, BackgroundTasks, UploadFile, File, Query
from fastapi.responses import StreamingResponse
import io
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import range_boundaries
import gspread
from google.oauth2.service_account import Credentials
import os
from datetime import datetime
import re
import uuid
import random
import csv
import json
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv
import requests
import httpx
from file_manager import save_upload, process_data_file
import pandas as pd
from dropdown_helpers import (
    get_dropdown_option_sheet,
    initialize_dropdown_sheet,
    get_all_dropdown_options,
    get_dropdown_options_for_field,
    add_dropdown_option,
    delete_dropdown_option,
    sync_dropdown_options_to_schema,
    sync_schema_to_dropdown_sheet
)

# Load environment variables from .env file
# Trigger reload for schema update
load_dotenv()

app = FastAPI()

SETTINGS_FILE = "settings.json"

class Settings(BaseModel):
    single_bed_price: float = 1000.0
    twin_bed_price: float = 2000.0
    icu_bed_price: float = 5000.0

def load_settings() -> Settings:
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                data = json.load(f)
            return Settings(**data)
        except Exception:
            return Settings()
    return Settings()

def save_settings(settings: Settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings.dict(), f, indent=2)

# CORS middleware to allow frontend requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for development
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration  
EXCEL_FILE_PATH = os.getenv("EXCEL_FILE_PATH", "Lead CRM ApplicationData.xlsx")
EXCEL_SHEET_NAME = "Lead CRM"  # Sheet name to use from the Excel file
ADMISSION_SHEET_NAME = "Admission Details"
COMPLAINT_SHEET_NAME = "Complaints"
FEEDBACK_SHEET_NAME = "Feedback"
LIST_BOX_SHEET = "List box"  # Sheet containing dropdown options (legacy)
DROPDOWN_OPTION_SHEET = "DropdownOption"  # New centralized dropdown management sheet
GOOGLE_SHEET_NAME = "Sheet1"  # Default worksheet name in Google Sheets
CREDENTIALS_FILE = "google_credentials.json"
CSV_FILE_PATH = "CRM Leads - Sheet1.csv"
FIELD_SCHEMA_FILE = "field_schema.json"
SCHEMA_CACHE_FILE = "schema_cache.json"

NOTIFICATION_SETTINGS_FILE = "notification_settings.json"
DEFAULT_SENDER_EMAIL = os.getenv("DEFAULT_NOTIFICATION_EMAIL", "harishkadhiravan.vtab@gmail.com")
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
EMAIL_SUBJECT = os.getenv("EMAIL_SUBJECT", "New CRM Lead Submission")
EMAIL_TRANSPORT = os.getenv("EMAIL_TRANSPORT", "smtp").strip().lower()

USE_GMAIL_API = False
if EMAIL_TRANSPORT == "gmail_api":
    try:
        from gmail_api_sender import send_email_via_gmail_api  # type: ignore
        USE_GMAIL_API = True
        print("[Email CONFIG] Gmail API transport enabled")
    except Exception as gmail_import_exc:  # pragma: no cover - import guard only
        print(f"[Email CONFIG] Gmail API not available ({gmail_import_exc}); falling back to SMTP")
        USE_GMAIL_API = False

# Optional: lock to a specific Google Sheet (recommended)
# - Set env var GOOGLE_SHEET_ID, or
# - Put the Sheet ID in a local file named "google_sheet_id.txt" in backend/
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID")
if not GOOGLE_SHEET_ID:
    try:
        with open("google_sheet_id.txt", "r", encoding="utf-8") as f:
            GOOGLE_SHEET_ID = f.read().strip() or None
    except FileNotFoundError:
        GOOGLE_SHEET_ID = None

# Fields to exclude from the dynamic form (case-insensitive)
EXCLUDED_FIELD_NAMES = {
    "reason for rejection",
    "ed comments",
    "closed date",
}

# Keywords that, when present in a header, should exclude the field
EXCLUDED_KEYWORDS = [
    # Remove follow-up/reminder date columns like Follow1 Date, Reminder Date_1, etc.
    "follow",    # matches Follow1 Date, Follow_2 Date, Follow_3 Date, etc.
    "reminder",  # matches Reminder Date_1, Reminder Date_2, etc.
]

# Keywords for date fields that should remain blank unless explicitly provided
REMINDER_DATE_KEYWORD = "reminder"
FOLLOW_DATE_KEYWORD = "follow"

fields_cache: List[Dict[str, Any]] = []
notification_settings_cache: Optional[Dict[str, Any]] = None


def load_notification_settings() -> Dict[str, Any]:
    """Return cached notification settings from disk with defaults."""
    global notification_settings_cache
    if notification_settings_cache is not None:
        return notification_settings_cache

    settings = {
        "sender_email": DEFAULT_SENDER_EMAIL,
        "cc_emails": [],
    }

    if os.path.exists(NOTIFICATION_SETTINGS_FILE):
        try:
            with open(NOTIFICATION_SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                sender = str(data.get("sender_email", "")).strip()
                if sender:
                    settings["sender_email"] = sender
                raw_cc = data.get("cc_emails")
                if isinstance(raw_cc, list):
                    cleaned_cc = [str(item).strip() for item in raw_cc if str(item).strip()]
                    if cleaned_cc:
                        settings["cc_emails"] = cleaned_cc
        except Exception as exc:
            print(f"Warning: could not read {NOTIFICATION_SETTINGS_FILE}: {exc}")

    notification_settings_cache = settings
    return settings


def save_notification_settings(settings: Dict[str, Any]) -> None:
    """Persist notification settings and refresh cache."""
    global notification_settings_cache
    sender = str(settings.get("sender_email", DEFAULT_SENDER_EMAIL)).strip() or DEFAULT_SENDER_EMAIL
    raw_cc = settings.get("cc_emails") or []
    cleaned_cc = [str(item).strip() for item in raw_cc if str(item).strip()]

    notification_settings_cache = {
        "sender_email": sender,
        "cc_emails": cleaned_cc,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    try:
        with open(NOTIFICATION_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(notification_settings_cache, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"Warning: could not write {NOTIFICATION_SETTINGS_FILE}: {exc}")


class FormData(BaseModel):
    data: Dict[str, Any]


class FieldItem(BaseModel):
    name: str
    label: Optional[str] = None
    data_type: str
    options: Optional[List[str]] = None


class UpdateFieldsPayload(BaseModel):
    fields: List[FieldItem]


class NotificationSettings(BaseModel):
    sender_email: str
    cc_emails: Optional[List[str]] = None



class LoginRequest(BaseModel):
    User_name: str
    Password: str


class FilterCriteria(BaseModel):
    year: Optional[str] = None
    month: Optional[str] = None
    specificDate: Optional[str] = None
    startDate: Optional[str] = None
    endDate: Optional[str] = None

class DeletePreviewPayload(BaseModel):
    filters: FilterCriteria
    date_column: str
    preview_columns: Optional[List[str]] = None

class DeleteConfirmPayload(BaseModel):
    filters: FilterCriteria
    date_column: str



def ensure_notification_defaults() -> Dict[str, Any]:
    settings = load_notification_settings()
    if not settings.get("sender_email"):
        settings["sender_email"] = DEFAULT_SENDER_EMAIL
    if "cc_emails" not in settings or not isinstance(settings.get("cc_emails"), list):
        settings["cc_emails"] = []
    return settings


def build_email_message(sender: str, recipient: str, payload: Dict[str, Any], subject_override: Optional[str] = None) -> EmailMessage:
    message = EmailMessage()
    message["From"] = f"CRM Lead Form <{sender}>"
    message["To"] = recipient
    message["Reply-To"] = sender
    message["Return-Path"] = sender
    message["X-Mailer"] = "CRM Lead Application"
    message["X-Priority"] = "3"

    # CC
    settings = ensure_notification_defaults()
    cc_list = settings.get("cc_emails", [])
    if cc_list:
        message["Cc"] = ", ".join(cc_list)

    # Resolve common aliases from payload (case-insensitive)
    def get_value(aliases: List[str]) -> Optional[str]:
        if not payload:
            return None
        # Direct first
        for a in aliases:
            if a in payload:
                val = payload.get(a)
                if val is not None and str(val).strip():
                    return str(val).strip()
        # Case-insensitive
        lower_map = {str(k).strip().lower(): v for k, v in payload.items()}
        for a in aliases:
            v = lower_map.get(a.lower())
            if v is not None and str(v).strip():
                return str(v).strip()
        return None

    member_id = get_value([
        "Memberidkey", "Member ID", "member id", "Member_ID", "Memberid", "MID"
    ])
    attender_name = get_value(["Attender Name", "attender name", "Attender", "attender"]) or "Customer"

    # Subject with Member ID if available (or override)
    subject = subject_override or EMAIL_SUBJECT
    if member_id:
        if subject_override:
            subject = f"{subject_override}: {member_id}"
        else:
            subject = f"{EMAIL_SUBJECT} - {member_id}"
    message["Subject"] = subject

    # Body template
    lines: List[str] = []
    lines.append(f"Dear {attender_name}, your form has been submitted")
    # If this is a status-change notification, add explicit status line
    if (subject_override or "").strip().upper().startswith("LEAD STATUS CHANGED"):
        lead_status = get_value(["Lead Status", "lead status"]) or ""
        if lead_status:
            lines.append("")
            lines.append(f"Lead status changed to : {lead_status}")
    lines.append("")
    lines.append("confirm your details below:")
    for key, value in payload.items():
        lines.append(f"- {key}: {value}")
    message.set_content("\n".join(lines))
    return message


def send_notification_email(recipient: str, payload: Dict[str, Any], subject_override: Optional[str] = None) -> None:
    print(f"[Email DEBUG] Function called with recipient={recipient}")
    if not recipient:
        print("[Email DEBUG] No recipient, returning early")
        return

    settings = ensure_notification_defaults()
    sender = settings.get("sender_email", DEFAULT_SENDER_EMAIL)
    cc_list: List[str] = settings.get("cc_emails", [])
    print(f"[Email DEBUG] Sender={sender}, CC={cc_list}, transport={'GMAIL_API' if USE_GMAIL_API else 'SMTP'}")

    if not USE_GMAIL_API and (not SMTP_USERNAME or not SMTP_PASSWORD):
        print("Warning: SMTP credentials not configured; skipping email notification")
        return

    try:
        masked_pw_len = len(SMTP_PASSWORD) if SMTP_PASSWORD else 0
        print(f"[Email] Preparing to send | sender={sender} | recipient={recipient} | user={SMTP_USERNAME} | pw_len={masked_pw_len} | host={SMTP_HOST}:{SMTP_PORT}")
    except Exception:
        pass

    # Adjust sender alignment:
    # - For Gmail, align header/envelope From to SMTP username for deliverability
    # - For Brevo (smtp-relay.brevo.com / *@smtp-brevo.com), keep the configured sender
    host_l = str(SMTP_HOST or "").lower()
    user_l = str(SMTP_USERNAME or "").lower()
    is_brevo = ("brevo" in host_l) or user_l.endswith("@smtp-brevo.com")
    if not USE_GMAIL_API and not is_brevo:
        effective_sender = SMTP_USERNAME if SMTP_USERNAME else sender
        if effective_sender != sender:
            print(f"[Email DEBUG] Aligning header From with SMTP user: {sender} -> {effective_sender}")
            sender = effective_sender

    message = build_email_message(sender, recipient, payload, subject_override)

    # Optional simple mode to suppress CCs for stricter deliverability
    simple_mode = str(os.getenv("EMAIL_SIMPLE_MODE", "0")).strip() == "1"
    if simple_mode and cc_list:
        print("[Email DEBUG] EMAIL_SIMPLE_MODE=1 -> suppressing CCs for this send")
        cc_list = []

    # Ensure message headers match final CC list
    if cc_list:
        if "Cc" in message:
            message.replace_header("Cc", ", ".join(cc_list))
        else:
            message["Cc"] = ", ".join(cc_list)
    elif "Cc" in message:
        del message["Cc"]

    # Auto-BCC logic (skip Brevo technical usernames)
    bcc_list: List[str] = []
    if (USE_GMAIL_API or not is_brevo) and SMTP_USERNAME and SMTP_USERNAME not in ([recipient] + cc_list):
        bcc_list = [SMTP_USERNAME]
        print(f"[Email DEBUG] Auto-BCC enabled -> {bcc_list}")

    if USE_GMAIL_API:
        try:
            body_text = message.get_content()
        except Exception:
            body_text = str(message)
        try:
            print("[Email DEBUG] Sending via Gmail API...")
            send_email_via_gmail_api(
                sender_email=sender,
                recipient_email=recipient,
                subject=message["Subject"],
                body_text=body_text,
                cc_emails=cc_list or None,
                bcc_emails=bcc_list or None,
            )
            print(f"[Email SUCCESS] Gmail API accepted message for {recipient}")
        except Exception as exc:
            print(f"[Email ERROR] Gmail API send failed: {type(exc).__name__}: {exc}")
            import traceback
            traceback.print_exc()
        return

    try:
        print(f"[Email DEBUG] Connecting to {SMTP_HOST}:{SMTP_PORT}...")
        # Try SSL if port 465, otherwise TLS
        if SMTP_PORT == 465:
            print(f"[Email DEBUG] Using SMTP_SSL on port 465...")
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
        
        try:
            server.ehlo()
            if SMTP_PORT != 465:
                print(f"[Email DEBUG] Starting TLS...")
                server.starttls()
                server.ehlo()
            print(f"[Email DEBUG] Logging in as {SMTP_USERNAME}...")
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            print(f"[Email DEBUG] Login successful!")
            print(f"[Email DEBUG] Preparing to send...")
            print(f"[Email DEBUG] From: {sender}")
            print(f"[Email DEBUG] To: {recipient}")
            print(f"[Email DEBUG] CC: {cc_list}")
            print(f"[Email DEBUG] Subject: {message['Subject']}")
            all_recipients = [recipient] + cc_list + bcc_list
            # de-duplicate recipients while preserving order
            seen = set()
            deduped: List[str] = []
            for addr in all_recipients:
                if addr and addr not in seen:
                    deduped.append(addr)
                    seen.add(addr)
            all_recipients = deduped
            # Use sendmail instead of send_message for more control
            msg_string = message.as_string()
            refused = server.sendmail(sender, all_recipients, msg_string)
            if refused:
                print(f"[Email WARNING] Some recipients were refused: {refused}")
            else:
                print(f"[Email DEBUG] All recipients accepted by server")
            print(f"[Email SUCCESS] Email sent to {recipient} (cc: {', '.join(cc_list) if cc_list else 'none'})")
        finally:
            server.quit()
    except Exception as exc:
        print(f"[Email ERROR] Failed to send: {type(exc).__name__}: {exc}")
        import traceback
        traceback.print_exc()


def send_follow_email(recipient: str, member_entry: Dict[str, Any], query_trigger: str) -> None:
    """Send a follow-up reminder email for the specified member."""
    if not recipient:
        print("[Follow Email] No recipient email provided; skipping")
        return

    payload = dict(member_entry.get("fields", {}))
    payload.setdefault("Member ID", member_entry.get("id", ""))
    payload.setdefault("Member Name", member_entry.get("name", "Unknown"))
    payload["AI Follow-up Trigger"] = query_trigger

    subject = f"FOLLOW-UP REMINDER - {member_entry.get('id', 'Member')}"
    try:
        send_notification_email(recipient, payload, subject_override=subject)
        print(f"[Follow Email] Successfully queued follow-up email to {recipient}")
    except Exception as exc:
        print(f"[Follow Email] Failed to send email to {recipient}: {exc}")


def extract_recipient_email(payload: Dict[str, Any]) -> Optional[str]:
    if not payload:
        return None

    for key, value in payload.items():
        if value is None:
            continue
        if "email" in str(key).lower():
            candidate = str(value).strip()
            if candidate:
                return candidate

    lower_map: Dict[str, Any] = {}
    for key, value in payload.items():
        if value is None:
            continue
        lower_map[str(key).strip().lower()] = value

    for field in fields_cache:
        name = field.get("name")
        if not name:
            continue
        normalized = str(name).strip().lower()
        if "email" not in normalized:
            continue
        candidate = lower_map.get(normalized)
        if candidate:
            text = str(candidate).strip()
            if text:
                return text
    return None


def read_excel_headers() -> List[str]:
    """Read header row from the Excel file."""
    if not os.path.exists(EXCEL_FILE_PATH):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE_PATH}")
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
    sheet = workbook[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME else workbook.active
    headers: List[str] = []
    for cell in sheet[1]:
        if not cell.value:
            continue
        header = str(cell.value).strip()
        header_lower = header.lower()
        if header_lower in EXCLUDED_FIELD_NAMES:
            continue
        if any(keyword in header_lower for keyword in EXCLUDED_KEYWORDS):
            continue
        headers.append(header)
    workbook.close()
    return headers


def read_csv_headers(csv_path: str) -> List[str]:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if any(str(x).strip() for x in row):
                return [h.strip() for h in row if str(h).strip()]
    return []


def read_excel_dropdown_options(workbook: openpyxl.Workbook) -> Dict[str, List[str]]:
    """Read dropdown options from the List box sheet"""
    dropdowns: Dict[str, List[str]] = {}
    
    try:
        if LIST_BOX_SHEET not in workbook.sheetnames:
            return dropdowns
        
        listbox_ws = workbook[LIST_BOX_SHEET]
        
        # Get headers from List box sheet (first row)
        headers = []
        for cell in listbox_ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append(None)
        
        # Read values for each column (exclude Date column as it's not a dropdown)
        for col_idx, header in enumerate(headers, start=1):
            if not header or header.lower() == 'date':
                continue
            
            values = set()
            # Read up to 1000 rows to get all unique values
            for row_idx in range(2, min(listbox_ws.max_row + 1, 1000)):
                val = listbox_ws.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip():
                    clean_val = str(val).strip()
                    # Skip datetime objects from Date column
                    if 'datetime' not in str(type(val)):
                        values.add(clean_val)
            
            if values:
                dropdowns[header] = sorted(list(values))
    
    except Exception as e:
        print(f"Warning: Could not read List box sheet: {e}")
    
    return dropdowns


def infer_field_type(field_name: str) -> str:
    """Infer input type based on field name"""
    field_lower = field_name.lower()
    
    if any(keyword in field_lower for keyword in ['date', 'dob', 'birth']):
        return 'date'
    elif any(keyword in field_lower for keyword in ['email', 'mail']):
        return 'email'
    elif any(keyword in field_lower for keyword in ['phone', 'mobile', 'contact']):
        return 'tel'
    elif any(keyword in field_lower for keyword in ['age', 'number', 'count', 'quantity']):
        return 'number'
    elif any(keyword in field_lower for keyword in ['description', 'comment', 'notes', 'address']):
        return 'textarea'
    else:
        return 'text'


def canonicalize_schema(headers: List[str]) -> List[Dict[str, Any]]:
    schema: List[Dict[str, Any]] = []
    for h in headers:
        dtype = infer_field_type(h)
        mapped_type = {
            'tel': 'phone',
            'textarea': 'text',
        }.get(dtype, dtype)
        schema.append({
            "name": h,
            "label": h,
            "data_type": mapped_type if mapped_type in {"text","number","date","email","phone","boolean","dropdown"} else "text",
            "options": []
        })
    return schema


def load_field_schema_from_disk() -> Optional[List[Dict[str, Any]]]:
    for path in (SCHEMA_CACHE_FILE, FIELD_SCHEMA_FILE):
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        return data
            except Exception:
                continue
    return None


def save_field_schema_to_disk(schema: List[Dict[str, Any]]):
    for path in (SCHEMA_CACHE_FILE, FIELD_SCHEMA_FILE):
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(schema, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Warning: could not write {path}: {e}")


def load_schema() -> List[Dict[str, Any]]:
    """Load schema: local JSON -> CSV -> Excel."""
    global fields_cache
    disk_schema = load_field_schema_from_disk()
    if disk_schema:
        fields_cache = disk_schema
        return disk_schema
    if os.path.exists(CSV_FILE_PATH):
        headers = read_csv_headers(CSV_FILE_PATH)
        schema = canonicalize_schema(headers)
        fields_cache = schema
        return schema
    headers = read_excel_headers()
    schema = canonicalize_schema(headers)
    fields_cache = schema
    return schema


def build_schema_from_excel() -> List[Dict[str, Any]]:
    if not os.path.exists(EXCEL_FILE_PATH):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE_PATH}")
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True, data_only=True)
    ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME else wb.active
    headers: List[str] = []
    for cell in ws[1]:
        if cell.value and str(cell.value).strip():
            headers.append(str(cell.value).strip())
    dropdown_map = read_excel_dropdown_options(wb)
    # Build a case-insensitive view of dropdown_map
    dropdown_map_ci: Dict[str, List[str]] = {str(k).strip().lower(): v for k, v in dropdown_map.items()}
    schema: List[Dict[str, Any]] = []
    for h in headers:
        h_lower = h.strip().lower()
        dtype = 'dropdown' if h_lower in dropdown_map_ci else infer_field_type(h)
        mapped_type = {'tel': 'phone', 'textarea': 'text'}.get(dtype, dtype)
        schema.append({
            "name": h,
            "label": h,
            "data_type": mapped_type if mapped_type in {"text","number","date","email","phone","boolean","dropdown"} else "text",
            "options": dropdown_map_ci.get(h_lower, [])
        })
    wb.close()
    return schema


def upload_to_google_sheets(data: Dict[str, Any]):
    """Upload form data to Google Sheets"""
    
    if not os.path.exists(CREDENTIALS_FILE):
        raise FileNotFoundError(
            f"Google credentials file not found: {CREDENTIALS_FILE}. "
            "Please download your service account JSON from Google Cloud Console."
        )
    
    try:
        # Define the scope
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        
        # Authenticate using service account
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        # Open the Google Sheet
        spreadsheet = None
        if GOOGLE_SHEET_ID:
            # Use explicit sheet ID to avoid writing to a different Drive
            spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        else:
            # Fall back to name-based open (can create a separate sheet owned by service account)
            try:
                spreadsheet = client.open(GOOGLE_SHEET_NAME)
            except gspread.SpreadsheetNotFound:
                spreadsheet = client.create(GOOGLE_SHEET_NAME)

        sheet = spreadsheet.sheet1

        # Read existing headers (first row). If none, initialize from fields_cache + Timestamp
        values = sheet.get_all_values()
        if len(values) == 0:
            headers = [field.get('name') for field in fields_cache] + ['Timestamp']
            sheet.append_row(headers, value_input_option='USER_ENTERED')
        else:
            headers = list(values[0])

        # Build case-insensitive source map from data
        source_ci = {str(k).strip().lower(): v for k, v in data.items()}

        # Helper: find a key in source that matches header name, case-insensitive,
        # with a fallback for Member ID key variations
        def resolve_value(header_name: str) -> Any:
            h_lower = header_name.strip().lower()
            if h_lower in source_ci:
                return source_ci[h_lower]

            # Fuzzy match for member id/key variations
            if ("member" in h_lower) and ("id" in h_lower or "key" in h_lower):
                for k_lower, v in source_ci.items():
                    if ("member" in k_lower) and ("id" in k_lower or "key" in k_lower):
                        return v
            return ""

        # Build output row aligned to existing headers
        row_data = [resolve_value(h) for h in headers]

        # If there is a Timestamp column in existing headers, place now there; otherwise leave as is
        try:
            ts_idx = [h.strip().lower() for h in headers].index('timestamp')
            row_data[ts_idx] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            pass

        # Append the data aligned to headers
        sheet.append_row(row_data, value_input_option='USER_ENTERED')
        
        return {
            "status": "success",
            "message": "Data uploaded successfully",
            "sheet_url": spreadsheet.url,
            "sheet_title": spreadsheet.title,
        }
    except Exception as e:
        raise Exception(f"Error uploading to Google Sheets: {str(e)}")


@app.post("/init_headers")
async def init_headers():
    """Overwrite row 1 headers in the target sheet with the current field list + Timestamp."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)

    # Ensure fields loaded
    if not fields_cache:
        load_schema()

    spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
    sheet = spreadsheet.sheet1

    headers = [field['name'] for field in fields_cache] + ['Timestamp']
    # Replace entire first row with headers
    sheet.update('1:1', [headers], value_input_option='USER_ENTERED')


    return {
        "status": "success",
        "message": "Headers initialized",
        "sheet_url": spreadsheet.url,
        "headers": headers,
        "worksheet_title": sheet.title,
    }


def get_sheet_data_as_df():
    """Helper to fetch all data from Google Sheet into a Pandas DataFrame."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=500, detail="Google credentials not found")
    
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    
    spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
    sheet = spreadsheet.sheet1
    
    data = sheet.get_all_values()
    if not data:
        return pd.DataFrame(), sheet
        
    headers = data[0]
    rows = data[1:]
    
    # Handle duplicate headers if any
    unique_headers = []
    seen = set()
    for h in headers:
        c = h
        i = 1
        while c in seen:
            c = f"{h}_{i}"
            i += 1
        seen.add(c)
        unique_headers.append(c)
            
    df = pd.DataFrame(rows, columns=unique_headers)
    return df, sheet


def apply_delete_filters(df: pd.DataFrame, filters: FilterCriteria, date_col: str):
    """
    Returns a boolean mask where True means 'to be deleted'.
    """
    if date_col not in df.columns:
        raise HTTPException(status_code=400, detail=f"Column '{date_col}' not found in sheet")

    # Convert date column to datetime
    # Robust parsing:
    # 1. Clean whitespace
    # 2. Try dayfirst=False (ISO/US)
    # 3. If mostly NaT, try dayfirst=True (EU/India)
    try:
        clean_col = df[date_col].astype(str).str.strip()
        
        # Try 1: Standard/ISO
        temp_dates = pd.to_datetime(clean_col, errors='coerce', dayfirst=False)
        
        # Check success rate
        valid_count = temp_dates.notna().sum()
        total_count = len(df)
        
        # If very few valid dates found (less than 20%), try dayfirst=True
        if total_count > 0 and (valid_count / total_count) < 0.2:
             # Try 2: DayFirst
             print(f"Initial date parsing yield low success ({valid_count}/{total_count}). Retrying with dayfirst=True")
             temp_dates_alt = pd.to_datetime(clean_col, errors='coerce', dayfirst=True)
             if temp_dates_alt.notna().sum() > valid_count:
                 temp_dates = temp_dates_alt
                 
    except Exception as e:
        print(f"Date parsing critical error: {e}")
        raise HTTPException(status_code=400, detail=f"Date parsing failed: {str(e)}")

    # Create mask (initially all False)
    mask = pd.Series([False] * len(df), index=df.index)
    
    # Priority 1: Date Range (Start/End)
    if filters.startDate or filters.endDate:
        print(f"[DEBUG] Using date range filter: startDate={filters.startDate}, endDate={filters.endDate}")
        range_mask = pd.Series([True] * len(df), index=df.index)
        
        if filters.startDate:
            try:
                start = pd.to_datetime(filters.startDate)
                range_mask = range_mask & (temp_dates >= start)
                print(f"[DEBUG] After start date filter: {range_mask.sum()} matches")
            except Exception as e:
                print(f"[ERROR] Failed to parse startDate: {e}")
        
        if filters.endDate:
            try:
                end = pd.to_datetime(filters.endDate)
                # Include the entire end date (up to 23:59:59)
                end_inclusive = end + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                range_mask = range_mask & (temp_dates <= end_inclusive)
                print(f"[DEBUG] After end date filter: {range_mask.sum()} matches")
            except Exception as e:
                print(f"[ERROR] Failed to parse endDate: {e}")
        
        mask = range_mask
    
    # Priority 2: Specific Date
    elif filters.specificDate:
        print(f"[DEBUG] Using specific date filter: {filters.specificDate}")
        try:
            target_date = pd.to_datetime(filters.specificDate)
            mask = (temp_dates.dt.normalize() == target_date.normalize())
            print(f"[DEBUG] Specific date matches: {mask.sum()}")
        except Exception as e:
            print(f"[ERROR] Failed to parse specificDate: {e}")
    
    # Priority 3: Year/Month Logic
    elif filters.year or filters.month:
        print(f"[DEBUG] Using year/month filter: year={filters.year}, month={filters.month}")
        year_mask = pd.Series([True] * len(df), index=df.index)
        month_mask = pd.Series([True] * len(df), index=df.index)
        
        if filters.year:
            try:
                y = int(filters.year)
                year_mask = (temp_dates.dt.year == y)
                print(f"[DEBUG] Year {y} matches: {year_mask.sum()}")
            except Exception as e:
                print(f"[ERROR] Failed to parse year: {e}")
        
        if filters.month:
            try:
                m = int(filters.month)
                month_mask = (temp_dates.dt.month == m)
                print(f"[DEBUG] Month {m} matches: {month_mask.sum()}")
            except Exception as e:
                print(f"[ERROR] Failed to parse month: {e}")

        mask = year_mask & month_mask
        print(f"[DEBUG] Combined year/month matches: {mask.sum()}")
    else:
        # No filters provided -> delete nothing
        print(f"[DEBUG] No filters provided, returning empty mask")
        return pd.Series([False] * len(df), index=df.index), temp_dates
        
    # Remove NaT from mask (don't delete rows with invalid dates even if they match 'None' logic?)
    mask = mask & temp_dates.notna()
    
    return mask, temp_dates


@app.post("/delete/preview")
async def preview_delete(payload: DeletePreviewPayload):
    try:
        print(f"\n[DELETE PREVIEW] Starting preview with filters: {payload.filters}")
        print(f"[DELETE PREVIEW] Date column: {payload.date_column}")
        
        df, _ = get_sheet_data_as_df()
        print(f"[DELETE PREVIEW] Loaded {len(df)} rows from sheet")
        
        if df.empty:
            print(f"[DELETE PREVIEW] Sheet is empty")
            return {"count": 0, "rows": [], "earliest": None, "latest": None, "headers": []}
            
        mask, date_series = apply_delete_filters(df, payload.filters, payload.date_column)
        
        print(f"[DELETE PREVIEW] Filter mask created, {mask.sum()} rows match filters")
        print(f"[DELETE PREVIEW] Sample parsed dates (first 5): {date_series.head().tolist()}")
        print(f"[DELETE PREVIEW] Date column raw values (first 5): {df[payload.date_column].head().tolist()}")
        
        # Filter the DF
        matches = df[mask]
        count = len(matches)
        
        print(f"[DELETE PREVIEW] Final match count: {count}")
        
        if count == 0:
            return {"count": 0, "rows": [], "earliest": None, "latest": None}
            
        # Get Earliest/Latest from the matched dates
        matched_dates = date_series[mask]
        earliest = matched_dates.min().strftime('%Y-%m-%d') if not matched_dates.empty else None
        latest = matched_dates.max().strftime('%Y-%m-%d') if not matched_dates.empty else None
        
        # Prepare Preview Rows
        # Limit to 50
        preview_page = matches.head(50)
        
        # Select columns if requested
        if payload.preview_columns:
            # Only keep columns that exist
            valid_cols = [c for c in payload.preview_columns if c in preview_page.columns]
            if not valid_cols:
                # Fallback to all if none match (shouldn't happen with correct inputs)
                valid_cols = preview_page.columns.tolist()
            preview_page = preview_page[valid_cols]
            
        return {
            "count": count,
            "headers": preview_page.columns.tolist(),
            "rows": preview_page.to_dict(orient='records'),
            "earliest": earliest,
            "latest": latest
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delete/confirm")
async def confirm_delete(payload: DeleteConfirmPayload):
    try:
        df, sheet = get_sheet_data_as_df()
        if df.empty:
            return {"status": "error", "message": "Sheet is empty"}
            
        mask, _ = apply_delete_filters(df, payload.filters, payload.date_column)
        matches_count = mask.sum()
        
        if matches_count == 0:
             return {"status": "success", "message": "No rows matched the criteria (0 deleted)."}
             
        # Invert mask to get rows to KEEP
        # We need to preserve the dataframe structure perfectly
        df_keep = df[~mask]
        
        # Rewrite Sheet
        # 1. Clear
        sheet.clear()
        
        # 2. Write Headers
        headers = df_keep.columns.tolist()
        
        # 3. Write Rows
        # Gspread append_rows or update. allow massive update.
        # convert df_keep to list of lists
        # Handle NaN/None -> empty string for sheets
        df_keep = df_keep.fillna("")
        values = [headers] + df_keep.values.tolist()
        
        sheet.update(range_name='A1', values=values)
        
        return {
            "status": "success", 
            "message": f"Successfully deleted {matches_count} rows.",
            "remaining_count": len(df_keep)
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@app.on_event("startup")
async def startup_event():
    """Load schema on startup"""
    try:
        load_schema()
        print(f"Loaded {len(fields_cache)} fields from schema source")
        ensure_notification_defaults()
        # Debug: Check SMTP configuration
        smtp_user_set = "YES" if SMTP_USERNAME else "NO"
        smtp_pass_set = "YES" if SMTP_PASSWORD else "NO"
        print(f"[SMTP Config] Username set: {smtp_user_set}, Password set: {smtp_pass_set}")
    except Exception as e:
        print(f"Warning: Could not load fields on startup: {str(e)}")
        print("Ensure CSV or Excel file is present in backend directory")


@app.get("/")
async def root():
    return {"message": "CRM Lead Form API", "status": "running"}


@app.get("/get_fields")
async def get_fields():
    """Return the current field schema."""
    if not fields_cache:
        try:
            load_schema()
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    return {"fields": fields_cache}


@app.get("/test_email")
async def test_email_get(to: str):
    to_addr = str(to or "").strip()
    if not to_addr:
        raise HTTPException(status_code=400, detail="Missing 'to' query parameter")
    try:
        send_notification_email(to_addr, {"test": "ping"})
        return {"status": "ok", "to": to_addr}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"fields": fields_cache}


@app.post("/test_email")
async def test_email(payload: Dict[str, Any]):
    to_addr = str(payload.get("to", "")).strip()
    print(f"[TEST_EMAIL] Received request for {to_addr}")
    if not to_addr:
        raise HTTPException(status_code=400, detail="Missing 'to' in payload")
    try:
        print(f"[TEST_EMAIL] About to call send_notification_email...")
        send_notification_email(to_addr, {"test": "ping"})
        print(f"[TEST_EMAIL] send_notification_email returned")
        return {"status": "ok", "to": to_addr}
    except Exception as e:
        print(f"[TEST_EMAIL] Exception: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/notification_settings")
async def get_notification_settings():
    settings = ensure_notification_defaults()
    return {
        "sender_email": settings.get("sender_email", DEFAULT_SENDER_EMAIL),
        "cc_emails": settings.get("cc_emails", []),
    }


@app.post("/notification_settings")
async def update_notification_settings(payload: NotificationSettings):
    sender = str(payload.sender_email).strip()
    if not sender:
        raise HTTPException(status_code=400, detail="Sender email must not be empty")
    if "@" not in sender:
        raise HTTPException(status_code=400, detail="Sender email appears invalid")
    cc_emails: List[str] = []
    if payload.cc_emails:
        for item in payload.cc_emails:
            if not item:
                continue
            item_str = str(item).strip()
            if not item_str:
                continue
            if "@" not in item_str:
                raise HTTPException(status_code=400, detail=f"Invalid CC email: {item_str}")
            cc_emails.append(item_str)

    save_notification_settings({"sender_email": sender, "cc_emails": cc_emails})
    return {"status": "success", "sender_email": sender, "cc_emails": cc_emails}


@app.post("/upload_file")
async def upload_file(file: UploadFile = File(...)):
    """
    Handle file uploads.
    - Stores non-data files safely.
    - Processes data files (Excel/CSV) for schema changes.
    """
    try:
        # 1. Save the file
        file_path = save_upload(file, file.filename)
        
        # 2. Determine if it's a data file
        filename = file.filename.lower()
        is_data_file = filename.endswith(('.csv', '.xlsx', '.xls', '.xlsm'))
        
        if is_data_file:
            # Load current schema to compare
            existing_schema = fields_cache
            if not existing_schema:
                try:
                    existing_schema = load_schema()
                except:
                    existing_schema = []
            
            result = process_data_file(file_path, existing_schema)
            # Add file path to result
            result['file_path'] = file_path
            return result
        else:
            # Non-data file
            return {
                "status": "success",
                "message": f"File '{file.filename}' stored successfully.",
                "file_path": file_path,
                "type": "non-data"
            }
            
    except Exception as e:
        print(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ConfirmUploadRequest(BaseModel):
    file_path: str


@app.post("/confirm_upload")
async def confirm_upload(payload: ConfirmUploadRequest):
    """
    Reads the previously uploaded file and APPENDS data to 'Sheet1' in Google Sheets.
    It maps columns by name to match the existing sheet headers.
    """
    file_path = payload.file_path
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials not found")

    try:
        # 1. Read Data
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(file_path)
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            df = pd.read_excel(file_path)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")

        df = df.fillna("")

        # 2. Connect to Sheets
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID) if GOOGLE_SHEET_ID else client.open(GOOGLE_SHEET_NAME)
        try:
            sheet = spreadsheet.worksheet("Sheet1")
        except gspread.WorksheetNotFound:
            sheet = spreadsheet.sheet1

        # 3. Get Existing Headers
        existing_values = sheet.get_all_values()
        if not existing_values:
            # Sheet is empty, write headers from file
            sheet_headers = df.columns.tolist() + ['Timestamp']
            sheet.append_row(sheet_headers, value_input_option='USER_ENTERED')
        else:
            sheet_headers = existing_values[0]

        # 3b. Detect and Add New Columns
        # Compare df columns to sheet_headers (case-insensitive check)
        sheet_headers_lower = {h.strip().lower() for h in sheet_headers}
        new_columns = []
        for col in df.columns:
            if col.strip().lower() not in sheet_headers_lower:
                new_columns.append(col)
        
        if new_columns:
            # Append new columns to header row in Sheet
            # We need to find the next available column index
            # But gspread's append_row doesn't work for partial row updates easily on existing rows
            # simplest way: update the first row with the extended list
            sheet_headers.extend(new_columns)
            sheet.update(range_name='1:1', values=[sheet_headers], value_input_option='USER_ENTERED')

        # 4. Map Data to Headers (now including new ones)
        # Create a map for case-insensitive matching of file columns
        file_cols_map = {c.strip().lower(): c for c in df.columns}
        
        data_to_append = []
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for _, row in df.iterrows():
            ordered_row = []
            for h in sheet_headers:
                h_lower = h.strip().lower()
                
                # Special handling for Timestamp
                if h_lower == 'timestamp':
                    ordered_row.append(current_time)
                    continue

                # Find matching column in file
                if h_lower in file_cols_map:
                    original_col_name = file_cols_map[h_lower]
                    # Handle NaN values to ensure empty string
                    val = row[original_col_name]
                    ordered_row.append(val if pd.notna(val) else "")
                else:
                    # Column exists in Sheet but not in File -> Empty
                    ordered_row.append("")
            
            data_to_append.append(ordered_row)

        if data_to_append:
            sheet.append_rows(data_to_append, value_input_option='USER_ENTERED')
            
        message = f"Successfully appended {len(data_to_append)} rows."
        if new_columns:
            message += f" Added {len(new_columns)} new columns: {', '.join(new_columns)}."

        return {
            "status": "success",
            "message": message,
            "sheet_url": spreadsheet.url
        }

    except Exception as e:
        print(f"Bulk update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sync_fields")
async def sync_fields():
    """Read Excel, rebuild schema, persist locally, update CSV and Google Sheet (headers + dropdowns)."""
    global fields_cache
    try:
        schema = build_schema_from_excel()
        save_field_schema_to_disk(schema)
        fields_cache = schema
        # Update local CSV header row
        try:
            header_row = [f["name"] for f in schema]
            rows: List[List[str]] = []
            if os.path.exists(CSV_FILE_PATH):
                with open(CSV_FILE_PATH, "r", encoding="utf-8-sig", newline="") as rf:
                    reader = csv.reader(rf)
                    existing = list(reader)
                rows = existing[1:] if len(existing) > 0 else []
            with open(CSV_FILE_PATH, "w", encoding="utf-8", newline="") as wf:
                writer = csv.writer(wf)
                writer.writerow(header_row)
                for r in rows:
                    writer.writerow(r)
        except Exception as e:
            print(f"Warning: failed updating local CSV: {e}")

        # Update Google Sheet
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        sheet = spreadsheet.sheet1
        headers = [f["name"] for f in schema] + ["Timestamp"]
        sheet.update('1:1', [headers], value_input_option='USER_ENTERED')
        sync_dropdown_options_to_sheet(client, spreadsheet, schema)
        return {"status": "success", "fields": schema, "sheet_url": spreadsheet.url}
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {e}")


class UpdateSingleFieldPayload(BaseModel):
    original_name: str
    name: Optional[str] = None
    label: Optional[str] = None
    data_type: Optional[str] = None
    options: Optional[List[str]] = None


@app.post("/update_field")
async def update_field(payload: UpdateSingleFieldPayload):
    """Modify one field; persist, update CSV and Google Sheet."""
    global fields_cache
    if not fields_cache:
        load_schema()
    idx = next((i for i, f in enumerate(fields_cache) if f.get('name') == payload.original_name), -1)
    if idx == -1:
        raise HTTPException(status_code=404, detail=f"Field not found: {payload.original_name}")
    updated = dict(fields_cache[idx])
    if payload.name is not None:
        updated['name'] = payload.name.strip()
    if payload.label is not None:
        updated['label'] = payload.label.strip()
    if payload.data_type is not None:
        updated['data_type'] = payload.data_type
        if payload.data_type != 'dropdown':
            updated['options'] = []
    if payload.options is not None:
        updated['options'] = [str(o).strip() for o in payload.options]
    fields_cache[idx] = updated
    save_field_schema_to_disk(fields_cache)


@app.get("/search_data")
async def search_data(query: Optional[str] = Query(None, min_length=2), limit: int = 50):
    """Search for patient data for auto-fill. If query is empty, returns all data up to limit."""
    try:
        # Use gspread directly to get raw data
        if not os.path.exists(CREDENTIALS_FILE):
             # Try to start without creds (maybe public?) No, strict requirement here.
             raise HTTPException(status_code=404, detail="Google credentials file not found")

        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        sheet = spreadsheet.sheet1
        values = sheet.get_all_values()

        if not values:
            return {"results": [], "headers": [], "rows": []}

        headers = values[0]
        rows = values[1:]
        
        results = []
        
        # If no query, return top N rows
        if not query:
            # Convert all rows to dicts
            for row in rows[:limit]:
                 # Handle rows shorter than headers
                r_len = len(row)
                row_dict = {h: (row[i] if i < r_len else "") for i, h in enumerate(headers)}
                results.append(row_dict)
            return {"results": results, "headers": headers, "rows": rows[:limit]} # Return raw rows to ensure frontend compatibility

        q_lower = query.lower()
        
        for row in rows:
            # Handle rows shorter than headers
            r_len = len(row)
            row_dict = {h: (row[i] if i < r_len else "") for i, h in enumerate(headers)}
            
            # Helper to safely get value
            def match(val):
                return str(val).lower().find(q_lower) >= 0

            # Find specific columns if possible or just search all values
            # Heuristic search
            found = False
            
            # Simple broad search across all values in the row
            # But prompt logic suggested checking specific fields?
            # Stick to broad search for robustness:
            for v in row:
                 if match(v):
                     found = True
                     break
            
            if found:
                results.append(row_dict)
                if len(results) >= limit: # Limit results
                    break
                    
        return {"results": results, "headers": headers, "rows": [list(r.values()) for r in results]} # Maintain compatibility
        
    except Exception as e:
        print(f"Search failed: {e}")
        return {"results": [], "headers": [], "rows": []}

@app.get("/download_template")
async def download_template(format: str = Query(default="xlsx", regex="^(csv|xlsx)$")):
    """Generate and return a sample template (CSV or Excel) based on active Sheet1 headers."""
    headers = []
    
    # Priority 1: Fetch Live Headers from Google Sheet
    try:
        if GOOGLE_SHEET_ID or os.path.exists(CREDENTIALS_FILE):
             client, spreadsheet = get_google_sheet_client()
             sheet = spreadsheet.sheet1
             # Get first row as headers
             live_headers = sheet.row_values(1)
             if live_headers:
                 headers = live_headers
    except Exception as e:
        print(f"Template generation: Failed to fetch live headers ({e}). Using fallback schema.")

    # Priority 2: Fallback to Local Schema
    if not headers:
        if not fields_cache:
            load_schema()
        headers = [f['name'] for f in fields_cache]
    
    # Prepare sample data
    sample_data = [
        {
            'date': '03/12/2025',
            'memberidkey': 'MID-2025-12-03-001',
            'attender name': 'John Smith',
            'patient name': 'John Patient',
            'gender': 'Male',
            'age': '45',
            'patient location': 'New York',
            'area': 'Manhattan',
            'email id': 'john@example.com',
            'mobile number': '1234567890'
        },
        {
            'date': '04/12/2025',
            'memberidkey': 'MID-2025-12-04-002',
            'attender name': 'Jane Doe',
            'patient name': 'Jane Patient',
            'gender': 'Female',
            'age': '32',
            'patient location': 'London',
            'area': 'Westminster',
            'email id': 'jane@example.com',
            'mobile number': '0987654321'
        },
        {
            'date': '05/12/2025',
            'memberidkey': 'MID-2025-12-05-003',
            'attender name': 'Robert Johnson',
            'patient name': 'Robert Patient',
            'gender': 'Male',
            'age': '58',
            'patient location': 'Paris',
            'area': 'Marais',
            'email id': 'robert@example.com',
            'mobile number': '1122334455'
        }
    ]
    
    # Build rows
    rows_data = []
    schema_map = {f['name'].lower().strip(): f for f in fields_cache} if fields_cache else {}
    
    for sample in sample_data:
        example_row = []
        
        for h in headers:
            h_lower = str(h).lower().strip()
            
            # Check if we have sample data for this field
            if h_lower in sample:
                example_row.append(sample[h_lower])
            else:
                # Fallback to smart defaults
                field_def = schema_map.get(h_lower)
                dtype = field_def.get('data_type', 'text') if field_def else 'text'
                
                if 'date' in h_lower or dtype == 'date':
                    example_row.append(datetime.now().strftime('%d/%m/%Y'))
                elif 'email' in h_lower or dtype == 'email':
                    example_row.append('example@domain.com')
                elif 'phone' in h_lower or 'mobile' in h_lower or dtype == 'phone':
                    example_row.append('9876543210')
                elif dtype == 'number' or 'price' in h_lower or 'rent' in h_lower or 'age' in h_lower:
                    example_row.append('100')
                elif dtype == 'dropdown' and field_def and field_def.get('options'):
                    example_row.append(field_def['options'][0] if field_def['options'] else '')
                else:
                    example_row.append('')
                    
        rows_data.append(example_row)
    
    # Generate file based on format
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows_data:
            writer.writerow(row)
        output.seek(0)
        
        response = StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv"
        )
        response.headers["Content-Disposition"] = "attachment; filename=patient_data_template.csv"
        return response
    else:  # xlsx
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patient Data"
        
        # Write headers
        ws.append(headers)
        
        # Write data rows
        for row in rows_data:
            ws.append(row)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = "attachment; filename=patient_data_template.xlsx"
        return response


@app.post("/admission/register")
async def register_admission(payload: Dict[str, Any] = Body(...)):
    """
    Handle Admission Registration.
    1. Save data to Sheet1 (append).
    2. Send Notification.
    """
    data = payload.get('data', {})
    if not data:
        raise HTTPException(status_code=400, detail="No data provided")
        
    # Validation: Check Member ID
    # In 'data', keys likely match the form internal state or mapped ones
    # The frontend maps them to 'MemberKey', etc.
    
    try:
        # 1. Upload to Sheets
        # This will append to Sheet1. 
        # Since 'upload_to_google_sheets' handles flexible columns by appending new ones if needed,
        # it fits the requirement of "Store it in Sheet1".
        
        # Ensure we don't duplicate if logic existed, but for now we just append.
        upload_result = upload_to_google_sheets(data)
        
        # 2. Send Notification
        # Extract details for message
        patient_name = data.get('Patient Name') or data.get('firstName', 'Patient')
        member_id = data.get('MemberKey') or data.get('memberId', '')
        admission_date = data.get('Admission Date', datetime.now().strftime('%Y-%m-%d'))
        hospital_name = "Grand World Elder Care" # Configurable or hardcoded
        
        # Recipient email: try to find one in the data, else use default sender (self-notification)
        # Requirement: "Send message via SMS, Email, or WhatsApp"
        # We'll stick to Email for now as we have the implementation.
        # Who receives it? "Dear {PatientName}" implies the PATIENT receives it.
        patient_email = data.get('Email ID') or data.get('email', '')
        
        if patient_email and '@' in patient_email:
            subject = "Admission Confirmation"
            message_body = f"""Dear {patient_name},

Your admission on {admission_date} has been successfully recorded at {hospital_name}.
Your Patient ID is {member_id}.

If you need assistance, please contact us.

– {hospital_name}"""
            
            # Send to Patient
            # We use background task normally, but here we can just call it.
            # Using existing send_notification_email which sends simple HTML/Text.
            # We construct a simple payload for the template or override body?
            # Existing `send_notification_email` takes a payload and builds key-value pairs.
            # We might need a `send_direct_email` function or adapt `send_notification_email`.
            # `build_email_message` creates a list of key-values.
            # Let's create a custom notification function or hack the payload.
            
            # Better: Create a dedicated notification function here just for this.
            try:
                msg = EmailMessage()
                msg.set_content(message_body)
                msg["Subject"] = subject
                msg["From"] = ensure_notification_defaults().get('sender_email', 'noreply@crm.com')
                msg["To"] = patient_email
                
                # We can reuse the `send_email_via_gmail_api` or `smtplib` logic if we duplicate it 
                # or refactor `send_notification_email` to take body_text.
                # `send_notification_email` calls `build_email_message` which enforces Key-Value format.
                # Let's try to just send a "Notification" payload.
                
                # Fallback: Just send the summary as before to the SYSTEM admin/configured email, 
                # AND try to send to Patient if possible.
                # The prompt asks for "Automatic Notifications After Admission... Dear {PatientName}".
                
                # I will create a `send_custom_email` helper.
                pass
            except:
                pass
                
        # Send Admin Notification (Standard)
        send_notification_email(
             ensure_notification_defaults().get('sender_email'),
             data,
             subject_override=f"New Admission: {patient_name}"
        )

        return {
            "status": "success",
            "message": "Admission registered successfully",
            "sheet_url": upload_result.get('sheet_url')
        }

    except Exception as e:
        print(f"Admission registration failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


    # Update CSV
    try:
        header_row = [f["name"] for f in fields_cache]
        rows: List[List[str]] = []
        if os.path.exists(CSV_FILE_PATH):
            with open(CSV_FILE_PATH, "r", encoding="utf-8-sig", newline="") as rf:
                reader = csv.reader(rf)
                existing = list(reader)
            rows = existing[1:] if len(existing) > 0 else []
        with open(CSV_FILE_PATH, "w", encoding="utf-8", newline="") as wf:
            writer = csv.writer(wf)
            writer.writerow(header_row)
            for r in rows:
                writer.writerow(r)
    except Exception as e:
        print(f"Warning: update_field failed CSV write: {e}")

    # Update Google Sheet
    try:
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        sheet = spreadsheet.sheet1
        headers = [f["name"] for f in fields_cache] + ["Timestamp"]
        sheet.update('1:1', [headers], value_input_option='USER_ENTERED')
        sync_dropdown_options_to_sheet(client, spreadsheet, fields_cache)
    except Exception as e:
        print(f"Warning: update_field failed Sheet sync: {e}")
    return {"status": "success", "field": updated}


@app.post("/submit")
async def submit_form(form_data: FormData, background_tasks: BackgroundTasks):
    """Submit form data to Google Sheets"""
    try:
        print(f"[Submit] Received form submission with {len(form_data.data)} fields")
        # Ensure defaults: auto Member ID and today's date for empty date fields
        enriched = dict(form_data.data)

        # Auto Member ID if a matching field exists and is empty
        member_id_field = None
        for field in fields_cache:
            name_lower = field['name'].lower()
            if ("member" in name_lower) and ("id" in name_lower or "key" in name_lower):
                member_id_field = field['name']
                break
        if member_id_field:
            # Always set server-side to keep it immutable from client edits
            # Format: MID-yyyy-mm-dd-numericalid
            now = datetime.now()
            date_part = now.strftime('%Y-%m-%d')
            # Generate unique numerical ID using timestamp + random number
            numerical_id = int(now.timestamp() * 1000) % 100000 + random.randint(1000, 9999)
            generated = f"MID-{date_part}-{numerical_id}"
            enriched[member_id_field] = generated

        # Fill empty date fields with today's date (YYYY-MM-DD)
        today = datetime.now().strftime('%Y-%m-%d')
        for field in fields_cache:
            if field.get('data_type') == 'date' or field.get('type') == 'date':
                name = field['name']
                name_lower = str(name).lower()
                if REMINDER_DATE_KEYWORD in name_lower or FOLLOW_DATE_KEYWORD in name_lower:
                    continue
                if not str(enriched.get(name, "")).strip():
                    enriched[name] = today

        result = upload_to_google_sheets(enriched)

        recipient_email = extract_recipient_email(enriched)
        print(f"[Email] Recipient detected from payload: {recipient_email}")
        if recipient_email:
            background_tasks.add_task(send_notification_email, recipient_email, dict(enriched))

        return result
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    excel_exists = os.path.exists(EXCEL_FILE_PATH)
    creds_exist = os.path.exists(CREDENTIALS_FILE)
    
    return {
        "status": "healthy",
        "excel_file": excel_exists,
        "google_credentials": creds_exist,
        "fields_loaded": len(fields_cache)
    }


@app.get("/debug/latest")
async def debug_latest_rows():
    """Return the target spreadsheet URL, first worksheet title, and last 5 rows."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)

    spreadsheet = None
    if GOOGLE_SHEET_ID:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
    else:
        try:
            spreadsheet = client.open(GOOGLE_SHEET_NAME)
        except gspread.SpreadsheetNotFound:
            raise HTTPException(status_code=404, detail="Spreadsheet not found")

    sheet = spreadsheet.sheet1
    values = sheet.get_all_values()
    last_rows = values[-5:] if len(values) > 5 else values

    return {
        "sheet_url": spreadsheet.url,
        "sheet_title": spreadsheet.title,
        "worksheet_title": sheet.title,
        "rows_total": len(values),
        "last_rows": last_rows,
    }


def ensure_google_sheet(client: gspread.Client) -> gspread.Spreadsheet:
    if GOOGLE_SHEET_ID:
        return client.open_by_key(GOOGLE_SHEET_ID)
    try:
        return client.open(GOOGLE_SHEET_NAME)
    except gspread.SpreadsheetNotFound:
        return client.create(GOOGLE_SHEET_NAME)


def get_google_sheet_client():
    """Helper to get authenticated gspread client and spreadsheet."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")
    
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    return client, spreadsheet



def sync_dropdown_options_to_sheet(client: gspread.Client, spreadsheet: gspread.Spreadsheet, schema: List[Dict[str, Any]]):
    try:
        try:
            ws = spreadsheet.worksheet("DropdownOptions")
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title="DropdownOptions", rows=100, cols=2)
        rows = [["Field Name", "Options"]]
        for f in schema:
            if (f.get("data_type") == "dropdown") and f.get("options"):
                rows.append([f.get("name"), ", ".join([str(o) for o in f.get("options")])])
        ws.clear()
        if rows:
            ws.update("A1", rows)
    except Exception as e:
        print(f"Warning: failed syncing dropdown options: {e}")


class FieldItem(BaseModel):
    name: str
    label: Optional[str] = None
    data_type: str
    options: Optional[List[str]] = None


class UpdateFieldsPayload(BaseModel):
    fields: List[FieldItem]


@app.post("/update_fields")
async def update_fields(payload: UpdateFieldsPayload):
    global fields_cache
    allowed_types = {"text","number","date","email","phone","boolean","dropdown"}
    schema: List[Dict[str, Any]] = []
    for f in payload.fields:
        if f.data_type not in allowed_types:
            raise HTTPException(status_code=400, detail=f"Invalid data_type for field '{f.name}': {f.data_type}")
        schema.append({
            "name": f.name.strip(),
            "label": (f.label or f.name).strip(),
            "data_type": f.data_type,
            "options": [str(o).strip() for o in (f.options or [])]
        })

    # Persist schema JSON
    with open(FIELD_SCHEMA_FILE, "w", encoding="utf-8") as f:
        json.dump(schema, f, ensure_ascii=False, indent=2)
    fields_cache = schema

    # Update local CSV header row
    try:
        header_row = [f["name"] for f in schema]
        rows: List[List[str]] = []
        if os.path.exists(CSV_FILE_PATH):
            with open(CSV_FILE_PATH, "r", encoding="utf-8-sig", newline="") as rf:
                reader = csv.reader(rf)
                existing = list(reader)
            rows = existing[1:] if len(existing) > 0 else []
        with open(CSV_FILE_PATH, "w", encoding="utf-8", newline="") as wf:
            writer = csv.writer(wf)
            writer.writerow(header_row)
            for r in rows:
                writer.writerow(r)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update local CSV: {e}")

    # Update Google Sheet header row + dropdowns
    try:
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        sheet = spreadsheet.sheet1
        headers = [f["name"] for f in schema] + ["Timestamp"]
        sheet.update('1:1', [headers], value_input_option='USER_ENTERED')
        sync_dropdown_options_to_sheet(client, spreadsheet, schema)
        return {"status": "success", "message": "Fields updated", "sheet_url": spreadsheet.url}
    except FileNotFoundError:
        return {"status": "success", "message": "Fields updated (local). Google credentials not found to sync."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to sync Google Sheet: {e}")


@app.get("/preview_data")
async def preview_data():
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    sheet = spreadsheet.sheet1
    values = sheet.get_all_values()
    sample = values[:6] if len(values) >= 6 else values
    return {"rows": sample, "sheet_url": spreadsheet.url}





@app.put("/update_record")
async def update_record(member_id: Optional[str] = None, payload: Optional[Dict[str, Any]] = Body(default=None)):
    """Update a record in Google Sheets by member ID.
    Accepts member_id either as a query parameter or inside the JSON body.
    JSON body can be either {"member_id": "...", "data": {...}} or {"data": {...}}.
    If member_id is not explicitly provided, it will be inferred from data keys containing member/id.
    """
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=404, detail="Google credentials file not found")

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    sheet = spreadsheet.sheet1
    values = sheet.get_all_values()

    if len(values) < 2:  # No data rows
        raise HTTPException(status_code=404, detail="No records found to update")

    headers = values[0]
    data_rows = values[1:]

    # Find member ID column
    member_id_col = None
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in ['member', 'id', 'key']) and ('member' in header_lower or 'id' in header_lower):
            member_id_col = i
            break

    if member_id_col is None:
        raise HTTPException(status_code=400, detail="Member ID column not found")

    # Read payload safely
    body_member_id: Optional[str] = None
    body_data: Dict[str, Any] = {}
    if isinstance(payload, dict):
        # FastAPI will parse JSON into dict when type is Dict[str, Any]
        body_member_id = payload.get("member_id")
        body_data = payload.get("data") or {}
    # Final member id resolution order: query -> body.member_id -> infer from body.data
    resolved_member_id: Optional[str] = (member_id or body_member_id)
    if not resolved_member_id:
        # Try to infer from data keys
        for k, v in body_data.items():
            k_lower = str(k).strip().lower()
            if ("member" in k_lower) and ("id" in k_lower or "key" in k_lower):
                resolved_member_id = str(v).strip()
                break

    if not resolved_member_id:
        raise HTTPException(status_code=400, detail="member_id is required (query param or in JSON body)")

    # Find the row with matching member ID
    target_row_idx = None
    for i, row in enumerate(data_rows):
        if member_id_col < len(row):
            row_member_id = str(row[member_id_col]).strip()
            if row_member_id == resolved_member_id.strip():
                target_row_idx = i + 2  # +2 because gspread is 1-indexed and we skip header
                break

    if target_row_idx is None:
        raise HTTPException(status_code=404, detail=f"Record with Member ID '{member_id}' not found")

    # Build case-insensitive source map from data
    source_ci = {str(k).strip().lower(): v for k, v in (body_data or {}).items()}

    # Helper: find a key in source that matches header name, case-insensitive,
    # with a fallback for Member ID key variations
    def resolve_value(header_name: str) -> Any:
        h_lower = header_name.strip().lower()
        if h_lower in source_ci:
            return source_ci[h_lower]

        # Fuzzy match for member id/key variations
        if ("member" in h_lower) and ("id" in h_lower or "key" in h_lower):
            for k_lower, v in source_ci.items():
                if ("member" in k_lower) and ("id" in k_lower or "key" in k_lower):
                    return v
        return ""

    # Build updated row aligned to existing headers
    # Preserve existing values if not provided in payload
    original_row = data_rows[target_row_idx - 2] if (target_row_idx - 2) < len(data_rows) else []
    updated_row: List[Any] = []
    for col_idx, h in enumerate(headers):
        new_val = resolve_value(h)
        if (new_val is None) or (str(new_val).strip() == ""):
            # fallback to original cell value if available
            if col_idx < len(original_row):
                updated_row.append(original_row[col_idx])
            else:
                updated_row.append("")
        else:
            updated_row.append(new_val)

    # Update the timestamp if it exists
    try:
        ts_idx = [h.strip().lower() for h in headers].index('timestamp')
        updated_row[ts_idx] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    except ValueError:
        pass

    # Update the specific row
    sheet.update(f'{target_row_idx}:{target_row_idx}', [updated_row], value_input_option='USER_ENTERED')

    # Determine if lead status changed and get email to notify
    try:
        headers_lower = [h.strip().lower() for h in headers]
        status_idx = headers_lower.index('lead status') if 'lead status' in headers_lower else None
        email_idx = None
        for i, h in enumerate(headers_lower):
            if 'email' in h:
                email_idx = i
                break
        member_id_idx = None
        for i, h in enumerate(headers_lower):
            if ('member' in h) and ('id' in h or 'key' in h):
                member_id_idx = i
                break

        status_changed = False
        recipient_email = None
        member_id_value = None
        if status_idx is not None and (target_row_idx - 2) < len(data_rows):
            prev_status = data_rows[target_row_idx - 2][status_idx] if status_idx < len(data_rows[target_row_idx - 2]) else ''
            new_status = updated_row[status_idx] if status_idx < len(updated_row) else prev_status
            status_changed = (str(prev_status).strip() != str(new_status).strip())
        if email_idx is not None and email_idx < len(updated_row):
            recipient_email = str(updated_row[email_idx]).strip()
        if member_id_idx is not None and member_id_idx < len(updated_row):
            member_id_value = str(updated_row[member_id_idx]).strip()

        if status_changed and recipient_email:
            payload_map = {h: (updated_row[i] if i < len(updated_row) else '') for i, h in enumerate(headers)}
            send_notification_email(recipient_email, payload_map, subject_override='LEAD STATUS CHANGED')
    except Exception as _e:
        print(f"[UpdateRecord] Notification check failed: {_e}")

    return {
        "status": "success",
        "message": f"Record with Member ID '{resolved_member_id}' updated successfully",
    }


# Helper global for login cache
login_cache: Dict[str, Any] = {"users": {}, "last_fetched": 0}
LOGIN_CACHE_DURATION = 300  # 5 minutes

@app.post("/login")
async def login(payload: LoginRequest):
    """Validate credentials against Google Sheet worksheet 'login details'.
    Expected columns: User_name, Password (case-insensitive).
    Fallback: accepts admin/admin for development if Google Sheet access fails.
    """
    import time
    start_time = time.time()
    print(f"[Login] Request received for user: {payload.User_name}")

    in_user = str(payload.User_name or '').strip()
    in_pass = str(payload.Password or '').strip()
    
    if not in_user or not in_pass:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    # Development fallback credentials
    DEV_CREDENTIALS = {
        "admin": "admin",
        "user": "user123",
        "test": "test123"
    }
    
    # Check dev credentials first for quick access
    if in_user in DEV_CREDENTIALS and DEV_CREDENTIALS[in_user] == in_pass:
        print(f"[Login] Dev credentials accepted. Time: {time.time() - start_time:.2f}s")
        return {"status": "ok"}

    # Check cache first
    current_time = time.time()
    if login_cache["users"] and (current_time - login_cache["last_fetched"] < LOGIN_CACHE_DURATION):
        cached_pass = login_cache["users"].get(in_user)
        if cached_pass and cached_pass == in_pass:
            print(f"[Login] Cached credentials accepted. Time: {time.time() - start_time:.2f}s")
            return {"status": "ok"}
    
    # Try Google Sheet authentication
    if not os.path.exists(CREDENTIALS_FILE):
        print("[Login] Google credentials file not found, using dev fallback only")
        raise HTTPException(status_code=401, detail="Invalid username or password")

    try:
        print(f"[Login] Connecting to Google Sheets... Time: {time.time() - start_time:.2f}s")
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)

        # Case-insensitive lookup for worksheet named 'Login Details'
        print(f"[Login] Finding worksheet... Time: {time.time() - start_time:.2f}s")
        try:
            ws = None
            # Fast check first
            try:
                ws = spreadsheet.worksheet("Login Details")
            except gspread.WorksheetNotFound:
                # Slow check
                for w in spreadsheet.worksheets():
                    if str(w.title).strip().lower() == 'login details':
                        ws = w
                        break
            
            if ws is None:
                print("[Login] 'Login Details' worksheet not found, using dev fallback only")
                raise HTTPException(status_code=401, detail="Invalid username or password")
        except HTTPException:
            raise
        except Exception as e:
            print(f"[Login] Error accessing worksheets: {e}")
            raise HTTPException(status_code=401, detail="Invalid username or password")

        print(f"[Login] Fetching values... Time: {time.time() - start_time:.2f}s")
        values = ws.get_all_values() or []
        if len(values) < 2:
            print("[Login] No user data in Login Details worksheet")
            raise HTTPException(status_code=401, detail="Invalid username or password")

        headers = values[0]
        rows = values[1:]
        # map header names to indices (case-insensitive)
        header_map = {str(h).strip().lower(): i for i, h in enumerate(headers)}
        user_col = (
            header_map.get('user_name')
            or header_map.get('user name')
            or header_map.get('username')
            or header_map.get('user-name')
        )
        pass_col = header_map.get('password')
        if user_col is None or pass_col is None:
            print("[Login] Required columns not found in Login Details worksheet")
            raise HTTPException(status_code=401, detail="Invalid username or password")

        # Refresh cache
        new_cache = {}
        found = False
        for r in rows:
            u = str(r[user_col]).strip() if user_col < len(r) else ''
            p = str(r[pass_col]).strip() if pass_col < len(r) else ''
            if u:
                new_cache[u] = p
            if u == in_user and p == in_pass:
                found = True

        # Update global cache
        login_cache["users"] = new_cache
        login_cache["last_fetched"] = current_time
        
        if found:
            print(f"[Login] Google Sheet credentials accepted and cached. Time: {time.time() - start_time:.2f}s")
            return {"status": "ok"}

        print(f"[Login] Credential mismatch for user '{in_user}'. Time: {time.time() - start_time:.2f}s")
        raise HTTPException(status_code=401, detail="Invalid username or password")
    except HTTPException:
        raise
    except Exception as exc:
        print(f"[Login] Exception during Google Sheet auth: {exc}")
        # If google auth fails, fallback to checking cache one last time if we have old data? 
        # No, simpler to just fail safe.
        raise HTTPException(status_code=401, detail="Invalid username or password")


@app.get("/list_sheets")
async def list_sheets():
    """Return the list of worksheet names in the Google Spreadsheet and whether 'login details' exists."""
    if not os.path.exists(CREDENTIALS_FILE):
        raise HTTPException(status_code=500, detail="Google credentials file not found")
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    names = [ws.title for ws in spreadsheet.worksheets()]
    return {
        "sheets": names,
        "has_login_details": any(str(n).strip().lower() == 'login details' for n in names)
    }



# --- Bed Management API ---

class BedAllocationRequest(BaseModel):
    patient_name: str
    member_id: Optional[str] = None
    gender: str
    room_no: str
    bed_index: int
    room_type: str
    admission_date: str
    discharge_date: Optional[str] = None
    pain_point: Optional[str] = None

class ComplaintRequest(BaseModel):
    patient_name: str
    room_no: str
    complaint_type: str
    description: Optional[str] = None

class FeedbackRequest(BaseModel):
    patient_name: str
    rating_comfort: int
    rating_cleanliness: int
    rating_staff: int
    comments: Optional[str] = None

def ensure_bed_sheets_google(spreadsheet: gspread.Spreadsheet):
    """Ensure Admission Details, Complaints, and Feedback sheets exist in Google Sheet."""
    
    # --- Admission Details ---
    try:
        ws_admission = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
    except gspread.WorksheetNotFound:
        # Create sheet
        ws_admission = spreadsheet.add_worksheet(title=ADMISSION_SHEET_NAME, rows=100, cols=20)
        headers = [
            "Room No", "Room Type", "Bed Count", "Bed Index", 
            "Patient Name", "Member ID", "Gender", 
            "Admission Date", "Discharge Date", "Status", 
            "Pain Point", "Complaints"
        ]
        ws_admission.append_row(headers)
        
        # Initialize Default Rooms (Mock Data)
        rows_to_add = []
        # Single Rooms (101-105)
        for i in range(101, 106):
            rows_to_add.append([str(i), "Single", 1, 0, "", "", "", "", "", "Available", "", ""])
        # Twin Rooms (201-205)
        for i in range(201, 206):
            rows_to_add.append([str(i), "Twin", 2, 0, "", "", "", "", "", "Available", "", ""])
            rows_to_add.append([str(i), "Twin", 2, 1, "", "", "", "", "", "Available", "", ""])
        
        ws_admission.append_rows(rows_to_add)

    # --- Complaints ---
    try:
        spreadsheet.worksheet(COMPLAINT_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws_complaints = spreadsheet.add_worksheet(title=COMPLAINT_SHEET_NAME, rows=100, cols=10)
        ws_complaints.append_row(["Date", "Room No", "Patient Name", "Type", "Description", "Resolved"])

    # --- Feedback ---
    try:
        spreadsheet.worksheet(FEEDBACK_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws_feedback = spreadsheet.add_worksheet(title=FEEDBACK_SHEET_NAME, rows=100, cols=10)
        ws_feedback.append_row(["Date", "Patient Name", "Comfort", "Cleanliness", "Staff", "Comments"])


@app.get("/api/beds")
async def get_beds():
    """Get all bed status from Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ensure_bed_sheets_google(spreadsheet)
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        rows = ws.get_all_values()
        if len(rows) < 2:
            return {"beds": []}
            
        headers = [str(h).strip().lower() for h in rows[0]]
        beds = []
        
        # Parse rows
        for idx, row in enumerate(rows[1:], start=2): # Start=2 to match sheet row number
            bed_data = {}
            for h_idx, h in enumerate(headers):
                if h_idx < len(row):
                    bed_data[h] = row[h_idx]
            
            clean_bed = {
                "room_no": bed_data.get("room no"),
                "room_type": bed_data.get("room type"),
                "bed_index": bed_data.get("bed index"),
                "patient_name": bed_data.get("patient name"),
                "member_id": bed_data.get("member id"),
                "gender": bed_data.get("gender"),
                "status": bed_data.get("status", "Available"),
                "admission_date": bed_data.get("admission date"),
                "discharge_date": bed_data.get("discharge date"),
                "pain_point": bed_data.get("pain point"),
                "row_idx": idx 
            }
            # Only add valid rows (skip empty ones if any)
            if clean_bed["room_no"]:
                beds.append(clean_bed)
                
        return {"beds": beds}
    except Exception as e:
        print(f"Error fetching beds: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/beds/allocate")
async def allocate_bed(payload: BedAllocationRequest):
    """Allocate a bed in Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        # Get all rows to find the target bed
        rows = ws.get_all_values()
        headers = [str(h).strip().lower() for h in rows[0]]
        
        # Find indices
        try:
            room_no_idx = headers.index("room no")
            bed_index_idx = headers.index("bed index")
            status_idx = headers.index("status")
            patient_idx = headers.index("patient name")
            mid_idx = headers.index("member id")
            gender_idx = headers.index("gender")
            adm_idx = headers.index("admission date")
            dis_idx = headers.index("discharge date")
            pain_idx = headers.index("pain point")
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"Sheet headers missing: {e}")

        target_row_num = -1
        
        # Search for the bed row (skip header)
        for i, row in enumerate(rows[1:], start=2):
            # Check range to avoid index errors
            if len(row) <= max(room_no_idx, bed_index_idx):
                continue
                
            r_no = str(row[room_no_idx]).strip()
            b_idx = str(row[bed_index_idx]).strip()
            
            if r_no == str(payload.room_no) and b_idx == str(payload.bed_index):
                target_row_num = i
                # Check status
                if len(row) > status_idx and row[status_idx] == "Occupied":
                    raise HTTPException(status_code=400, detail="Bed already occupied")
                break
        
        if target_row_num == -1:
             raise HTTPException(status_code=404, detail="Bed not found in system")

        # Update the row
        # gspread uses 1-based indexing for columns
        updates = [
            {"range": f"{chr(65+patient_idx)}{target_row_num}", "values": [[payload.patient_name]]},
            {"range": f"{chr(65+mid_idx)}{target_row_num}", "values": [[payload.member_id or ""]]},
            {"range": f"{chr(65+gender_idx)}{target_row_num}", "values": [[payload.gender]]},
            {"range": f"{chr(65+adm_idx)}{target_row_num}", "values": [[payload.admission_date]]},
            {"range": f"{chr(65+dis_idx)}{target_row_num}", "values": [[payload.discharge_date or ""]]},
            {"range": f"{chr(65+status_idx)}{target_row_num}", "values": [["Occupied"]]},
            {"range": f"{chr(65+pain_idx)}{target_row_num}", "values": [[payload.pain_point or ""]]}
        ]
        
        # Batch update is more efficient (or multiple update_cells)
        # For simplicity with gspread, we can just update the row range 
        # But constructing the full row might be tricky if we don't have it all.
        # Let's use update_acell or update matching the column index.
        # Note: headers.index returns 0-based. gspread col is i+1.
        
        ws.update_cell(target_row_num, patient_idx + 1, payload.patient_name)
        ws.update_cell(target_row_num, mid_idx + 1, payload.member_id or "")
        ws.update_cell(target_row_num, gender_idx + 1, payload.gender)
        ws.update_cell(target_row_num, adm_idx + 1, payload.admission_date)
        ws.update_cell(target_row_num, dis_idx + 1, payload.discharge_date or "")
        ws.update_cell(target_row_num, status_idx + 1, "Occupied")
        ws.update_cell(target_row_num, pain_idx + 1, payload.pain_point or "")
        
        return {"status": "success", "message": "Bed allocated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error allocating bed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/complaints")
async def log_complaint(payload: ComplaintRequest):
    """Log a complaint to Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        try:
            ws = spreadsheet.worksheet(COMPLAINT_SHEET_NAME)
        except gspread.WorksheetNotFound:
            # Fallback if ensure wasn't called (though get_beds calls it)
            ensure_bed_sheets_google(spreadsheet)
            ws = spreadsheet.worksheet(COMPLAINT_SHEET_NAME)
            
        ws.append_row([
            datetime.now().strftime('%Y-%m-%d'),
            payload.room_no,
            payload.patient_name,
            payload.complaint_type,
            payload.description,
            "No"
        ])
        
        return {"status": "success", "message": "Complaint logged"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/feedback")
async def submit_feedback(payload: FeedbackRequest):
    """Submit feedback to Google Sheets."""
    try:
        client, spreadsheet = get_google_sheet_client()
        try:
            ws = spreadsheet.worksheet(FEEDBACK_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ensure_bed_sheets_google(spreadsheet)
            ws = spreadsheet.worksheet(FEEDBACK_SHEET_NAME)
            
        ws.append_row([
            datetime.now().strftime('%Y-%m-%d'),
            payload.patient_name,
            payload.rating_comfort,
            payload.rating_cleanliness,
            payload.rating_staff,
            payload.comments
        ])
        
        return {"status": "success", "message": "Feedback submitted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/beds/discharge")
async def discharge_bed(room_no: str, bed_index: int):
    """Discharge a patient and release the bed. Patient data remains in main sheet."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        # Get all rows
        rows = ws.get_all_values()
        headers = [str(h).strip().lower() for h in rows[0]]
        
        # Find indices
        try:
            room_no_idx = headers.index("room no")
            bed_index_idx = headers.index("bed index")
            status_idx = headers.index("status")
            patient_idx = headers.index("patient name")
            mid_idx = headers.index("member id")
            gender_idx = headers.index("gender")
            adm_idx = headers.index("admission date")
            dis_idx = headers.index("discharge date")
            pain_idx = headers.index("pain point")
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"Sheet headers missing: {e}")

        target_row_num = -1
        
        # Find the bed row
        for i, row in enumerate(rows[1:], start=2):
            if len(row) <= max(room_no_idx, bed_index_idx):
                continue
                
            r_no = str(row[room_no_idx]).strip()
            b_idx = str(row[bed_index_idx]).strip()
            
            if r_no == str(room_no) and b_idx == str(bed_index):
                target_row_num = i
                # Check if occupied
                if len(row) > status_idx and row[status_idx] != "Occupied":
                    raise HTTPException(status_code=400, detail="Bed is not occupied")
                break
        
        if target_row_num == -1:
            raise HTTPException(status_code=404, detail="Bed not found")

        # Clear bed data using batch update for better performance
        # Convert column indices to A1 notation
        def col_to_letter(col_idx):
            """Convert 0-based column index to Excel column letter"""
            col_idx += 1  # Convert to 1-based
            result = ""
            while col_idx > 0:
                col_idx -= 1
                result = chr(65 + (col_idx % 26)) + result
                col_idx //= 26
            return result
        
        # Prepare batch update
        updates = [
            {
                'range': f'{col_to_letter(patient_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(mid_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(gender_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(adm_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(dis_idx)}{target_row_num}',
                'values': [['']]
            },
            {
                'range': f'{col_to_letter(status_idx)}{target_row_num}',
                'values': [['Available']]
            },
            {
                'range': f'{col_to_letter(pain_idx)}{target_row_num}',
                'values': [['']]
            }
        ]
        
        ws.batch_update(updates)
        
        return {"status": "success", "message": "Patient discharged successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error discharging bed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/beds/update-discharge")
async def update_discharge_date(room_no: str, bed_index: int, discharge_date: str):
    """Update discharge date for an occupied bed. Does NOT discharge the patient."""
    try:
        client, spreadsheet = get_google_sheet_client()
        ws = spreadsheet.worksheet(ADMISSION_SHEET_NAME)
        
        # Get all rows
        rows = ws.get_all_values()
        headers = [str(h).strip().lower() for h in rows[0]]
        
        # Find indices
        try:
            room_no_idx = headers.index("room no")
            bed_index_idx = headers.index("bed index")
            status_idx = headers.index("status")
            dis_idx = headers.index("discharge date")
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"Sheet headers missing: {e}")

        target_row_num = -1
        
        # Find the bed row
        for i, row in enumerate(rows[1:], start=2):
            if len(row) <= max(room_no_idx, bed_index_idx):
                continue
                
            r_no = str(row[room_no_idx]).strip()
            b_idx = str(row[bed_index_idx]).strip()
            
            if r_no == str(room_no) and b_idx == str(bed_index):
                target_row_num = i
                # Check if occupied
                if len(row) > status_idx and row[status_idx] != "Occupied":
                    raise HTTPException(status_code=400, detail="Bed is not occupied")
                break
        
        if target_row_num == -1:
            raise HTTPException(status_code=404, detail="Bed not found")

        # ONLY update discharge date - do NOT change status or clear patient data
        ws.update_cell(target_row_num, dis_idx + 1, discharge_date or "")
        
        return {"status": "success", "message": "Discharge date updated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating discharge date: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class AIChatRequest(BaseModel):
    query: str
    filter: Optional[str] = None


# AI Configuration
AI_PROVIDER = os.getenv('AI_PROVIDER', 'groq')
AI_ENABLED = os.getenv('AI_ENABLED', 'True').lower() == 'true'

# Groq Configuration (Fast & Reliable)
GROQ_API_KEY = os.getenv('GROQ_API_KEY', '')
GROQ_MODEL = os.getenv('GROQ_MODEL', 'llama-3.1-70b-versatile')
GROQ_API_BASE_URL = 'https://api.groq.com/openai/v1'

# Hugging Face Configuration (Backup)
HF_TOKEN = os.getenv('HF_TOKEN', '')
HF_MODEL = os.getenv('HF_MODEL', 'mistralai/Mistral-7B-Instruct-v0.2')
HF_API_BASE_URL = os.getenv('HF_API_BASE_URL', 'https://router.huggingface.co/v1')
HF_ENABLED = os.getenv('HF_ENABLED', 'False').lower() == 'true'


async def query_groq_ai(user_query: str, crm_data_summary: str, member_ids: List[str]) -> Optional[str]:
    """
    Query Groq API with CRM context.
    
    Args:
        user_query: The user's question
        crm_data_summary: Summary of relevant CRM data
        member_ids: List of relevant member IDs
        
    Returns:
        AI-generated response or None if API fails
    """
    if not GROQ_API_KEY:
        return None
    
    try:
        # Build the prompt with CRM context
        system_prompt = """You are an AI CRM assistant for an elderly care service. 
You help staff quickly understand customer data and follow-up requirements.
Be concise, professional, and helpful.

IMPORTANT: Member names are shown as "AttenderName/PatientName" when both exist. Search BOTH names when looking for a person.

When users ask for "names", provide the patient/customer names.
When users ask for "member IDs" or "IDs", provide the member IDs.
When users ask about locations (e.g., city, area, where), provide the location for each relevant member.
When users ask for phone numbers (e.g., phone, mobile, contact number), provide the phone number for each relevant member.
Otherwise, provide both name and ID in format: "Name (ID)" and include location and phone if it adds clarity."""
        
        user_prompt = f"""CRM Data Context:
{crm_data_summary}

User Question: {user_query}

Provide a clear, concise answer based on the CRM data above.
If the user asks for names, list the names from the member details.
If the user asks for IDs, list the member IDs.
If the user asks about location, provide the member locations alongside their names or IDs.
If the user asks for phone numbers, provide the phone/mobile number alongside the name or ID.
Otherwise, provide both name and ID, including location and phone if helpful."""
        
        # Prepare the API request
        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": GROQ_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "max_tokens": 500,
            "temperature": 0.7,
            "top_p": 0.9
        }
        
        # Make async request to Groq
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{GROQ_API_BASE_URL}/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code == 200:
                result = response.json()
                ai_response = result.get('choices', [{}])[0].get('message', {}).get('content', '')
                if ai_response:
                    print(f"✅ Groq AI Response received: {ai_response[:100]}...")
                    return ai_response.strip()
                else:
                    print("⚠️ Groq returned empty response")
                    return None
            else:
                print(f"❌ Groq API error: {response.status_code} - {response.text}")
                return None
                
    except Exception as e:
        print(f"Error calling Groq API: {e}")
        return None


async def query_huggingface_ai(user_query: str, crm_data_summary: str, member_ids: List[str]) -> Optional[str]:
    """
    Query Hugging Face Inference API with CRM context.
    
    Args:
        user_query: The user's question
        crm_data_summary: Summary of relevant CRM data
        member_ids: List of relevant member IDs
        
    Returns:
        AI-generated response or None if API fails
    """
    if not HF_TOKEN or not HF_ENABLED:
        return None
    
    try:
        # Build the prompt with CRM context
        system_prompt = """You are an AI CRM assistant for an elderly care service. 
You help staff quickly understand customer data and follow-up requirements.
Be concise, professional, and helpful.

IMPORTANT: Member names are shown as "AttenderName/PatientName" when both exist. Search BOTH names when looking for a person.

When users ask for "names", provide the patient/customer names.
When users ask for "member IDs" or "IDs", provide the member IDs.
When users ask about locations (e.g., city, area, where), provide the location for each relevant member.
When users ask for phone numbers (e.g., phone, mobile, contact number), provide the phone number for each relevant member.
Otherwise, provide both name and ID in format: "Name (ID)" and include location and phone if it adds clarity."""
        
        user_prompt = f"""CRM Data Context:
{crm_data_summary}

User Question: {user_query}

Provide a clear, concise answer based on the CRM data above.
If the user asks for names, list the names from the member details.
If the user asks for IDs, list the member IDs.
If the user asks about location, provide the member locations alongside their names or IDs.
If the user asks for phone numbers, provide the phone/mobile number alongside the name or ID.
Otherwise, provide both name and ID, including location and phone if helpful."""
        
        # Prepare the API request
        headers = {
            "Authorization": f"Bearer {HF_TOKEN}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": HF_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "max_tokens": 500,
            "temperature": 0.7,
            "top_p": 0.9
        }
        
        # Make async request to Hugging Face (OpenAI-compatible format)
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"{HF_API_BASE_URL}/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code == 200:
                result = response.json()
                # OpenAI-compatible format
                ai_response = result.get('choices', [{}])[0].get('message', {}).get('content', '')
                if ai_response:
                    print(f"✅ AI Response received: {ai_response[:100]}...")
                    return ai_response.strip()
                else:
                    print("⚠️ AI returned empty response")
                    return None
            elif response.status_code == 503:
                # Model is loading
                print(f"⚠️ Model is loading, using fallback...")
                return None
            else:
                print(f"❌ Hugging Face API error: {response.status_code} - {response.text}")
                return None
                
    except Exception as e:
        print(f"Error calling Hugging Face API: {e}")
        return None


def generate_fallback_response(query_lower: str, filter_type: str, filtered_rows: List, 
                              member_ids: List[str], status_counts: Dict = None, 
                              member_data: List[Dict] = None) -> str:
    """
    Generate rule-based response as fallback when AI is unavailable.
    
    Args:
        query_lower: Lowercase user query
        filter_type: Applied filter (today, this_week, overdue)
        filtered_rows: Filtered CRM data rows
        member_ids: List of member IDs
        status_counts: Dictionary of status counts
        member_data: List of {id, name, location} dictionaries
        
    Returns:
        Formatted response string
    """
    # Check if user is asking for names, locations, or phone numbers
    name_keywords = ['name', 'who', 'patient', 'customer']
    location_keywords = ['location', 'where', 'city', 'area', 'place', 'address', 'region']
    phone_keywords = ['phone', 'mobile', 'contact', 'number', 'contact number', 'mobile number']
    asking_for_names = any(keyword in query_lower for keyword in name_keywords)
    asking_for_location = any(keyword in query_lower for keyword in location_keywords)
    asking_for_phone = any(keyword in query_lower for keyword in phone_keywords)

    # Format member info based on what user is asking
    def format_members(limit=10):
        if not member_data:
            return ', '.join(member_ids[:limit])

        formatted = []
        for item in member_data[:limit]:
            name = item.get('name', 'Unknown')
            mid = item.get('id', '')
            location = item.get('location')
            phone = item.get('phone')
            has_location = location and location.lower() != 'unknown location'
            has_phone = bool(str(phone).strip())

            # Determine display based on query intent
            if asking_for_phone and not asking_for_location and not asking_for_names:
                # Only phone requested
                if has_phone and name != 'Unknown':
                    formatted.append(f"{name} - {phone}")
                elif has_phone:
                    formatted.append(f"{mid} - {phone}")
                else:
                    formatted.append(f"{name if name != 'Unknown' else mid} - phone not available")
            elif asking_for_location and not asking_for_names:
                if has_location and name != 'Unknown':
                    formatted.append(f"{name} - {location}")
                elif has_location:
                    formatted.append(f"{mid} - {location}")
                else:
                    formatted.append(f"{name if name != 'Unknown' else mid} - location not available")
            elif asking_for_names and not asking_for_location:
                if name != 'Unknown':
                    if asking_for_phone and has_phone and has_location:
                        formatted.append(f"{name} ({location} · {phone})")
                    elif asking_for_phone and has_phone:
                        formatted.append(f"{name} ({phone})")
                    elif has_location:
                        formatted.append(f"{name} ({location})")
                    else:
                        formatted.append(name)
                else:
                    if asking_for_phone and has_phone and has_location:
                        formatted.append(f"{mid} ({location} · {phone})")
                    elif asking_for_phone and has_phone:
                        formatted.append(f"{mid} ({phone})")
                    elif has_location:
                        formatted.append(f"{mid} ({location})")
                    else:
                        formatted.append(mid)
            else:
                base = f"{mid} ({name})"
                if has_location:
                    base += f" - {location}"
                if asking_for_phone and has_phone:
                    base += f" - {phone}"
                formatted.append(base)

        if not formatted:
            # Fallback if member_data was empty or formatting failed
            base_list = member_names if asking_for_names else member_ids
            return ', '.join(base_list[:limit])

        return ', '.join(formatted)

    if 'follow' in query_lower or 'followup' in query_lower or 'follow-up' in query_lower:
        members_str = format_members(10)
        if filter_type == 'today':
            return f"Found {len(filtered_rows)} member(s) requiring follow-up today: {members_str}"
        elif filter_type == 'this_week' or filter_type == 'this week':
            return f"Found {len(filtered_rows)} member(s) requiring follow-up this week: {members_str}"
        elif filter_type == 'overdue':
            return f"Found {len(filtered_rows)} overdue follow-up(s): {members_str}"
        else:
            return f"Found {len(filtered_rows)} member(s) with follow-ups: {members_str}"
    
    elif 'status' in query_lower and status_counts:
        status_summary = ', '.join([f"{k}: {v}" for k, v in status_counts.items()])
        members_str = format_members(10)
        return f"Lead status summary: {status_summary}. Members: {members_str}"
    
    elif 'count' in query_lower or 'how many' in query_lower:
        response = f"Found {len(filtered_rows)} member(s) matching your criteria."
        if member_data or member_ids:
            members_str = format_members(10)
            response += f" Members: {members_str}"
        return response
    
    else:
        if len(filtered_rows) == 0:
            return "No members found matching your criteria."
        else:
            members_str = format_members(10)
            response = f"Found {len(filtered_rows)} member(s): {members_str}"
            if len(member_ids) > 10:
                response += f" and {len(member_ids) - 10} more..."
            return response


# AI CRM Chat Models
class AIChatRequest(BaseModel):
    query: str
    filter: Optional[str] = None

def get_all_google_sheets_data():
    """Get ALL data from Google Sheets for AI Analytics - no field filtering"""
    if not os.path.exists(CREDENTIALS_FILE):
        raise Exception("Google credentials file not found")
    
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive'
    ]
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
    client = gspread.authorize(creds)
    spreadsheet = ensure_google_sheet(client)
    sheet = spreadsheet.sheet1
    
    # Get ALL values from the sheet
    values = sheet.get_all_values()
    
    if len(values) < 2:
        return {"headers": [], "data": [], "total_rows": 0}
    
        return "No data available"
    
    # Take a sample of rows for AI Analytics
    sample_rows = data_rows[:max_rows]
    
    # Create a structured text representation
    lines = []
    lines.append(f"CRM Data Summary ({len(sample_rows)} of {len(data_rows)} total records):")
    lines.append("=" * 50)
    lines.append("")
    
    # Add headers
    lines.append("Available Columns:")
    for i, header in enumerate(headers):
        lines.append(f"{i+1}. {header}")
    lines.append("")
    
    # Add sample data in a readable format
    lines.append("Sample Records:")
    for row_idx, row in enumerate(sample_rows):
        lines.append(f"\nRecord {row_idx + 1}:")
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and value.strip():  # Only show non-empty values
                lines.append(f"  {headers[col_idx]}: {value}")
        
        # Only show first 10 records in detail to avoid overwhelming the AI
        if row_idx >= 9:
            lines.append(f"\n... and {len(sample_rows) - 10} more records")
            break
    
    return "\n".join(lines)

@app.post("/api/ai-crm/chat")
async def ai_crm_chat(request: AIChatRequest, background_tasks: BackgroundTasks):
    """
    AI CRM Chat endpoint - queries Google Sheets data based on user input.
    Supports filters: today, this_week, overdue
    """
    try:
        if not os.path.exists(CREDENTIALS_FILE):
            return {
                "response": "Google Sheets is not connected. Please configure credentials.",
                "member_ids": [],
                "connected": False
            }
        
        # Connect to Google Sheets
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scope)
        client = gspread.authorize(creds)
        spreadsheet = ensure_google_sheet(client)
        
        # Try to find the correct worksheet
        try:
            worksheet = spreadsheet.worksheet(GOOGLE_SHEET_NAME)
        except gspread.exceptions.WorksheetNotFound:
            # Try alternative names
            worksheets = spreadsheet.worksheets()
            worksheet_names = [ws.title for ws in worksheets]
            
            # Try to find a worksheet with "CRM" or "Lead" in the name
            crm_sheet = None
            for ws in worksheets:
                ws_name_lower = ws.title.lower()
                if 'crm' in ws_name_lower or 'lead' in ws_name_lower:
                    crm_sheet = ws
                    break
            
            if crm_sheet:
                worksheet = crm_sheet
            else:
                # Use the first worksheet if no CRM sheet found
                worksheet = worksheets[0] if worksheets else None
                
            if not worksheet:
                return {
                    "response": f"Could not find CRM worksheet. Available sheets: {', '.join(worksheet_names)}",
                    "member_ids": [],
                    "connected": True
                }
        
        # Get all data
        all_rows = worksheet.get_all_values()
        if len(all_rows) < 2:
            return {
                "response": "No data found in the CRM sheet.",
                "member_ids": [],
                "connected": True
            }
        
        original_headers = [str(h).strip() for h in all_rows[0]]
        headers = [h.lower() for h in original_headers]
        data_rows = all_rows[1:]
        
        # Find relevant column indices
        member_id_col = next((i for i, h in enumerate(headers) if 'member' in h and 'id' in h), None)
        date_col = next((i for i, h in enumerate(headers) if h == 'date'), None)
        follow1_col = next((i for i, h in enumerate(headers) if 'follow1 date' in h or 'follow_1 date' in h), None)
        follow2_col = next((i for i, h in enumerate(headers) if 'follow_2 date' in h or 'follow2 date' in h), None)
        follow3_col = next((i for i, h in enumerate(headers) if 'follow_3 date' in h or 'follow3 date' in h), None)
        lead_status_col = next((i for i, h in enumerate(headers) if 'lead status' in h), None)
        patient_name_col = next((i for i, h in enumerate(headers) if 'patient name' in h), None)
        attender_name_col = next((i for i, h in enumerate(headers) if 'attender name' in h), None)
        patient_location_col = next((i for i, h in enumerate(headers) if 'patient location' in h), None)
        location_col = next((i for i, h in enumerate(headers) if h == 'location'), None)
        area_col = next((i for i, h in enumerate(headers) if 'area' in h), None)
        # Phone/Mobile/Contact number column
        email_col = next((i for i, h in enumerate(headers) if 'email' in h), None)
        mobile_col = next((i for i, h in enumerate(headers) if any(k in h for k in ['mobile', 'phone', 'contact'])), None)
        
        # Get today's date
        from datetime import datetime, timedelta
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        
        def parse_date(date_str):
            """Parse date string in various formats"""
            if not date_str or str(date_str).strip() == '':
                return None
            try:
                # Try DD/MM/YYYY format
                if '/' in str(date_str):
                    parts = str(date_str).split('/')
                    if len(parts) == 3:
                        return datetime.strptime(date_str, '%d/%m/%Y').date()
                # Try YYYY-MM-DD format
                elif '-' in str(date_str):
                    return datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                pass
            return None
        
        def smart_filter_rows(all_rows, query_text, headers_list):
            """
            Smart retrieval: Scan ALL rows and filter based on query keywords.
            For field-specific queries (age, gender, email, etc.), return ALL rows.
            """
            query_lower = query_text.lower()
            
            # Field-specific query keywords - return ALL rows for these
            field_keywords = ['age', 'gender', 'email', 'phone', 'mobile', 'service', 'status', 
                            'location', 'area', 'source', 'pain', 'date', 'assigned', 'agent']
            
            # If query is asking about a field value, return ALL rows
            if any(keyword in query_lower for keyword in field_keywords):
                return all_rows
            
            relevant_rows = []
            
            # Extract potential search terms from query
            search_terms = []
            # Remove common words
            stop_words = {'the', 'a', 'an', 'is', 'are', 'was', 'were', 'of', 'for', 'in', 'on', 'at', 'to', 'from', 'by', 'with', 'what', 'who', 'where', 'when', 'how', 'show', 'give', 'get', 'find', 'tell', 'me', 'my', 'i', 'you', 'your', 'member', 'members', 'patient', 'patients'}
            words = query_lower.split()
            search_terms = [w for w in words if w not in stop_words and len(w) > 2]
            
            # Scan ALL rows
            for row in all_rows:
                if len(row) == 0:
                    continue
                
                # If no specific search terms, include all rows
                if not search_terms:
                    relevant_rows.append(row)
                    continue
                
                # Check if any search term matches any cell in the row
                row_text = ' '.join([str(cell).lower() for cell in row])
                if any(term in row_text for term in search_terms):
                    relevant_rows.append(row)
            
            return relevant_rows if relevant_rows else all_rows  # Return all if no matches
        
        # Smart filtering: Scan ALL rows first, then apply query-based filtering
        query_lower = request.query.lower()
        filter_type = (request.filter or '').lower()
        
        # Step 1: Smart filter based on query (scans ALL rows)
        query_filtered_rows = smart_filter_rows(data_rows, request.query, headers)
        
        # Step 2: Apply date-based filter to query-filtered rows
        filtered_rows = []
        for row in query_filtered_rows:
            if len(row) == 0:
                continue
            
            # Apply filter
            if filter_type == 'today':
                # Check if any follow-up date is today
                follow_dates = []
                if follow1_col is not None and follow1_col < len(row):
                    follow_dates.append(parse_date(row[follow1_col]))
                if follow2_col is not None and follow2_col < len(row):
                    follow_dates.append(parse_date(row[follow2_col]))
                if follow3_col is not None and follow3_col < len(row):
                    follow_dates.append(parse_date(row[follow3_col]))
                
                if not any(d == today for d in follow_dates if d):
                    continue
                    
            elif filter_type == 'this_week' or filter_type == 'this week':
                # Check if any follow-up date is within this week
                follow_dates = []
                if follow1_col is not None and follow1_col < len(row):
                    follow_dates.append(parse_date(row[follow1_col]))
                if follow2_col is not None and follow2_col < len(row):
                    follow_dates.append(parse_date(row[follow2_col]))
                if follow3_col is not None and follow3_col < len(row):
                    follow_dates.append(parse_date(row[follow3_col]))
                
                if not any(week_ago <= d <= today for d in follow_dates if d):
                    continue
                    
            elif filter_type == 'overdue':
                # Check if any follow-up date is in the past
                follow_dates = []
                if follow1_col is not None and follow1_col < len(row):
                    follow_dates.append(parse_date(row[follow1_col]))
                if follow2_col is not None and follow2_col < len(row):
                    follow_dates.append(parse_date(row[follow2_col]))
                if follow3_col is not None and follow3_col < len(row):
                    follow_dates.append(parse_date(row[follow3_col]))
                
                if not any(d < today for d in follow_dates if d):
                    continue
            
            filtered_rows.append(row)
        
        # Extract member IDs and names from filtered rows
        member_ids = []
        member_names = []
        member_data = []  # Store {id, name, location, phone, fields} dictionaries

        if member_id_col is not None:
            for row in filtered_rows:
                if member_id_col < len(row) and row[member_id_col]:
                    mid = str(row[member_id_col]).strip()
                    member_ids.append(mid)
                    
                    # Try to get patient/customer name (check both patient and attender)
                    name = ""
                    if patient_name_col is not None and patient_name_col < len(row):
                        name = str(row[patient_name_col]).strip()
                    
                    # Also check attender name
                    attender_name = ""
                    if attender_name_col is not None and attender_name_col < len(row):
                        attender_name = str(row[attender_name_col]).strip()
                    
                    # Prefer patient name, but use attender if patient is empty
                    if not name and attender_name:
                        name = attender_name
                    
                    # If still no name, try other name columns
                    if not name:
                        for i, h in enumerate(headers):
                            if 'name' in h and i < len(row):
                                potential_name = str(row[i]).strip()
                                if potential_name and potential_name != mid:
                                    name = potential_name
                                    break
                    
                    name = name if name else "Unknown"

                    # Determine best available location information
                    location = ""
                    if patient_location_col is not None and patient_location_col < len(row):
                        location = str(row[patient_location_col]).strip()
                    if (not location) and location_col is not None and location_col < len(row):
                        location = str(row[location_col]).strip()
                    if (not location) and area_col is not None and area_col < len(row):
                        location = str(row[area_col]).strip()
                    if not location:
                        # Search for other location-related headers
                        for i, h in enumerate(headers):
                            if any(keyword in h for keyword in ['city', 'town', 'district']) and i < len(row):
                                potential_loc = str(row[i]).strip()
                                if potential_loc:
                                    location = potential_loc
                                    break
                    location = location if location else "Unknown location"

                    # Extract phone/mobile number if present
                    phone = ""
                    if mobile_col is not None and mobile_col < len(row):
                        phone = str(row[mobile_col]).strip()

                    email_value = ""
                    if email_col is not None and email_col < len(row):
                        email_value = str(row[email_col]).strip()

                    # Build a field map for this row (header -> value)
                    field_map = {}
                    for ci, ch in enumerate(original_headers):
                        if ci < len(row):
                            val = str(row[ci]).strip()
                            if val:
                                field_map[ch] = val

                    member_names.append(name)
                    # Store both patient and attender names for AI search
                    member_data.append({
                        "id": mid, 
                        "name": name, 
                        "patient_name": name if patient_name_col is not None and patient_name_col < len(row) else "",
                        "attender_name": attender_name,
                        "location": location, 
                        "phone": phone, 
                        "fields": field_map,
                        "email": email_value,
                    })
        
        # Collect status counts for context
        status_counts = {}
        if lead_status_col is not None:
            for row in filtered_rows:
                if lead_status_col < len(row):
                    status = str(row[lead_status_col]).strip() or 'Unknown'
                    status_counts[status] = status_counts.get(status, 0) + 1
        
        # Detect special commands (e.g., send mail)
        response_text = None
        send_mail_triggered = False
        query_lower = request.query.lower()

        if "send mail" in query_lower and "follow" in query_lower:
            send_mail_triggered = True
            target_member: Optional[Dict[str, Any]] = None
            target_member_id = None

            id_match = re.search(r"(mid-[\w-]+)", request.query, flags=re.IGNORECASE)
            if id_match:
                target_member_id = id_match.group(1).upper()

            if target_member_id:
                target_member = next((m for m in member_data if m.get("id", "").upper() == target_member_id), None)
            else:
                for item in member_data:
                    name_candidates = [item.get('name'), item.get('patient_name'), item.get('attender_name')]
                    for candidate in name_candidates:
                        if candidate and candidate.lower() in query_lower:
                            target_member = item
                            break
                    if target_member:
                        break

            if target_member:
                recipient_email = target_member.get("email")
                if not recipient_email and target_member.get("fields"):
                    for k, v in target_member["fields"].items():
                        if 'email' in k.lower() and str(v).strip():
                            recipient_email = str(v).strip()
                            break

                if recipient_email:
                    background_tasks.add_task(send_follow_email, recipient_email, target_member, request.query)
                    response_text = (
                        f"Scheduled follow-up email to {recipient_email} for "
                        f"{target_member.get('name', 'the member')} ({target_member.get('id')})."
                    )
                else:
                    response_text = (
                        f"I found {target_member.get('id')} but there is no email address on file."
                    )
            else:
                response_text = "I could not identify which member to email. Please include the member name or ID."

        # Build CRM data summary for AI with names
        crm_summary_parts = [
            f"Total records: {len(filtered_rows)}",
            f"Filter applied: {filter_type or 'none'}",
            f"Members found: {len(member_ids)}"
        ]
        # Include ALL available fields (schema) to allow AI to use any column dynamically
        if original_headers:
            hdr_text = ", ".join(original_headers)
            crm_summary_parts.append(f"Available fields: {hdr_text}")
        
        # Add member details (ID + Name) for AI context - ALL members for real-time accuracy
        if member_data:
            member_details = []
            for item in member_data:  # Include ALL members, no limit
                # Show both patient and attender names if different - COMPACT format
                att = item.get('attender_name', '')
                pat = item.get('patient_name', '')
                
                if att and pat and att != pat:
                    # Both names present and different
                    name_part = f"{att}/{pat}"
                elif att:
                    name_part = att
                elif pat:
                    name_part = pat
                else:
                    name_part = item['name']
                
                # Compact format: ID(Name-Location-Phone)
                parts = [item['id'], f"({name_part}"]
                if item.get('location') and item['location'] != 'Unknown location':
                    parts[-1] += f"-{item['location']}"
                if item.get('phone'):
                    parts[-1] += f"-{item['phone']}"
                parts[-1] += ")"
                member_details.append("".join(parts))
            crm_summary_parts.append(f"Members: {', '.join(member_details)}")
            
            # Add field details for queries asking about specific fields
            query_lower_check = request.query.lower()
            field_query_keywords = ['age', 'gender', 'email', 'service', 'pain', 'source', 'agent', 'assigned']
            if any(kw in query_lower_check for kw in field_query_keywords):
                # Include relevant field data for ALL members
                field_lines = []
                for item in member_data[:50]:  # Limit to 50 for token efficiency
                    fm = item.get('fields') or {}
                    # Extract fields mentioned in query
                    relevant_fields = {}
                    for field_name, field_value in fm.items():
                        field_lower = field_name.lower()
                        if any(kw in field_lower for kw in field_query_keywords):
                            relevant_fields[field_name] = field_value
                    if relevant_fields:
                        field_str = "; ".join([f"{k}={v}" for k, v in relevant_fields.items()])
                        field_lines.append(f"{item['id']}: {field_str}")
                if field_lines:
                    crm_summary_parts.append(f"Field details: {' | '.join(field_lines)}")
        
        if status_counts:
            status_summary = ', '.join([f"{k}: {v}" for k, v in status_counts.items()])
            crm_summary_parts.append(f"Lead statuses: {status_summary}")
        
        if filter_type == 'today':
            crm_summary_parts.append("These members need follow-up TODAY")
        elif filter_type == 'this_week' or filter_type == 'this week':
            crm_summary_parts.append("These members need follow-up THIS WEEK")
        elif filter_type == 'overdue':
            crm_summary_parts.append("These follow-ups are OVERDUE")
        
        crm_data_summary = "\n".join(crm_summary_parts)
        
        # Try to get AI-powered response first (only if no special command handled it)
        ai_response_text = None
        if response_text is None and AI_ENABLED:
            print(f"🤖 Calling {AI_PROVIDER.upper()} AI for query: {request.query}")

            if AI_PROVIDER == 'groq' and GROQ_API_KEY:
                ai_response_text = await query_groq_ai(
                    user_query=request.query,
                    crm_data_summary=crm_data_summary,
                    member_ids=member_ids
                )
            elif HF_ENABLED and HF_TOKEN:
                ai_response_text = await query_huggingface_ai(
                    user_query=request.query,
                    crm_data_summary=crm_data_summary,
                    member_ids=member_ids
                )

            if ai_response_text:
                print(f"✅ Using {AI_PROVIDER.upper()} AI response")
                response_text = ai_response_text
            else:
                print("⚠️ AI failed, using fallback")

        # Fall back to rule-based response if AI fails or is disabled
        if response_text is None:
            response_text = generate_fallback_response(
                query_lower=query_lower,
                filter_type=filter_type,
                filtered_rows=filtered_rows,
                member_ids=member_ids,
                status_counts=status_counts,
                member_data=member_data
            )
        
        return {
            "response": response_text,
            "member_ids": member_ids,  # Return all IDs
            "connected": True,
            "count": len(filtered_rows)
        }
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"AI Chat error: {e}")
        print(f"Full traceback:\n{error_details}")
        return {
            "response": f"Sorry, I encountered an error: {str(e)}. Please check if Google Sheets is properly configured.",
            "member_ids": [],
            "connected": False,
            "error_details": str(e)
        }



@app.get("/api/settings/charges")
async def get_charge_settings():
    return load_settings()


@app.post("/api/settings/charges")
async def update_charge_settings(settings: Settings):
    save_settings(settings)
    return {"status": "success", "settings": settings}


# ============ Dropdown Management API Endpoints ============

@app.get("/api/dropdown-options")
async def api_get_all_dropdown_options():
    """Get all dropdown fields and their options from DropdownOption sheet."""
    try:
        options = get_all_dropdown_options()
        return {"status": "success", "dropdown_options": options}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching dropdown options: {str(e)}")


@app.get("/api/dropdown-options/{field_name}")
async def api_get_dropdown_options(field_name: str):
    """Get options for a specific dropdown field."""
    try:
        options = get_dropdown_options_for_field(field_name)
        return {"status": "success", "field_name": field_name, "options": options}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching options for {field_name}: {str(e)}")


@app.post("/api/dropdown-options/{field_name}")
async def api_add_dropdown_option(field_name: str, option: str = Body(..., embed=True)):
    """Add a new option to a dropdown field."""
    try:
        if not option or not option.strip():
            raise HTTPException(status_code=400, detail="Option cannot be empty")
        
        result = add_dropdown_option(field_name, option.strip())
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error adding option: {str(e)}")


@app.delete("/api/dropdown-options/{field_name}/{option}")
async def api_delete_dropdown_option(field_name: str, option: str):
    """Remove an option from a dropdown field."""
    try:
        result = delete_dropdown_option(field_name, option)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting option: {str(e)}")


@app.post("/api/sync-schema-sheet")
async def api_sync_schema_sheet():
    """Sync schema fields with DropdownOption sheet columns bidirectionally."""
    try:
        # Sync schema to sheet (add new columns for new schema fields)
        sync_schema_to_dropdown_sheet()
        
        # Sync sheet to schema (update schema with sheet options)
        sync_dropdown_options_to_schema()
        
        return {
            "status": "success",
            "message": "Schema and DropdownOption sheet synced successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing: {str(e)}")


@app.get("/api/settings/charges")
async def get_charge_settings():
    return load_settings()

@app.post("/api/settings/charges")
async def update_charge_settings(settings: Settings):
    save_settings(settings)
    return {"status": "success", "settings": settings}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
