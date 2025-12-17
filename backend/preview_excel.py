"""
Preview Excel data in browser
"""
import openpyxl
from fastapi import FastAPI
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

app = FastAPI()

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "CRM_Lead_Template (1).xlsm"
SHEET_NAME = "Sheet"

@app.get("/", response_class=HTMLResponse)
async def preview_excel():
    """Display Excel data in an HTML table"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True, data_only=True)
        ws = wb[SHEET_NAME]
        
        # Get headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        # Get sample data rows (first 20 rows)
        data_rows = []
        for row_idx in range(2, min(22, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            data_rows.append(row_data)
        
        wb.close()
        
        # Build HTML
        html = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Preview - CRM Lead Template</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
        }
        
        .header {
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            margin-bottom: 20px;
        }
        
        h1 {
            color: #667eea;
            font-size: 2.5em;
            margin-bottom: 10px;
        }
        
        .info {
            color: #666;
            font-size: 1.1em;
        }
        
        .info strong {
            color: #764ba2;
        }
        
        .table-container {
            background: white;
            padding: 20px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow-x: auto;
        }
        
        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }
        
        thead {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            position: sticky;
            top: 0;
            z-index: 10;
        }
        
        th {
            padding: 15px 12px;
            text-align: left;
            font-weight: 600;
            border: 1px solid rgba(255,255,255,0.3);
            white-space: nowrap;
        }
        
        td {
            padding: 12px;
            border: 1px solid #e0e0e0;
            color: #333;
        }
        
        tbody tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        
        tbody tr:hover {
            background-color: #e3f2fd;
            transition: background-color 0.2s;
        }
        
        .empty-cell {
            color: #ccc;
            font-style: italic;
        }
        
        .stats {
            display: flex;
            gap: 20px;
            margin-top: 15px;
            flex-wrap: wrap;
        }
        
        .stat-box {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px 25px;
            border-radius: 10px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
        }
        
        .stat-label {
            font-size: 0.9em;
            opacity: 0.9;
        }
        
        .stat-value {
            font-size: 1.8em;
            font-weight: bold;
            margin-top: 5px;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 CRM Lead Template - Excel Preview</h1>
            <p class="info">
                <strong>File:</strong> CRM_Lead_Template (1).xlsm &nbsp;|&nbsp; 
                <strong>Sheet:</strong> Sheet
            </p>
            <div class="stats">
                <div class="stat-box">
                    <div class="stat-label">Total Fields</div>
                    <div class="stat-value">""" + str(len(headers)) + """</div>
                </div>
                <div class="stat-box">
                    <div class="stat-label">Sample Rows</div>
                    <div class="stat-value">""" + str(len(data_rows)) + """</div>
                </div>
            </div>
        </div>
        
        <div class="table-container">
            <table>
                <thead>
                    <tr>
"""
        
        # Add headers
        for header in headers:
            html += f"                        <th>{header if header else '(empty)'}</th>\n"
        
        html += """                    </tr>
                </thead>
                <tbody>
"""
        
        # Add data rows
        if data_rows:
            for row in data_rows:
                html += "                    <tr>\n"
                for cell in row:
                    cell_class = ' class="empty-cell"' if not cell else ''
                    cell_value = cell if cell else "(empty)"
                    html += f"                        <td{cell_class}>{cell_value}</td>\n"
                html += "                    </tr>\n"
        else:
            html += f'                    <tr><td colspan="{len(headers)}" style="text-align:center; padding:40px; color:#999;">No data rows found. This is a template file.</td></tr>\n'
        
        html += """                </tbody>
            </table>
        </div>
    </div>
</body>
</html>
"""
        return html
        
    except Exception as e:
        error_html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Error</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            padding: 50px;
            background: #f5f5f5;
        }}
        .error {{
            background: white;
            padding: 30px;
            border-radius: 10px;
            border-left: 5px solid #dc3545;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{ color: #dc3545; }}
    </style>
</head>
<body>
    <div class="error">
        <h1>❌ Error Loading Excel File</h1>
        <p><strong>Error:</strong> {str(e)}</p>
    </div>
</body>
</html>
"""
        return error_html

@app.get("/api/fields")
async def get_fields():
    """Get field list as JSON"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True, data_only=True)
        ws = wb[SHEET_NAME]
        
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        wb.close()
        
        return {
            "success": True,
            "file": EXCEL_FILE,
            "sheet": SHEET_NAME,
            "fields": headers,
            "count": len(headers)
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

if __name__ == "__main__":
    print("=" * 60)
    print("🚀 Starting Excel Preview Server")
    print("=" * 60)
    print(f"\n📁 File: {EXCEL_FILE}")
    print(f"📋 Sheet: {SHEET_NAME}")
    print("\n🌐 Open in browser: http://localhost:8080")
    print("\n💡 API endpoint: http://localhost:8080/api/fields")
    print("\nPress CTRL+C to stop the server\n")
    print("=" * 60)
    
    uvicorn.run(app, host="0.0.0.0", port=8080)
