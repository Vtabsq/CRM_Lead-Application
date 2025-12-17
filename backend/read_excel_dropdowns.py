import openpyxl
from openpyxl.utils import range_boundaries

EXCEL_FILE = "Lead CRM ApplicationData.xlsx"

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)

# Find the sheet - try "Lead CRM ApplicationData" or use active/first sheet
print("Available sheets:", wb.sheetnames)
if "Lead CRM ApplicationData" in wb.sheetnames:
    SHEET_NAME = "Lead CRM ApplicationData"
elif "Lead CRM" in wb.sheetnames:
    SHEET_NAME = "Lead CRM"
else:
    SHEET_NAME = wb.sheetnames[0]

print(f"Using sheet: {SHEET_NAME}\n")
ws = wb[SHEET_NAME]

# Get headers
print("=" * 80)
print("HEADERS")
print("=" * 80)
headers = []
for idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        header = str(cell.value).strip()
        headers.append(header)
        print(f"{idx}. {header}")

print(f"\nTotal headers: {len(headers)}")

# Get dropdowns/data validations
print("\n" + "=" * 80)
print("DROPDOWNS (Data Validations)")
print("=" * 80)

header_map = {idx: str(cell.value).strip() for idx, cell in enumerate(ws[1], start=1) if cell.value}

dv_list = getattr(ws, 'data_validations', None)
if dv_list and dv_list.dataValidation:
    dropdown_count = 0
    for dv in dv_list.dataValidation:
        if dv.type == "list":
            formula = dv.formula1 or ""
            
            # Parse options
            options = []
            if formula.startswith('='):
                # Reference to a range
                ref = formula[1:]
                if '!' in ref:
                    sheet_name, cell_range = ref.split('!', 1)
                    sheet_name = sheet_name.replace("'", "")
                    try:
                        ref_sheet = wb[sheet_name]
                    except KeyError:
                        continue
                else:
                    ref_sheet = ws
                    cell_range = ref
                
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            val = ref_sheet.cell(row=r, column=c).value
                            if val is not None and str(val).strip():
                                options.append(str(val).strip())
                except Exception:
                    pass
            else:
                # Direct list
                raw = formula.strip().strip('"')
                options = [x.strip() for x in raw.split(',') if x.strip()]
            
            if options:
                # Find which column this applies to
                for rng in str(dv.sqref).split():
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(rng)
                        if min_col == max_col:
                            col_idx = min_col
                            if col_idx in header_map:
                                dropdown_count += 1
                                print(f"\n{dropdown_count}. Field: {header_map[col_idx]}")
                                print(f"   Options ({len(options)}): {', '.join(options[:10])}")
                                if len(options) > 10:
                                    print(f"   ... and {len(options) - 10} more")
                    except Exception:
                        pass
    
    if dropdown_count == 0:
        print("No dropdowns found")
else:
    print("No data validations found")

wb.close()

print("\n" + "=" * 80)
