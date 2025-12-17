import openpyxl

print("=== CRM_Lead_Template (1).xlsm ===")
wb1 = openpyxl.load_workbook("CRM_Lead_Template (1).xlsm", keep_vba=True, data_only=True)
print("Available sheets:")
for sheet_name in wb1.sheetnames:
    print(f"  - {sheet_name}")
    
# Get first sheet
ws1 = wb1.active
print(f"\nFirst sheet: {ws1.title}")
print(f"Max row: {ws1.max_row}, Max col: {ws1.max_column}")
print("\nHeaders (first 10):")
for idx, cell in enumerate(list(ws1[1])[:10], start=1):
    if cell.value:
        print(f"  {idx}. {cell.value}")
wb1.close()

print("\n\n=== Lead CRM ApplicationData.xlsx ===")
wb2 = openpyxl.load_workbook("Lead CRM ApplicationData.xlsx", data_only=True)
print("Available sheets:")
for sheet_name in wb2.sheetnames:
    print(f"  - {sheet_name}")
    
ws2 = wb2.active
print(f"\nFirst sheet: {ws2.title}")
print(f"Max row: {ws2.max_row}, Max col: {ws2.max_column}")
print("\nHeaders (first 10):")
for idx, cell in enumerate(list(ws2[1])[:10], start=1):
    if cell.value:
        print(f"  {idx}. {cell.value}")
wb2.close()
