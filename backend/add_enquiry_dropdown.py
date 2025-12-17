"""Utility script to ensure the 'Enquiry made for' column exists in the main sheet
and that the corresponding dropdown options are populated in the List box sheet."""
from openpyxl import load_workbook

EXCEL_FILE = "Lead CRM ApplicationData.xlsx"
MAIN_SHEET = "Lead CRM"
LIST_BOX_SHEET = "List box"
MAIN_HEADER = "Enquiry made for"
LIST_BOX_HEADER = "Enquiry Made For"

DROPDOWN_OPTIONS = [
    "Job Enquiry",
    "Service Availability",
    "Service Types Offered",
    "Cost & Packages",
    "Appointment Booking",
    "Eligibility & Requirements",
    "Trial Services",
    "Caregiver Profiles",
    "Emergency Services",
    "Language Support",
    "Duration & Frequency",
    "Customization Options",
    "Transportation Assistance",
    "Insurance & Reimbursements",
    "Feedback & Escalations",
    "Cancellation & Refund Policies",
]

def ensure_column_and_dropdown():
    wb = load_workbook(EXCEL_FILE, data_only=False)
    try:
        # Ensure main sheet column exists
        ws_main = wb[MAIN_SHEET]
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws_main[1]]
        header_index_map = {header.lower(): idx for idx, header in enumerate(headers, start=1)}

        column_index = header_index_map.get(MAIN_HEADER.lower())
        if not column_index:
            column_index = ws_main.max_column + 1
            ws_main.cell(row=1, column=column_index, value=MAIN_HEADER)
        else:
            ws_main.cell(row=1, column=column_index, value=MAIN_HEADER)

        # Make sure each data row has at least an empty string for the new column
        for row in range(2, ws_main.max_row + 1):
            cell = ws_main.cell(row=row, column=column_index)
            if cell.value is None:
                cell.value = ""

        # Ensure List box sheet has the dropdown options
        ws_list = wb[LIST_BOX_SHEET]
        list_headers = [str(cell.value).strip() if cell.value else "" for cell in ws_list[1]]
        list_header_index_map = {header.lower(): idx for idx, header in enumerate(list_headers, start=1)}
        list_col_index = list_header_index_map.get(LIST_BOX_HEADER.lower())
        if not list_col_index:
            list_col_index = ws_list.max_column + 1
            ws_list.cell(row=1, column=list_col_index, value=LIST_BOX_HEADER)
        else:
            ws_list.cell(row=1, column=list_col_index, value=LIST_BOX_HEADER)

        # Clear existing values in the column
        max_rows = max(ws_list.max_row, len(DROPDOWN_OPTIONS) + 1)
        for row in range(2, max_rows + 1):
            ws_list.cell(row=row, column=list_col_index, value=None)

        # Populate dropdown options (one per row, starting at row 2)
        for offset, option in enumerate(DROPDOWN_OPTIONS, start=2):
            ws_list.cell(row=offset, column=list_col_index, value=option)

        wb.save(EXCEL_FILE)
        print("✅ 'Enquiry made for' column and dropdown options ensured.")
    finally:
        wb.close()


if __name__ == "__main__":
    ensure_column_and_dropdown()
