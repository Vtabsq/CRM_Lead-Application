import openpyxl

EXCEL_FILE = "Lead CRM ApplicationData.xlsx"

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Read List box sheet
if "List box" in wb.sheetnames:
    ws = wb["List box"]
    print("=" * 80)
    print("LIST BOX SHEET - Dropdown Options")
    print("=" * 80)
    
    # Get headers from List box sheet
    headers_row = []
    for cell in ws[1]:
        if cell.value:
            headers_row.append(str(cell.value).strip())
        else:
            headers_row.append("")
    
    print(f"\nColumns in List box sheet ({len(headers_row)}):")
    for idx, header in enumerate(headers_row, start=1):
        if header:
            print(f"{idx}. {header}")
    
    # Read values under each column (first 50 rows)
    print("\n" + "=" * 80)
    print("DROPDOWN OPTIONS BY COLUMN")
    print("=" * 80)
    
    for col_idx, header in enumerate(headers_row, start=1):
        if not header:
            continue
        
        values = set()
        for row_idx in range(2, min(ws.max_row + 1, 1000)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip():
                values.add(str(val).strip())
        
        if values:
            sorted_values = sorted(values)
            print(f"\n{header} ({len(sorted_values)} options):")
            print(f"   {', '.join(sorted_values[:15])}")
            if len(sorted_values) > 15:
                print(f"   ... and {len(sorted_values) - 15} more")

wb.close()
